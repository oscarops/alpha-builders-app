# ==============================================================================
# PARTE 1 DE 5: CONFIGURACIÓN DE PÁGINA, ESTILOS CSS Y CONSTANTES
# ==============================================================================
import base64
import datetime
import io
import json
import os
import zoneinfo
import openpyxl
import pandas as pd
from PIL import Image, ImageOps
import reportlab.lib.colors as colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from supabase import Client, create_client


def safe_json_dumps(obj):
    """Serializador JSON seguro para datos provenientes de Pandas/NumPy/fechas."""
    return json.dumps(obj, ensure_ascii=False, default=str, sort_keys=True)


# ==============================================================================
# 1. CONFIGURACIÓN DE PÁGINA Y ESTILOS GLOBALES (SOPORTE MODO CLARO Y OSCURO)
# ==============================================================================
st.set_page_config(
    page_title="Alpha Builders | Portal Ejecutivo",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Montserrat:wght@500;600;700;800&display=swap');

    :root {
        --bg-app: #f8fafc;
        --text-app: #0f172a;
        --card-bg: #ffffff;
        --card-border: #cbd5e1;
        --header-bg: #0f172a;
        --header-text: #ffffff;
        --sidebar-bg: #0b0f19;
        --sidebar-card: #111827;
        --sidebar-border: #1f2937;
        --subtext: #64748b;
        --hover-bg: #f1f5f9;
        --table-row-even: #f8fafc;
    }

    @media (prefers-color-scheme: dark) {
        :root {
            --bg-app: #0f172a;
            --text-app: #f8fafc;
            --card-bg: #1e293b;
            --card-border: #334155;
            --header-bg: #020617;
            --header-text: #ffffff;
            --sidebar-bg: #030712;
            --sidebar-card: #0f172a;
            --sidebar-border: #1e293b;
            --subtext: #94a3b8;
            --hover-bg: #334155;
            --table-row-even: #1e293b;
        }
    }

    html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }
    h1, h2, h3, .brand-title { font-family: 'Montserrat', sans-serif !important; letter-spacing: -0.03em !important; }

    .block-container { 
        padding-top: 3.2rem !important; 
        padding-bottom: 1.2rem !important; 
        padding-left: 1rem !important; 
        padding-right: 1rem !important; 
        max-width: 100% !important; 
    }
    .stApp { background-color: var(--bg-app) !important; color: var(--text-app) !important; }
    .stApp p, .stApp label, .stApp span, .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 { color: var(--text-app); }
    .stCaption, caption, small, [data-testid="stCaptionContainer"] { color: var(--subtext) !important; }

    [data-testid="stInputInstructions"], div[data-testid="stInputInstructions"] { display: none !important; visibility: hidden !important; }
    [data-testid="stHeader"] { background: transparent !important; z-index: 100 !important; }

    /* BOTONES DE TOGGLE SIDEBAR */
    [data-testid="stSidebarCollapseButton"] { display: block !important; visibility: visible !important; opacity: 1 !important; z-index: 999999 !important; }
    [data-testid="collapsedControl"] { display: block !important; visibility: visible !important; opacity: 1 !important; position: fixed !important; top: 12px !important; left: 15px !important; z-index: 999999 !important; }

    [data-testid="stSidebarCollapseButton"] button, [data-testid="collapsedControl"] button {
        background-color: var(--header-bg) !important; 
        border: 1px solid var(--card-border) !important; 
        border-radius: 50% !important; 
        width: 34px !important; 
        height: 34px !important; 
        color: #ffffff !important; 
        box-shadow: 0 4px 12px rgba(0,0,0,0.3) !important; 
        transition: all 0.2s ease !important;
    }
    [data-testid="stSidebarCollapseButton"] button:hover, [data-testid="collapsedControl"] button:hover {
        background-color: #3b82f6 !important; 
        border-color: #3b82f6 !important; 
        transform: scale(1.08);
    }
    [data-testid="stSidebarCollapseButton"] svg, [data-testid="collapsedControl"] svg { fill: #ffffff !important; color: #ffffff !important; }

    /* SIDEBAR */
    [data-testid="stSidebar"] { 
        background-color: var(--sidebar-bg) !important; 
        border-right: 1px solid var(--sidebar-border) !important; 
        padding-top: 0px !important; 
        padding-left: 10px !important; 
        padding-right: 10px !important; 
        padding-bottom: 12px !important; 
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.4rem !important; padding-top: 0px !important; }
    [data-testid="stSidebar"] label, [data-testid="stSidebar"] p, [data-testid="stSidebar"] span, [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3, [data-testid="stSidebar"] div { color: #ffffff !important; }

    .sidebar-logo-card { 
        background-color: #ffffff; 
        border-radius: 10px; 
        padding: 6px 8px; 
        margin-top: 0px !important; 
        margin-bottom: 14px !important; 
        box-shadow: 0 4px 12px rgba(0,0,0,0.3); 
        width: 100% !important; 
        box-sizing: border-box; 
        text-align: center; 
        display: block; 
    }
    [data-testid="stSidebar"] [data-testid="stImage"] { 
        width: 100% !important; 
        display: block !important; 
        margin-top: 4px !important; 
        margin-bottom: 8px !important; 
        clear: both !important; 
    }
    [data-testid="stImage"] img { 
        border-radius: 10px !important; 
        width: 100% !important; 
        height: auto !important; 
        max-width: 100% !important; 
        object-fit: cover !important; 
        border: 1px solid var(--card-border) !important; 
        margin: 0 !important; 
        display: block !important; 
    }

    .sidebar-profile-box { 
        background: var(--sidebar-card); 
        border: 1px solid var(--sidebar-border); 
        border-radius: 10px; 
        padding: 10px 8px !important; 
        text-align: center; 
        margin-top: 2px; 
        margin-bottom: 6px; 
        width: 100% !important; 
        box-shadow: 0 4px 10px rgba(0,0,0,0.3); 
        box-sizing: border-box; 
    }
    .sidebar-user-nombres { font-size: 0.84rem; font-weight: 800; color: #ffffff !important; line-height: 1.2; }
    .sidebar-user-apellidos { font-size: 0.80rem; font-weight: 700; color: #e2e8f0 !important; margin-bottom: 3px !important; line-height: 1.2; }
    .sidebar-user-email { font-size: 0.65rem; color: #60a5fa !important; font-weight: 600; margin-bottom: 4px !important; word-break: break-all; }
    .sidebar-user-cargo { 
        display: inline-block; 
        background: #1f2937 !important; 
        color: #ffffff !important; 
        border: 1px solid #374151 !important; 
        font-size: 0.58rem !important; 
        font-weight: 800 !important; 
        padding: 2px 7px !important; 
        border-radius: 12px !important; 
        text-transform: uppercase !important; 
    }

    [data-testid="stSidebar"] hr { margin: 4px 0 !important; border-color: var(--sidebar-border) !important; }

    /* SMART DASHBOARD GLASSMORPHISM */
    .smart-dashboard-container {
        background: radial-gradient(120% 120% at 50% 0%, #1e293b 0%, #0f172a 60%, #090d16 100%);
        border: 1px solid rgba(255, 255, 255, 0.12);
        border-radius: 14px;
        padding: 12px 14px;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.35);
        color: #ffffff !important;
        margin-top: 4px;
        margin-bottom: 12px;
        width: 100%;
        box-sizing: border-box;
    }
    .smart-dashboard-container * { color: #ffffff !important; }

    .smart-header-bar {
        display: flex;
        justify-content: space-between;
        align-items: center;
        flex-wrap: wrap;
        gap: 8px;
        border-bottom: 1px solid rgba(255, 255, 255, 0.08);
        padding-bottom: 8px;
        margin-bottom: 10px;
    }
    .smart-title { font-size: 1.1rem; font-weight: 800; letter-spacing: -0.02em; line-height: 1.1; }
    .smart-user-sub {
        font-size: 0.78rem;
        color: #94a3b8 !important;
        margin-top: 3px;
        display: flex;
        align-items: center;
        flex-wrap: wrap;
        gap: 6px;
    }

    .edificio-tag-badge {
        display: inline-flex;
        align-items: center;
        background: rgba(59, 130, 246, 0.20);
        border: 1px solid rgba(96, 165, 250, 0.45);
        color: #93c5fd !important;
        font-size: 0.64rem;
        font-weight: 800;
        padding: 2px 7px;
        border-radius: 6px;
        letter-spacing: 0.02em;
    }

    .smart-pill {
        display: inline-flex;
        align-items: center;
        gap: 4px;
        background: rgba(255, 255, 255, 0.06);
        border: 1px solid rgba(255, 255, 255, 0.12);
        backdrop-filter: blur(8px);
        padding: 3px 8px;
        border-radius: 16px;
        font-size: 0.70rem;
        font-weight: 700;
    }

    .widgets-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(135px, 1fr));
        gap: 8px;
    }

    .widget-glass-card {
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 10px;
        padding: 8px 10px;
        backdrop-filter: blur(10px);
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        box-sizing: border-box;
    }

    .w-card-title {
        font-size: 0.65rem;
        font-weight: 800;
        text-transform: uppercase;
        color: #94a3b8 !important;
        letter-spacing: 0.04em;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 4px;
    }

    .milestone-item {
        display: flex;
        justify-content: space-between;
        align-items: center;
        background: rgba(255, 255, 255, 0.03);
        border: 1px solid rgba(255, 255, 255, 0.05);
        border-radius: 6px;
        padding: 4px 6px;
        margin-bottom: 3px;
    }
    .milestone-name { font-size: 0.68rem; font-weight: 700; color: #e2e8f0 !important; }
    .milestone-status-done {
        background: #10b981;
        color: #ffffff !important;
        font-size: 0.58rem;
        font-weight: 800;
        padding: 1px 6px;
        border-radius: 10px;
    }
    .milestone-status-pending {
        background: rgba(239, 68, 68, 0.2);
        border: 1px solid rgba(239, 68, 68, 0.4);
        color: #fca5a5 !important;
        font-size: 0.58rem;
        font-weight: 800;
        padding: 1px 6px;
        border-radius: 10px;
    }

    /* DONA SVG */
    .donut-container { display: flex; align-items: center; gap: 8px; }
    .donut-chart-svg { width: 44px; height: 44px; transform: rotate(-90deg); flex-shrink: 0; }
    .donut-bg { fill: none; stroke: #334155 !important; stroke-width: 4.5; }
    .donut-progress { fill: none; stroke-width: 4.5; stroke-linecap: round; }
    .donut-info-val { font-size: 0.98rem; font-weight: 900; line-height: 1; }
    .donut-info-lbl { font-size: 0.60rem; color: #94a3b8 !important; font-weight: 600; margin-top: 1px; }
    .stat-hero-number { font-size: 1.35rem; font-weight: 900; line-height: 1; margin-top: 2px; margin-bottom: 2px; }

    /* PESTAÑAS */
    .stTabs [data-baseweb="tab-list"] { 
        gap: 6px !important; 
        background-color: var(--card-border) !important; 
        padding: 4px !important; 
        border-radius: 10px !important; 
        border: 1px solid var(--card-border) !important; 
    }
    .stTabs [data-baseweb="tab"] { 
        border-radius: 8px !important; 
        padding: 6px 14px !important; 
        background-color: var(--card-bg) !important; 
        border: 1px solid var(--card-border) !important; 
        box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important; 
        transition: all 0.2s ease !important; 
    }
    .stTabs [data-baseweb="tab"]:hover { border-color: var(--subtext) !important; background-color: var(--hover-bg) !important; }
    .stTabs [data-baseweb="tab"] p, .stTabs [data-baseweb="tab"] span { color: var(--text-app) !important; font-weight: 700 !important; font-size: 0.78rem !important; }
    .stTabs [aria-selected="true"] { 
        background-color: var(--header-bg) !important; 
        border: 1px solid var(--header-bg) !important; 
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.25) !important; 
    }
    .stTabs [aria-selected="true"] p, .stTabs [aria-selected="true"] span, .stTabs [aria-selected="true"] div { 
        color: var(--header-text) !important; 
        font-weight: 800 !important; 
    }

    .stButton > button { 
        background-color: var(--header-bg) !important; 
        color: #ffffff !important; 
        border-radius: 980px !important; 
        border: 1px solid var(--card-border) !important; 
        font-weight: 700 !important; 
        padding: 5px 12px !important; 
        font-size: 0.78rem !important; 
    }
    .stButton > button p, .stButton > button span { color: #ffffff !important; }

    .banner-item-header { 
        background-color: var(--header-bg) !important; 
        border: 1px solid var(--card-border) !important; 
        border-bottom: 2px solid #ffffff !important; 
        border-radius: 6px 6px 0 0 !important; 
        padding: 6px 10px !important; 
        margin-top: 4px !important; 
        margin-bottom: 0px !important; 
    }
    .banner-item-header span { color: #ffffff !important; font-size: 0.80rem !important; font-weight: 800 !important; }

    .card-item-body-compact { 
        border: 1px solid var(--card-border); 
        border-top: none; 
        border-radius: 0 0 6px 6px; 
        padding: 6px 8px 2px 8px; 
        background-color: var(--card-bg); 
        margin-bottom: 6px; 
    }

    .worker-card-row { 
        background: var(--card-bg); 
        border: 1px solid var(--card-border); 
        border-radius: 8px; 
        padding: 8px 12px; 
        margin-bottom: 6px; 
        display: flex; 
        align-items: center; 
        justify-content: space-between; 
        gap: 8px; 
        box-shadow: 0 1px 3px rgba(0,0,0,0.03); 
    }
    .worker-info-block { display: flex; flex-direction: column; flex: 1; min-width: 0; }
    .worker-name-title { font-size: 0.82rem; font-weight: 800; color: var(--text-app); line-height: 1.2; word-break: break-word; }
    .worker-meta-tags { display: flex; flex-wrap: wrap; gap: 4px; margin-top: 4px; align-items: center; }
    .tag-cargo-chip { 
        background: var(--card-border); 
        color: var(--text-app); 
        font-size: 0.68rem; 
        font-weight: 800; 
        padding: 2px 8px; 
        border-radius: 12px; 
        border: 1px solid var(--card-border); 
    }
    .tag-edif-chip { 
        background: rgba(59, 130, 246, 0.15); 
        color: #3b82f6; 
        font-size: 0.68rem; 
        font-weight: 800; 
        padding: 2px 8px; 
        border-radius: 12px; 
        border: 1px solid rgba(59, 130, 246, 0.35); 
    }

    @media (max-width: 768px) { 
        .block-container { padding-top: 3.6rem !important; padding-left: 0.4rem !important; padding-right: 0.4rem !important; }
        .widgets-grid { grid-template-columns: 1fr 1fr; }
        .smart-title { font-size: 0.95rem !important; }
        .worker-card-row { padding: 6px 8px; }
        .worker-name-title { font-size: 0.78rem; }
    }

    .incidencias-table { 
        width: 100%; 
        border-collapse: collapse !important; 
        margin-top: 4px; 
        margin-bottom: 8px; 
    }
    .incidencias-table th { 
        background-color: var(--header-bg) !important; 
        color: #ffffff !important; 
        padding: 6px 8px !important; 
        font-size: 0.75rem !important; 
        font-weight: 700 !important; 
        border: 1px solid var(--card-border) !important; 
        text-align: left; 
    }
    .incidencias-table th.center, .incidencias-table td.center { text-align: center !important; }
    .incidencias-table td { 
        padding: 5px 7px !important; 
        font-size: 0.76rem !important; 
        border: 1px solid var(--card-border) !important; 
        vertical-align: middle !important; 
        background-color: var(--card-bg) !important; 
        color: var(--text-app) !important; 
    }
    .incidencias-table tr:nth-child(even) td { 
        background-color: var(--table-row-even) !important; 
    }
    #MainMenu {visibility: hidden;} footer {visibility: hidden;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ==============================================================================
# 2. CONSTANTES INSTITUCIONALES Y ROLES
# ==============================================================================
CARGOS_DISPONIBLES = ["Residente", "Asistente", "Maestro Mayor"]

EDIFICIOS_ALPHA = [
    "Tesla",
    "Lafuente",
    "Imagine",
    "Asimov",
    "Rubik",
    "Castle Rock",
    "Musk",
    "Wolf",
    "Dablanc",
    "Thomas Edison",
    "Westinghouse",
    "Smart",
    "Magnus",
    "Sparta",
]

NOMBRES_MESES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

UNIDADES_RUBRO = {"Enlucidos": "m2", "Fijos": "m2", "Fajas": "m", "Dinteles": "m"}
RENDIMIENTOS_TEORICOS = {"Enlucidos": 0.75, "Fijos": 0.50, "Fajas": 0.30, "Dinteles": 0.40}

ACTIVIDADES_MANANA_CLEAN = [
    "Verificación de asistencia del personal",
    "Verificación de los trabajos y la calidad",
    "Coordinación con otras especialidades",
    "Detección de incidencias",
]

ACTIVIDADES_TARDE_CLEAN = [
    "Verificación del avance físico de las actividades",
    "Corrección de incidencias",
    "Confirmación de materiales para el siguiente día",
    "Revisión del cumplimiento de la meta diaria",
]

OFICIOS_NOMINA_FORMATO = [
    "ALBAÑILES", "AYUDANTE", "SOLDADOR", "OPERADOR", "FIERRERO",
    "PINTORES", "MAESTRO SUPERVISOR", "CARPINTERO", "GUACHIMAN",
    "GYPSEROS", "ELECTRICOS", "PLOMEROS", "ALUMINIO Y VIDRIO"
]

RUROS_ROTATIVOS_FORMATO = [
    "BLINDOBARRAS", "TUBERÍAS AGUA POTABLE", "SOLSYSTECH (ELÉCTRICOS)",
    "TUBERÍAS HIDROSANITARIAS", "ESTRUSA (VIDRIOS)", "ASCENSORES INTERNACIONALES",
    "TUBERÍAS CONTRA INCENDIOS", "PISO FLOTANTE", "SOLDEINSA",
    "MÁRMOL Y PORCELÁNICO COVEÑA", "TERRAMODA (JARDINERAS)"
]

MAQUINARIAS_FORMATO = [
    "EXTENSION 110", "SIERRA CIRCULAR", "AMOLADORA", "SOLDADORA 220 + CARETA",
    "ARNES", "VIBRADOR", "TALADRO", "TIJERA"
]
# ==============================================================================
# PARTE 2 DE 5: CARGA RÁPIDA DE DATOS BLINDADA Y EXPORTADORES EN CACHÉ
# ==============================================================================

# ==============================================================================
# 3. SERVICIO EN TIEMPO REAL: HORA LOCAL Y ESTADO DE CLIMA INSTANTÁNEO
# ==============================================================================
def get_local_datetime_ecuador():
    try:
        tz = zoneinfo.ZoneInfo("America/Guayaquil")
        return datetime.datetime.now(tz)
    except Exception:
        tz_offset = datetime.timezone(datetime.timedelta(hours=-5))
        return datetime.datetime.now(tz_offset)


def get_realtime_weather():
    local_dt = get_local_datetime_ecuador()
    h = local_dt.hour
    if 6 <= h < 12:
        return "☀️ 17°C Mañana Despejada"
    elif 12 <= h < 18:
        return "⛅ 21°C Tarde Templada"
    else:
        return "🌙 14°C Noche Fresca"


# ==============================================================================
# 4. BASE DE DATOS SUPABASE - INICIALIZACIÓN Y CARGA LIGERA BLINDADA
# ==============================================================================
@st.cache_resource
def init_supabase():
    url = st.secrets.get("SUPABASE_URL", "")
    key = st.secrets.get("SUPABASE_KEY", "")
    if not url or not key:
        st.error("⚠️ Credenciales SUPABASE_URL / SUPABASE_KEY no configuradas en st.secrets.")
        st.stop()
    return create_client(url, key)


supabase = init_supabase()


def load_db_from_supabase():
    access_pin = "1254"
    try:
        res_pin = supabase.table("app_config").select("value").eq("key", "access_pin").execute()
        if res_pin.data and len(res_pin.data) > 0:
            access_pin = res_pin.data[0]["value"]
    except Exception as e:
        print(f"[Warn] No se pudo leer access_pin: {e}")

    fallback_edificios_map = {}
    fallback_trabajadores_map = {}
    try:
        res_cfg = supabase.table("app_config").select("key, value").execute()
        if res_cfg.data:
            for r_cfg in res_cfg.data:
                k = r_cfg.get("key", "")
                if k.startswith("user_edificios_"):
                    u_k = k.replace("user_edificios_", "").lower().strip()
                    try:
                        fallback_edificios_map[u_k] = json.loads(r_cfg["value"])
                    except Exception:
                        fallback_edificios_map[u_k] = [r_cfg["value"]]
                elif k.startswith("user_trabajadores_"):
                    u_k = k.replace("user_trabajadores_", "").lower().strip()
                    try:
                        parsed_val = json.loads(r_cfg["value"])
                        if isinstance(parsed_val, list):
                            fallback_trabajadores_map[u_k] = [
                                item for item in parsed_val
                                if str(item.get("usuario_email", "")).lower().strip() == u_k
                            ]
                    except Exception:
                        pass
    except Exception as e:
        print(f"[Warn] Error en app_config: {e}")

    db_usuarios = []
    admin_emails = ["oscarsebitas2013@gmail.com"]
    try:
        res_usr = supabase.table("usuarios").select("correo, nombres, apellidos, password, cargo, edificios, fecha_registro, estado, es_admin").execute()
        if res_usr.data:
            for row in res_usr.data:
                c = str(row.get("correo", "")).lower().strip()
                if not c:
                    continue
                edifs = row.get("edificios")
                if isinstance(edifs, list):
                    edifs_list = [str(x).strip() for x in edifs if str(x).strip()]
                elif isinstance(edifs, str):
                    try:
                        parsed_edifs = json.loads(edifs)
                        edifs_list = [str(x).strip() for x in parsed_edifs if str(x).strip()] if isinstance(parsed_edifs, list) else [edifs.strip()]
                    except Exception:
                        edifs_list = [edifs.strip()] if edifs.strip() else []
                else:
                    edifs_list = fallback_edificios_map.get(c, [])

                db_usuarios.append({
                    "Nombres": row.get("nombres", ""),
                    "Apellidos": row.get("apellidos", ""),
                    "Correo": c,
                    "Password": str(row.get("password", "")),
                    "Cargo": row.get("cargo", "Residente"),
                    "Edificios": edifs_list,
                    "Fecha_Registro": str(row.get("fecha_registro", "")),
                    "Estado": row.get("estado", "Activo")
                })
                if row.get("es_admin") and c not in admin_emails:
                    admin_emails.append(c)
    except Exception as e:
        print(f"[Warn] Error en consulta usuarios: {e}")

    db_trabajadores_por_usuario = {u["Correo"]: [] for u in db_usuarios}
    for u_k, t_list in fallback_trabajadores_map.items():
        if u_k not in db_trabajadores_por_usuario:
            db_trabajadores_por_usuario[u_k] = []
        for t_item in t_list:
            if not any(x.get("id") == t_item.get("id") or x.get("nombre") == t_item.get("nombre") for x in db_trabajadores_por_usuario[u_k]):
                db_trabajadores_por_usuario[u_k].append(t_item)

    try:
        res_trab = supabase.table("trabajadores").select("*").order("id", desc=False).execute()
        if res_trab.data:
            for r in res_trab.data:
                try:
                    u_owner = str(r.get("usuario_email") or r.get("correo") or r.get("user_email") or "").lower().strip()
                    if not u_owner or u_owner in ["none", "null", "", "undefined"]:
                        continue

                    if u_owner not in db_trabajadores_por_usuario:
                        db_trabajadores_por_usuario[u_owner] = []

                    edif_val = r.get("edificio") or "General"
                    if not any(x.get("id") == r.get("id") or x.get("nombre") == r.get("nombre") for x in db_trabajadores_por_usuario[u_owner]):
                        db_trabajadores_por_usuario[u_owner].append({
                            "id": r.get("id"),
                            "nombre": r.get("nombre", ""),
                            "cargo": r.get("cargo", ""),
                            "edificio": edif_val,
                            "usuario_email": u_owner
                        })
                except Exception as ex_item:
                    print(f"[Warn] Error en fila trabajador: {ex_item}")
    except Exception as e:
        print(f"[Warn] Error en tabla trabajadores: {e}")

    # Ordenamiento alfabético automático obligatorio de la nómina
    for u_k in db_trabajadores_por_usuario:
        db_trabajadores_por_usuario[u_k] = sorted(
            db_trabajadores_por_usuario[u_k], 
            key=lambda item: str(item.get("nombre", "")).upper()
        )

    # -------------------------------------------------------------------------
    # CARGA BLINDADA DE CHECKLISTS (FILA POR FILA Y RESILIENTE A FORMATOS ANTIGUOS)
    # -------------------------------------------------------------------------
    db_checklists = {}
    try:
        res_chk = supabase.table("checklists").select("*").execute()
        if res_chk.data:
            for r in res_chk.data:
                try:
                    c = str(r.get("usuario_email") or r.get("correo") or r.get("user_email") or r.get("usuario") or "").lower().strip()
                    if not c:
                        resp_nom = str(r.get("responsable", "")).lower().strip()
                        matched_u = next((u["Correo"] for u in db_usuarios if f"{u['Nombres']} {u['Apellidos']}".lower().strip() == resp_nom), None)
                        c = matched_u if matched_u else "general"

                    if c not in db_checklists:
                        db_checklists[c] = []

                    raw_d = r.get("datos")
                    if isinstance(raw_d, (list, dict)):
                        datos_parsed = raw_d
                    elif isinstance(raw_d, str):
                        try:
                            datos_parsed = json.loads(raw_d)
                        except Exception:
                            datos_parsed = {}
                    else:
                        datos_parsed = {}

                    db_checklists[c].append({
                        "db_id": r.get("id"),
                        "Fecha": str(r.get("fecha", "")),
                        "Hora_Inicio": r.get("hora_inicio", "07:00"),
                        "Hora_Fin": r.get("hora_fin", "17:00"),
                        "Edificio": r.get("edificio", ""),
                        "Responsable": r.get("responsable", ""),
                        "Cargo": r.get("cargo", ""),
                        "Observacion_General": r.get("observacion_general", ""),
                        "Datos": datos_parsed
                    })
                except Exception as ex_chk_row:
                    print(f"[Warn] Error parseando fila checklist {r.get('id')}: {ex_chk_row}")
    except Exception as e:
        print(f"[Warn] Error en consulta tabla checklists: {e}")

    # -------------------------------------------------------------------------
    # CARGA BLINDADA DE INSPECCIONES / LIBRO DE OBRA
    # -------------------------------------------------------------------------
    db_inspecciones = {}
    try:
        res_insp = supabase.table("inspecciones").select("*").execute()
        if res_insp.data:
            for r in res_insp.data:
                try:
                    c = str(r.get("usuario_email") or r.get("correo") or r.get("user_email") or r.get("usuario") or "").lower().strip()
                    if not c:
                        res_nom = str(r.get("residente", "")).lower().strip()
                        matched_u = next((u["Correo"] for u in db_usuarios if f"{u['Nombres']} {u['Apellidos']}".lower().strip() == res_nom), None)
                        c = matched_u if matched_u else "general"

                    if c not in db_inspecciones:
                        db_inspecciones[c] = []

                    raw_d = r.get("datos")
                    if isinstance(raw_d, (list, dict)):
                        datos_parsed = raw_d
                    elif isinstance(raw_d, str):
                        try:
                            datos_parsed = json.loads(raw_d)
                        except Exception:
                            datos_parsed = {}
                    else:
                        datos_parsed = {}

                    db_inspecciones[c].append({
                        "db_id": r.get("id"),
                        "Fecha": str(r.get("fecha", "")),
                        "Dia": r.get("dia", ""),
                        "Proyecto": r.get("proyecto", ""),
                        "Residente": r.get("residente", ""),
                        "Frente": r.get("frente", ""),
                        "Clima": r.get("clima", ""),
                        "Hora_Inicio": r.get("hora_inicio", "07:00"),
                        "Hora_Fin": r.get("hora_fin", "17:00"),
                        "Datos": datos_parsed
                    })
                except Exception as ex_insp_row:
                    print(f"[Warn] Error parseando fila inspeccion {r.get('id')}: {ex_insp_row}")
    except Exception as e:
        print(f"[Warn] Error en consulta tabla inspecciones: {e}")

    # -------------------------------------------------------------------------
    # CARGA BLINDADA DE INCIDENCIAS
    # -------------------------------------------------------------------------
    db_incidencias_all = []
    try:
        res_inc = supabase.table("incidencias").select("*").execute()
        if res_inc.data:
            for r in res_inc.data:
                try:
                    c_inc = str(r.get("usuario_email") or r.get("correo") or r.get("user_email") or r.get("usuario") or "").lower().strip()
                    db_incidencias_all.append({
                        "db_id": r.get("id"),
                        "Area": r.get("area", ""),
                        "Descripcion": r.get("descripcion", ""),
                        "Responsable": r.get("responsable", ""),
                        "Prioridad": r.get("prioridad", "Media"),
                        "Fecha_Compromiso": str(r.get("fecha_compromiso", "")),
                        "Estado": r.get("estado", "Abierta"),
                        "Proyecto": r.get("proyecto", ""),
                        "Usuario": c_inc
                    })
                except Exception as ex_inc:
                    print(f"[Warn] Error parseando incidencia {r.get('id')}: {ex_inc}")
    except Exception as e:
        print(f"[Warn] Error en tabla incidencias: {e}")

    # -------------------------------------------------------------------------
    # CARGA BLINDADA DE RENDIMIENTOS
    # -------------------------------------------------------------------------
    db_rendimientos = {}
    try:
        res_rnd = supabase.table("rendimientos").select("*").execute()
        if res_rnd.data:
            for r in res_rnd.data:
                try:
                    c = str(r.get("usuario_email") or r.get("correo") or r.get("user_email") or "").lower().strip()
                    if not c:
                        continue
                    if c not in db_rendimientos:
                        db_rendimientos[c] = []
                    db_rendimientos[c].append({
                        "db_id": r.get("id"),
                        "Usuario_Registro": c,
                        "Cargo_Registrador": r.get("cargo_obrero", ""),
                        "Fecha": str(r.get("fecha", "")),
                        "Trabajador": r.get("trabajador", ""),
                        "Cargo_Obrero": r.get("cargo_obrero", ""),
                        "Rubro": r.get("rubro", ""),
                        "Intervalo": r.get("intervalo", "Jornada"),
                        "Horas Trabajadas (HH)": float(r.get("horas_hh") or 0.0),
                        "Avance": float(r.get("avance") or 0.0),
                        "Esperado": float(r.get("esperado") or 0.0),
                        "Unidad": r.get("unidad", "m2"),
                        "Rend. Real (HH/Unid)": float(r.get("rend_real") or 0.0),
                        "Rend. Teórico": float(r.get("rend_teorico") or 1.0),
                        "Estado": r.get("estado", "EFICIENTE")
                    })
                except Exception as ex_rnd:
                    print(f"[Warn] Error parseando rendimiento {r.get('id')}: {ex_rnd}")
    except Exception as e:
        print(f"[Warn] Error en tabla rendimientos: {e}")

    return {
        "access_pin": access_pin,
        "admin_emails": admin_emails,
        "db_usuarios": db_usuarios,
        "db_checklists": db_checklists,
        "db_inspecciones": db_inspecciones,
        "db_incidencias_all": db_incidencias_all,
        "db_rendimientos": db_rendimientos,
        "db_trabajadores_por_usuario": db_trabajadores_por_usuario,
    }


if "db_loaded" not in st.session_state or not st.session_state.db_loaded:
    p_data = load_db_from_supabase()
    st.session_state.access_pin = p_data["access_pin"]
    st.session_state.admin_emails = p_data["admin_emails"]
    st.session_state.db_usuarios = p_data["db_usuarios"]
    st.session_state.db_checklists = p_data["db_checklists"]
    st.session_state.db_inspecciones = p_data["db_inspecciones"]
    st.session_state.db_incidencias_all = p_data["db_incidencias_all"]
    st.session_state.db_rendimientos = p_data["db_rendimientos"]
    st.session_state.db_trabajadores_por_usuario = p_data["db_trabajadores_por_usuario"]
    st.session_state.db_loaded = True

if "db_incidencias_all" not in st.session_state:
    st.session_state.db_incidencias_all = []
if "db_trabajadores_por_usuario" not in st.session_state:
    st.session_state.db_trabajadores_por_usuario = {}
if "db_checklists" not in st.session_state:
    st.session_state.db_checklists = {}
if "db_inspecciones" not in st.session_state:
    st.session_state.db_inspecciones = {}
if "db_rendimientos" not in st.session_state:
    st.session_state.db_rendimientos = {}
if "db_usuarios" not in st.session_state:
    st.session_state.db_usuarios = []


# ==============================================================================
# 5. FUNCIONES DE FORMATO Y EXPORTADORES EN CACHÉ (RETROCOMPATIBLES)
# ==============================================================================
def render_estado_badge(estado_str):
    if not estado_str:
        return '<span style="color: #64748b; font-weight: 600;">Sin Responder</span>'
    if "Cumple" in estado_str or estado_str in ["Sí", "Operativo", "Completado", "Cerrada", "EFICIENTE", "CUMPLE META"]:
        return f'<span style="background-color: #dcfce7; color: #16a34a; font-weight: 800; padding: 2px 8px; border-radius: 6px; border: 1px solid #bbf7d0; font-size: 0.76rem;">{estado_str}</span>'
    elif "No" in estado_str or estado_str in ["Fuera de servicio", "Retrasado", "Abierta", "EXCESO DE HH", "BAJO RENDIMIENTO"]:
        return f'<span style="background-color: #fee2e2; color: #dc2626; font-weight: 800; padding: 2px 8px; border-radius: 6px; border: 1px solid #fca5a5; font-size: 0.76rem;">{estado_str}</span>'
    else:
        return f'<span style="background-color: #f1f5f9; color: #121318; font-weight: 800; padding: 2px 8px; border-radius: 6px; border: 1px solid #cbd5e1; font-size: 0.76rem;">{estado_str}</span>'


def image_to_base64(image_file):
    if image_file is not None:
        try:
            img = Image.open(image_file)
            img = ImageOps.exif_transpose(img)
            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            return base64.b64encode(buffered.getvalue()).decode("utf-8")
        except Exception:
            return None
    return None


def base64_to_image(b64_str):
    if b64_str:
        try:
            img_data = base64.b64decode(b64_str)
            img = Image.open(io.BytesIO(img_data))
            img = ImageOps.exif_transpose(img)
            return img
        except Exception:
            return None
    return None


def get_repo_image_b64(filenames):
    for filename in filenames:
        if os.path.exists(filename):
            try:
                with open(filename, "rb") as f:
                    return base64.b64encode(f.read()).decode("utf-8")
            except Exception:
                pass
    return None


def export_dataframe_to_excel_csv(df):
    df_clean = df.drop(columns=["Foto_B64", "Fotos", "db_id", "id", "usuario_email", "Responsables", "Encargados", "Observaciones"], errors="ignore")
    return df_clean.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")


def export_incidencias_to_excel(incidencias_list, proyecto_nombre="General"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    fill_header = PatternFill(start_color="121318", end_color="121318", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"LEVANTAMIENTO DE INCIDENCIAS - {proyecto_nombre.upper()}"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = fill_header
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    headers = ["N°", "Área", "Descripción", "Responsable", "Prioridad", "Fecha compromiso", "Estado"]
    ws.append(headers)
    ws.row_dimensions[2].height = 24

    for col_i in range(1, 8):
        c = ws.cell(row=2, column=col_i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    for idx, inc in enumerate(incidencias_list or [], 1):
        prioridad = str(inc.get("Prioridad", "Media"))
        estado = str(inc.get("Estado", "Abierta"))
        prio_str = f"Alta {'[X]' if prioridad == 'Alta' else '[ ]'}\nMedia {'[X]' if prioridad == 'Media' else '[ ]'}\nBaja {'[X]' if prioridad == 'Baja' else '[ ]'}"
        est_str = f"Abierta {'[X]' if estado == 'Abierta' else '[ ]'}\nCerrada {'[X]' if estado == 'Cerrada' else '[ ]'}"

        ws.append([
            idx,
            inc.get("Area", ""),
            inc.get("Descripcion", ""),
            inc.get("Responsable", ""),
            prio_str,
            str(inc.get("Fecha_Compromiso", "")),
            est_str,
        ])
        r_i = ws.max_row
        ws.row_dimensions[r_i].height = 55
        for c_idx in range(1, 8):
            cell = ws.cell(row=r_i, column=c_idx)
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if c_idx in [1, 5, 6, 7] else "left",
                vertical="center",
                wrap_text=True,
            )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_incidencias_to_pdf(incidencias_list, proyecto_nombre="General"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=25, bottomMargin=25)
    story = []

    title_style = ParagraphStyle("IncTitle", fontName="Helvetica-Bold", fontSize=12, textColor=colors.HexColor("#121318"), spaceAfter=8)
    header_style = ParagraphStyle("IncHeader", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle("IncCell", fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#121318"))
    cell_center = ParagraphStyle("IncCenter", fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#121318"), alignment=1)

    story.append(Paragraph(f"<b>LEVANTAMIENTO DE INCIDENCIAS — {proyecto_nombre.upper()}</b>", title_style))
    story.append(Spacer(1, 6))

    table_data = [[
        Paragraph("<b>N°</b>", header_style),
        Paragraph("<b>Área</b>", header_style),
        Paragraph("<b>Descripción</b>", header_style),
        Paragraph("<b>Responsable</b>", header_style),
        Paragraph("<b>Prioridad</b>", header_style),
        Paragraph("<b>Fecha compromiso</b>", header_style),
        Paragraph("<b>Estado</b>", header_style),
    ]]

    for idx, item in enumerate(incidencias_list or [], 1):
        prioridad = str(item.get("Prioridad", "Media"))
        estado = str(item.get("Estado", "Abierta"))
        prio_text = f"{'☑' if prioridad == 'Alta' else '☐'} Alta<br/>{'☑' if prioridad == 'Media' else '☐'} Media<br/>{'☑' if prioridad == 'Baja' else '☐'} Baja"
        est_text = f"{'☑' if estado == 'Abierta' else '☐'} Abierta<br/>{'☑' if estado == 'Cerrada' else '☐'} Cerrada"

        table_data.append([
            Paragraph(str(idx), cell_center),
            Paragraph(str(item.get("Area", "")), cell_style),
            Paragraph(str(item.get("Descripcion", "")), cell_style),
            Paragraph(str(item.get("Responsable", "")), cell_style),
            Paragraph(prio_text, cell_style),
            Paragraph(str(item.get("Fecha_Compromiso", "")), cell_center),
            Paragraph(est_text, cell_style),
        ])

    table = Table(table_data, colWidths=[25, 85, 175, 95, 65, 75, 65])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#121318")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_checklist_excel(jornada_dict_str):
    jornada_dict = json.loads(jornada_dict_str)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"CHECKLIST DE OBRA - {jornada_dict.get('Edificio', '')} ({jornada_dict.get('Fecha', '')})"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill(start_color="121318", end_color="121318", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = f"Hora Inicio: {jornada_dict.get('Hora_Inicio', 'N/A')}"
    ws["B2"] = f"Hora Fin: {jornada_dict.get('Hora_Fin', 'N/A')}"
    ws["C2"] = f"Responsable: {jornada_dict.get('Responsable', '')}"

    headers = ["Jornada / Sección", "N°", "Actividad", "Estado / Encargados", "Observaciones del Ítem"]
    ws.append([])
    ws.append(headers)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="121318", end_color="121318", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[4].height = 25
    datos_raw = jornada_dict.get("Datos", [])

    if isinstance(datos_raw, dict):
        items_verif = datos_raw.get("Verificaciones", [])
        items_sup = datos_raw.get("Supervision_Trabajos", [])
    elif isinstance(datos_raw, list):
        items_verif = datos_raw
        items_sup = []
    else:
        items_verif = []
        items_sup = []

    current_r = 5
    for item in items_verif:
        obs_val = item.get("Observaciones", [])
        obs_str = " | ".join([f"• {o}" for o in obs_val if o]) if isinstance(obs_val, list) else str(obs_val or "")
        estado_str = str(item.get("Estado", ""))
        if item.get("Actividad") == "Coordinación con otras especialidades" and estado_str == "✗ No Cumple":
            responsables = item.get("Responsables", [])
            if responsables:
                resp_str = " | ".join([f"{r.get('nombre', '')} ({r.get('area', '')})" for r in responsables if r.get('nombre')])
                extra = f" | {resp_str}"
            else:
                extra = ""
        else:
            extra = ""
        ws.append([item.get("Jornada", ""), item.get("N°", ""), item.get("Actividad", ""), estado_str + extra, obs_str])
        ws.row_dimensions[current_r].height = 28
        for c_i in range(1, 6):
            cell_txt = ws.cell(row=current_r, column=c_i)
            cell_txt.border = thin_border
            cell_txt.font = Font(name="Arial", size=9)
            cell_txt.alignment = Alignment(vertical="center", wrap_text=True)
        current_r += 1

    if items_sup:
        ws.append([])
        current_r += 1
        ws.merge_cells(f"A{current_r}:E{current_r}")
        ws[f"A{current_r}"] = "SUPERVISIÓN DE LA EJECUCIÓN DE TRABAJOS"
        ws[f"A{current_r}"].font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ws[f"A{current_r}"].fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        current_r += 1

        for idx_s, sup in enumerate(items_sup, 1):
            ws.append(["Supervisión", idx_s, sup.get("Actividad", ""), f"Encargados: {sup.get('Encargados', '')}", sup.get("Observaciones", "")])
            ws.row_dimensions[current_r].height = 120
            for c_i in range(1, 6):
                cell_txt = ws.cell(row=current_r, column=c_i)
                cell_txt.border = thin_border
                cell_txt.font = Font(name="Arial", size=9)
                cell_txt.alignment = Alignment(vertical="center", wrap_text=True)

            fotos = sup.get("Fotos", [])
            if fotos and len(fotos) > 0:
                try:
                    img_data = base64.b64decode(fotos[0])
                    img_pil = Image.open(io.BytesIO(img_data))
                    img_pil = ImageOps.exif_transpose(img_pil)
                    img_pil = img_pil.resize((500, 375), Image.Resampling.LANCZOS)
                    img_stream = io.BytesIO()
                    img_pil.save(img_stream, format="PNG", quality=100)
                    img_stream.seek(0)

                    img_xlsx = OpenpyxlImage(img_stream)
                    img_xlsx.width = 220
                    img_xlsx.height = 110
                    ws.add_image(img_xlsx, f"E{current_r}")
                except Exception:
                    pass
            current_r += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_checklist_pdf(jornada_dict_str):
    jornada_dict = json.loads(jornada_dict_str)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    title_style = ParagraphStyle('TitleStyle', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#121318'), alignment=1, spaceAfter=8)
    sub_style = ParagraphStyle('SubStyle', fontName='Helvetica', fontSize=8.5, textColor=colors.HexColor('#333333'), spaceAfter=8)
    sec_style = ParagraphStyle('SecStyle', fontName='Helvetica-Bold', fontSize=9.5, textColor=colors.HexColor('#1e293b'), spaceBefore=8, spaceAfter=4)
    header_style = ParagraphStyle('HeaderStyle', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle('CellStyle', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#121318'))

    story.append(Paragraph(f"CHECKLIST DE OBRA — {jornada_dict.get('Edificio', '').upper()}", title_style))
    story.append(Paragraph(f"<b>Fecha:</b> {jornada_dict.get('Fecha', '')} | <b>Horario:</b> {jornada_dict.get('Hora_Inicio', '')} - {jornada_dict.get('Hora_Fin', '')} | <b>Responsable:</b> {jornada_dict.get('Responsable', '')}", sub_style))

    datos_raw = jornada_dict.get("Datos", [])
    if isinstance(datos_raw, dict):
        items_verif = datos_raw.get("Verificaciones", [])
        items_sup = datos_raw.get("Supervision_Trabajos", [])
    elif isinstance(datos_raw, list):
        items_verif = datos_raw
        items_sup = []
    else:
        items_verif = []
        items_sup = []

    story.append(Paragraph("1. VERIFICACIONES DE JORNADA", sec_style))
    data_v = [[
        Paragraph("<b>Jornada</b>", header_style),
        Paragraph("<b>N°</b>", header_style),
        Paragraph("<b>Actividad</b>", header_style),
        Paragraph("<b>Estado</b>", header_style),
        Paragraph("<b>Observaciones Integradas</b>", header_style)
    ]]
    for item in items_verif:
        obs_val = item.get("Observaciones", [])
        obs_str = "<br/>".join([f"• {o}" for o in obs_val if o]) if isinstance(obs_val, list) else str(obs_val or "")
        estado_str = str(item.get("Estado", ""))
        if item.get("Actividad") == "Coordinación con otras especialidades" and estado_str == "✗ No Cumple":
            responsables = item.get("Responsables", [])
            if responsables:
                resp_str = "<br/>".join([f"<b>Responsable:</b> {r.get('nombre', '')} | <b>Área:</b> {r.get('area', '')}" for r in responsables if r.get('nombre')])
                extra = f"<br/>{resp_str}"
            else:
                extra = ""
        else:
            extra = ""
        data_v.append([
            Paragraph(str(item.get("Jornada", "")), cell_style),
            Paragraph(str(item.get("N°", "")), cell_style),
            Paragraph(str(item.get("Actividad", "")), cell_style),
            Paragraph(estado_str + extra, cell_style),
            Paragraph(obs_str, cell_style)
        ])

    table_v = Table(data_v, colWidths=[60, 25, 195, 75, 195])
    table_v.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#121318')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(table_v)

    if items_sup:
        story.append(Spacer(1, 6))
        story.append(Paragraph("2. SUPERVISIÓN DE LA EJECUCIÓN DE TRABAJOS", sec_style))
        data_s = [[
            Paragraph("<b>N°</b>", header_style),
            Paragraph("<b>Actividad a Ejecutar</b>", header_style),
            Paragraph("<b>Personal Encargado</b>", header_style),
            Paragraph("<b>Observaciones</b>", header_style)
        ]]
        for idx_s, sup in enumerate(items_sup, 1):
            data_s.append([
                Paragraph(str(idx_s), cell_style),
                Paragraph(str(sup.get("Actividad", "")), cell_style),
                Paragraph(str(sup.get("Encargados", "")), cell_style),
                Paragraph(str(sup.get("Observaciones", "")), cell_style)
            ])

        table_s = Table(data_s, colWidths=[25, 220, 160, 145])
        table_s.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        story.append(table_s)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_libro_maestro_excel(insp_dict_str):
    insp_dict = json.loads(insp_dict_str)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Maestro"

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )
    fill_header = PatternFill(start_color="121318", end_color="121318", fill_type="solid")

    ws.merge_cells("A1:E1")
    ws["A1"] = f"LIBRO DE OBRA — MAESTRO MAYOR ({insp_dict.get('Proyecto', '')})"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = fill_header
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append([f"Fecha: {insp_dict.get('Fecha', '')}", f"Proyecto: {insp_dict.get('Proyecto', '')}", f"Maestro: {insp_dict.get('Residente', '')}", "", ""])
    ws.append([])
    ws.append(["N°", "Actividad Ejecutada", "Personal a Cargo", "Cantidad Realizada", "Observaciones"])

    for col_i in range(1, 6):
        c = ws.cell(row=4, column=col_i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9.5)
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    raw_dm = insp_dict.get("Datos", {})
    d_parsed = raw_dm if isinstance(raw_dm, dict) else json.loads(raw_dm or "{}") if isinstance(raw_dm, str) else {}
    acts = d_parsed.get("Actividades_Maestro", []) if isinstance(d_parsed, dict) else d_parsed if isinstance(d_parsed, list) else []

    for idx_m, a_m in enumerate(acts, 1):
        pers_str = ", ".join(a_m.get("Personal_A_Cargo", [])) if isinstance(a_m.get("Personal_A_Cargo"), list) else str(a_m.get("Personal_A_Cargo", ""))
        ws.append([idx_m, a_m.get("Actividad", ""), pers_str, str(a_m.get("Cantidad", "")), a_m.get("Observaciones", "")])
        r_idx = ws.max_row
        for col_i in range(1, 6):
            cell = ws.cell(row=r_idx, column=col_i)
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border
            if col_i in [1, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_libro_maestro_pdf(insp_dict_str):
    insp_dict = json.loads(insp_dict_str)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    story = []

    title_style = ParagraphStyle('TitleM', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#121318'), alignment=1, spaceAfter=6)
    sub_style = ParagraphStyle('SubM', fontName='Helvetica', fontSize=8.5, textColor=colors.HexColor('#333333'), spaceAfter=8)
    hdr_tbl = ParagraphStyle('HdrM', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle('CellM', fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor('#0f172a'))
    cell_center = ParagraphStyle('CellMC', fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor('#0f172a'), alignment=1)

    story.append(Paragraph(f"LIBRO DE OBRA — MAESTRO MAYOR ({insp_dict.get('Proyecto', '').upper()})", title_style))
    story.append(Paragraph(f"<b>Fecha:</b> {insp_dict.get('Fecha', '')} | <b>Maestro Mayor:</b> {insp_dict.get('Residente', '')}", sub_style))

    data_m = [[Paragraph("<b>N°</b>", hdr_tbl), Paragraph("<b>Actividad</b>", hdr_tbl), Paragraph("<b>Personal a Cargo</b>", hdr_tbl), Paragraph("<b>Cantidad</b>", hdr_tbl), Paragraph("<b>Observaciones</b>", hdr_tbl)]]
    
    raw_dm = insp_dict.get("Datos", {})
    d_parsed = raw_dm if isinstance(raw_dm, dict) else json.loads(raw_dm or "{}") if isinstance(raw_dm, str) else {}
    acts = d_parsed.get("Actividades_Maestro", []) if isinstance(d_parsed, dict) else d_parsed if isinstance(d_parsed, list) else []

    for idx_m, a_m in enumerate(acts, 1):
        pers_str = ", ".join(a_m.get("Personal_A_Cargo", [])) if isinstance(a_m.get("Personal_A_Cargo"), list) else str(a_m.get("Personal_A_Cargo", ""))
        data_m.append([
            Paragraph(str(idx_m), cell_center),
            Paragraph(str(a_m.get("Actividad", "")), cell_style),
            Paragraph(pers_str, cell_style),
            Paragraph(str(a_m.get("Cantidad", "")), cell_center),
            Paragraph(str(a_m.get("Observaciones", "")), cell_style)
        ])

    table_m = Table(data_m, colWidths=[25, 175, 140, 95, 125])
    table_m.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#121318')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(table_m)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_libro_oficial_excel(insp_dict_str):
    insp_dict = json.loads(insp_dict_str)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Obra"

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    fill_header = PatternFill(start_color="121318", end_color="121318", fill_type="solid")
    fill_sub = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font_bold = Font(name="Arial", bold=True, color="000000", size=9)
    font_regular = Font(name="Arial", size=8.5)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"LIBRO DE OBRA — {insp_dict.get('Proyecto', '').upper()} (HOJA: {insp_dict.get('Hoja', '001')})"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = fill_header
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    meta = [
        ["CIUDAD:", "QUITO", "SUPERINTENDENTE:", insp_dict.get("Superintendente", "ING. PABLO ESPINOSA"), "", "FECHA:", f"{insp_dict.get('Fecha', '')} ({insp_dict.get('Dia', '')})"],
        ["UBICACIÓN:", insp_dict.get("Ubicacion", "CALLE LUXEMBURGO Y HOLANDA"), "RESIDENTE:", insp_dict.get("Residente", ""), "", "HORARIO:", f"{insp_dict.get('Hora_Inicio', '07:00')} - {insp_dict.get('Hora_Fin', '16:00')}"],
        ["BARRIO:", insp_dict.get("Barrio", "BENALCAZAR"), "FISCALIZADOR:", insp_dict.get("Fiscalizador", "ING. DIEGO CHARVET"), "", "PROYECTO:", insp_dict.get("Proyecto", "")]
    ]
    for row_m in meta:
        ws.append(row_m)
        r_i = ws.max_row
        for col_i in range(1, 8):
            c = ws.cell(row=r_i, column=col_i)
            c.font = font_bold if col_i in [1, 3, 6] else font_regular
            c.border = thin_border

    ws.append([])
    r_s = ws.max_row + 1
    ws.merge_cells(f"A{r_s}:G{r_s}")
    ws[f"A{r_s}"] = "1. JORNADA DE TRABAJO Y NÓMINA DE PERSONAL"
    ws[f"A{r_s}"].font = Font(name="Arial", bold=True, color="FFFFFF", size=9.5)
    ws[f"A{r_s}"].fill = fill_sub
    ws.row_dimensions[r_s].height = 20

    ws.append(["PERSONAL NÓMINA", "No.", "ENTRADA", "SALIDA", "PERSONAL ROTATIVO / SUBCONTRATO", "No.", "HORARIO"])
    r_hdr_p = ws.max_row
    for col_i in range(1, 8):
        c = ws.cell(row=r_hdr_p, column=col_i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=8.5)
        c.fill = fill_header
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    raw_dinsp = insp_dict.get("Datos", {})
    datos = raw_dinsp if isinstance(raw_dinsp, dict) else json.loads(raw_dinsp or "{}") if isinstance(raw_dinsp, str) else {}
    nomina_map = datos.get("Nomina_Conteo", {}) if isinstance(datos, dict) else {}
    rotativo_map = datos.get("Rotativo_Conteo", {}) if isinstance(datos, dict) else {}

    max_len = max(len(OFICIOS_NOMINA_FORMATO), len(RUROS_ROTATIVOS_FORMATO))
    for i in range(max_len):
        oficio_nom = OFICIOS_NOMINA_FORMATO[i] if i < len(OFICIOS_NOMINA_FORMATO) else ""
        cant_nom = nomina_map.get(oficio_nom, 0) if (oficio_nom and isinstance(nomina_map, dict)) else ""
        ent_nom = "7:00AM" if (oficio_nom and cant_nom and int(cant_nom) > 0) else ""
        sal_nom = "4:00PM" if (oficio_nom and cant_nom and int(cant_nom) > 0) else ""

        rubro_rot = RUROS_ROTATIVOS_FORMATO[i] if i < len(RUROS_ROTATIVOS_FORMATO) else ""
        cant_rot = rotativo_map.get(rubro_rot, 0) if (rubro_rot and isinstance(rotativo_map, dict)) else ""
        hor_rot = "7:00AM - 4:00PM" if (rubro_rot and cant_rot and int(cant_rot) > 0) else ""

        ws.append([oficio_nom, cant_nom if cant_nom else "", ent_nom, sal_nom, rubro_rot, cant_rot if cant_rot else "", hor_rot])
        r_row = ws.max_row
        for col_i in range(1, 8):
            c = ws.cell(row=r_row, column=col_i)
            c.font = font_regular
            c.border = thin_border
            if col_i in [2, 3, 4, 6, 7]:
                c.alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])
    r_act = ws.max_row + 1
    ws.merge_cells(f"A{r_act}:G{r_act}")
    ws[f"A{r_act}"] = "2. ACTIVIDADES REALIZADAS DENTRO DE LA JORNADA LABORAL"
    ws[f"A{r_act}"].font = Font(name="Arial", bold=True, color="FFFFFF", size=9.5)
    ws[f"A{r_act}"].fill = fill_sub
    ws.row_dimensions[r_act].height = 20

    ws.append(["N°", "DESCRIPCIÓN DE LA ACTIVIDAD", "ÁREA DE TRABAJO", "UNIDAD", "CANT. REALIZADA"])
    r_hdr_a = ws.max_row
    for col_i in range(1, 6):
        c = ws.cell(row=r_hdr_a, column=col_i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=8.5)
        c.fill = fill_header
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    acts = datos.get("Actividades_Ejecutadas", []) if isinstance(datos, dict) else []
    for idx_a, a in enumerate(acts, 1):
        ws.append([
            idx_a,
            a.get("Descripcion", a.get("actividad", "")),
            a.get("Area", ""),
            a.get("Unidad", "m2"),
            a.get("Cantidad", 0.0)
        ])
        r_row_a = ws.max_row
        for col_i in range(1, 6):
            c = ws.cell(row=r_row_a, column=col_i)
            c.font = font_regular
            c.border = thin_border
            if col_i in [1, 4, 5]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    ws.append([])
    r_eq = ws.max_row + 1
    ws.merge_cells(f"A{r_eq}:G{r_eq}")
    ws[f"A{r_eq}"] = "3. CONDICIONES CLIMÁTICAS, EQUIPOS Y SEGURIDAD"
    ws[f"A{r_eq}"].font = Font(name="Arial", bold=True, color="FFFFFF", size=9.5)
    ws[f"A{r_eq}"].fill = fill_sub

    clima_val = insp_dict.get("Clima", "Soleado")
    maq_map = datos.get("Maquinaria_Conteo", {}) if isinstance(datos, dict) else {}

    ws.append(["Clima Registrado:", clima_val, "Obs. Clima:", datos.get("Clima_Obs", "") if isinstance(datos, dict) else "", "Equipos Operativos:", ", ".join([f"{k}: {v}" for k, v in maq_map.items() if v > 0]), ""])
    r_row_e = ws.max_row
    for col_i in range(1, 8):
        ws.cell(row=r_row_e, column=col_i).font = font_regular
        ws.cell(row=r_row_e, column=col_i).border = thin_border

    ws.append([])
    r_nov = ws.max_row + 1
    ws.merge_cells(f"A{r_nov}:G{r_nov}")
    ws[f"A{r_nov}"] = "4. NOVEDADES Y RECOMENDACIONES"
    ws[f"A{r_nov}"].font = Font(name="Arial", bold=True, color="FFFFFF", size=9.5)
    ws[f"A{r_nov}"].fill = fill_sub

    ws.merge_cells(f"A{r_nov+1}:G{r_nov+2}")
    ws[f"A{r_nov+1}"] = datos.get("Novedades", "Sin novedades técnicas en la jornada laboral.") if isinstance(datos, dict) else "Sin novedades técnicas."
    ws[f"A{r_nov+1}"].font = font_regular
    ws[f"A{r_nov+1}"].alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@st.cache_data(show_spinner=False, max_entries=50)
def get_cached_libro_oficial_pdf(insp_dict_str):
    insp_dict = json.loads(insp_dict_str)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    story = []

    title_style = ParagraphStyle('TitleL', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#121318'), alignment=1, spaceAfter=4)
    hdr_tbl = ParagraphStyle('HdrL', fontName='Helvetica-Bold', fontSize=7, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle('CellL', fontName='Helvetica', fontSize=6.5, textColor=colors.HexColor('#0f172a'))
    cell_center = ParagraphStyle('CellC', fontName='Helvetica', fontSize=6.5, textColor=colors.HexColor('#0f172a'), alignment=1)

    story.append(Paragraph(f"<b>LIBRO DE OBRA — {insp_dict.get('Proyecto', '').upper()}</b> (HOJA N° {insp_dict.get('Hoja', '001')})", title_style))

    raw_dinsp = insp_dict.get("Datos", {})
    datos = raw_dinsp if isinstance(raw_dinsp, dict) else json.loads(raw_dinsp or "{}") if isinstance(raw_dinsp, str) else {}
    
    meta_data = [
        [Paragraph("<b>CIUDAD:</b> QUITO", cell_style), Paragraph(f"<b>SUPERINTENDENTE:</b> {insp_dict.get('Superintendente', 'ING. PABLO ESPINOSA')}", cell_style), Paragraph(f"<b>FECHA:</b> {insp_dict.get('Fecha', '')} ({insp_dict.get('Dia', '')})", cell_style)],
        [Paragraph(f"<b>UBICACIÓN:</b> {insp_dict.get('Ubicacion', 'CALLE LUXEMBURGO Y HOLANDA')}", cell_style), Paragraph(f"<b>RESIDENTE:</b> {insp_dict.get('Residente', '')}", cell_style), Paragraph(f"<b>HORARIO:</b> {insp_dict.get('Hora_Inicio', '07:00')} - {insp_dict.get('Hora_Fin', '16:00')}", cell_style)],
        [Paragraph(f"<b>BARRIO:</b> {insp_dict.get('Barrio', 'BENALCAZAR')}", cell_style), Paragraph(f"<b>FISCALIZADOR:</b> {insp_dict.get('Fiscalizador', 'ING. DIEGO CHARVET')}", cell_style), Paragraph(f"<b>PROYECTO:</b> {insp_dict.get('Proyecto', '')}", cell_style)]
    ]
    t_meta = Table(meta_data, colWidths=[240, 260, 250])
    t_meta.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 4))

    nomina_map = datos.get("Nomina_Conteo", {}) if isinstance(datos, dict) else {}
    rotativo_map = datos.get("Rotativo_Conteo", {}) if isinstance(datos, dict) else {}
    p_data = [[Paragraph("<b>PERSONAL NÓMINA</b>", hdr_tbl), Paragraph("<b>No.</b>", hdr_tbl), Paragraph("<b>HORARIO</b>", hdr_tbl), Paragraph("<b>PERSONAL ROTATIVO / SUBCONTRATO</b>", hdr_tbl), Paragraph("<b>No.</b>", hdr_tbl), Paragraph("<b>HORARIO</b>", hdr_tbl)]]

    max_len = max(len(OFICIOS_NOMINA_FORMATO), len(RUROS_ROTATIVOS_FORMATO))
    for i in range(max_len):
        oficio_nom = OFICIOS_NOMINA_FORMATO[i] if i < len(OFICIOS_NOMINA_FORMATO) else ""
        cant_nom = str(nomina_map.get(oficio_nom, 0)) if (oficio_nom and isinstance(nomina_map, dict) and int(nomina_map.get(oficio_nom, 0)) > 0) else ""
        hor_nom = "7:00AM - 4:00PM" if cant_nom else ""

        rubro_rot = RUROS_ROTATIVOS_FORMATO[i] if i < len(RUROS_ROTATIVOS_FORMATO) else ""
        cant_rot = str(rotativo_map.get(rubro_rot, 0)) if (rubro_rot and isinstance(rotativo_map, dict) and int(rotativo_map.get(rubro_rot, 0)) > 0) else ""
        hor_rot = "7:00AM - 4:00PM" if cant_rot else ""

        p_data.append([
            Paragraph(oficio_nom, cell_style),
            Paragraph(cant_nom, cell_center),
            Paragraph(hor_nom, cell_center),
            Paragraph(rubro_rot, cell_style),
            Paragraph(cant_rot, cell_center),
            Paragraph(hor_rot, cell_center)
        ])

    t_pers = Table(p_data, colWidths=[150, 40, 110, 290, 40, 120])
    t_pers.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#121318')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('TOPPADDING', (0,0), (-1,-1), 1.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1.5),
    ]))
    story.append(t_pers)
    story.append(Spacer(1, 4))

    acts = datos.get("Actividades_Ejecutadas", []) if isinstance(datos, dict) else []
    if acts:
        a_data = [[
            Paragraph("<b>N°</b>", hdr_tbl),
            Paragraph("<b>DESCRIPCIÓN DE LA ACTIVIDAD</b>", hdr_tbl),
            Paragraph("<b>ÁREA DE TRABAJO</b>", hdr_tbl),
            Paragraph("<b>UNIDAD</b>", hdr_tbl),
            Paragraph("<b>CANTIDAD</b>", hdr_tbl)
        ]]
        for idx_a, a in enumerate(acts, 1):
            a_data.append([
                Paragraph(str(idx_a), cell_center),
                Paragraph(a.get("Descripcion", a.get("actividad", "")), cell_style),
                Paragraph(a.get("Area", "Frente Principal"), cell_style),
                Paragraph(a.get("Unidad", "m2"), cell_center),
                Paragraph(str(a.get("Cantidad", 0.0)), cell_center)
            ])
        t_acts = Table(a_data, colWidths=[25, 295, 150, 70, 70])
        t_acts.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        story.append(t_acts)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
# ==============================================================================
# PARTE 3 DE 5: AUTENTICACIÓN PERSISTENTE, BARRA LATERAL Y SMART DASHBOARD
# ==============================================================================

# ==============================================================================
# PERSISTENCIA INMEDIATA Y ROBUSTA DE SESIÓN (LOCALSTORAGE SEGURO + QUERY PARAMS)
# ==============================================================================
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.usuario_email = ""
    st.session_state.usuario_nombres = ""
    st.session_state.usuario_apellidos = ""
    st.session_state.usuario_cargo = ""
    st.session_state.usuario_edificios = []

url_user = st.query_params.get("u")
if url_user and not st.session_state.autenticado:
    m_clean = str(url_user).strip().lower()
    u_match = next((u for u in st.session_state.db_usuarios if u["Correo"] == m_clean), None)
    if u_match:
        st.session_state.autenticado = True
        st.session_state.usuario_email = m_clean
        st.session_state.usuario_nombres = u_match["Nombres"]
        st.session_state.usuario_apellidos = u_match["Apellidos"]
        st.session_state.usuario_cargo = u_match["Cargo"]
        st.session_state.usuario_edificios = u_match.get("Edificios", [])

if not st.session_state.autenticado:
    components.html(
        """
        <script>
        try {
            const win = window.top || window.parent || window;
            const saved = win.localStorage.getItem('alpha_user_session');
            if (saved) {
                const currentUrl = new URL(win.location.href);
                if (!currentUrl.searchParams.get('u')) {
                    currentUrl.searchParams.set('u', saved);
                    win.location.href = currentUrl.href;
                }
            }
        } catch (e) {}
        </script>
        """,
        height=0,
        width=0
    )

# ==============================================================================
# 6. MÓDULO DE AUTENTICACIÓN: LOGIN DIRECTO, REGISTRO Y RECUPERACIÓN
# ==============================================================================
if not st.session_state.autenticado:
    col_l1, col_l2, col_l3 = st.columns([1, 2, 1])

    with col_l2:
        if os.path.exists("images.png"):
            with open("images.png", "rb") as image_file:
                encoded_logo = base64.b64encode(image_file.read()).decode("utf-8")
            st.markdown(
                f"""
                <div style="text-align: center; margin-top: 10px; margin-bottom: 10px;">
                    <img src="data:image/png;base64,{encoded_logo}" style="width: 280px; max-width: 100%; pointer-events: none;">
                </div>
                """,
                unsafe_allow_html=True,
            )

        tab_login, tab_register, tab_reset = st.tabs(["Iniciar Sesión", "Registrarse", "¿Olvidaste tu Contraseña?"])

        with tab_login:
            st.markdown("### Iniciar Sesión")
            st.caption("Ingrese sus credenciales registradas y el código de acceso.")

            with st.form("form_login_clean"):
                login_email = st.text_input("Correo electrónico:", placeholder="nombre@correo.com", key="log_email")
                login_pass = st.text_input("Contraseña:", type="password", key="log_pass")
                login_pin = st.text_input("Código de Seguridad (PIN de 4 dígitos):", type="password", max_chars=4, placeholder="****", key="log_pin")
                mantener_sesion = st.checkbox("🔒 Mantener sesión iniciada en este dispositivo", value=True, key="chk_keep_session")

                btn_log = st.form_submit_button("Entrar al Portal", type="primary", use_container_width=True)

            if btn_log:
                if login_email and login_pass and login_pin:
                    mail_clean = login_email.strip().lower()

                    u_match = None
                    try:
                        res_direct = supabase.table("usuarios").select("correo, nombres, apellidos, password, cargo, edificios").ilike("correo", mail_clean).execute()
                        if res_direct.data and len(res_direct.data) > 0:
                            row = res_direct.data[0]
                            edifs = row.get("edificios")
                            if isinstance(edifs, list):
                                edifs_list = [str(x).strip() for x in edifs if str(x).strip()]
                            elif isinstance(edifs, str):
                                try:
                                    parsed_edifs = json.loads(edifs)
                                    edifs_list = [str(x).strip() for x in parsed_edifs if str(x).strip()] if isinstance(parsed_edifs, list) else [edifs.strip()]
                                except Exception:
                                    edifs_list = [edifs.strip()] if edifs.strip() else []
                            else:
                                edifs_list = []

                            u_match = {
                                "Nombres": row.get("nombres", ""),
                                "Apellidos": row.get("apellidos", ""),
                                "Correo": str(row.get("correo", "")).lower().strip(),
                                "Password": str(row.get("password", "")),
                                "Cargo": row.get("cargo", "Residente"),
                                "Edificios": edifs_list
                            }
                    except Exception as e:
                        print(f"Consulta directa Supabase error: {e}")

                    if not u_match:
                        u_match = next((u for u in st.session_state.db_usuarios if u["Correo"].lower().strip() == mail_clean), None)

                    if u_match:
                        if u_match["Password"] == login_pass.strip():
                            current_pin = "1254"
                            try:
                                res_pin = supabase.table("app_config").select("value").eq("key", "access_pin").execute()
                                if res_pin.data and len(res_pin.data) > 0:
                                    current_pin = res_pin.data[0]["value"]
                            except Exception:
                                current_pin = st.session_state.get("access_pin", "1254")

                            if login_pin.strip() == current_pin.strip():
                                st.session_state.autenticado = True
                                st.session_state.usuario_email = mail_clean
                                st.session_state.usuario_nombres = u_match["Nombres"]
                                st.session_state.usuario_apellidos = u_match["Apellidos"]
                                st.session_state.usuario_cargo = u_match["Cargo"]
                                st.session_state.usuario_edificios = u_match.get("Edificios", [])
                                
                                st.query_params["u"] = mail_clean
                                
                                if mantener_sesion:
                                    components.html(
                                        f"""
                                        <script>
                                        try {{
                                            const win = window.top || window.parent || window;
                                            win.localStorage.setItem('alpha_user_session', '{mail_clean}');
                                        }} catch(e) {{}}
                                        </script>
                                        """,
                                        height=0,
                                        width=0
                                    )
                                
                                if "db_trabajadores_por_usuario" not in st.session_state:
                                    st.session_state.db_trabajadores_por_usuario = {}
                                
                                try:
                                    res_t_user = supabase.table("trabajadores").select("*").ilike("usuario_email", mail_clean).order("id", desc=False).execute()
                                    st.session_state.db_trabajadores_por_usuario[mail_clean] = [
                                        {
                                            "id": r.get("id"),
                                            "nombre": r.get("nombre", ""),
                                            "cargo": r.get("cargo", ""),
                                            "edificio": r.get("edificio") or "General",
                                            "usuario_email": mail_clean
                                        } for r in res_t_user.data
                                    ] if res_t_user.data else []
                                except Exception:
                                    if mail_clean not in st.session_state.db_trabajadores_por_usuario:
                                        st.session_state.db_trabajadores_por_usuario[mail_clean] = []

                                # Ordenamiento alfabético automático
                                st.session_state.db_trabajadores_por_usuario[mail_clean] = sorted(
                                    st.session_state.db_trabajadores_por_usuario[mail_clean],
                                    key=lambda it: str(it.get("nombre", "")).upper()
                                )

                                st.session_state.db_loaded = False
                                st.success("Acceso concedido...")
                                st.rerun()
                            else:
                                st.error("⚠️ Código de Seguridad (PIN) incorrecto.")
                        else:
                            st.error("Contraseña incorrecta.")
                    else:
                        st.error("El usuario no existe. Complete el registro.")
                else:
                    st.error("Por favor complete todos los campos, incluyendo el código PIN de 4 dígitos.")

        with tab_register:
            st.markdown("### Crear una Cuenta Nueva")
            st.caption("Complete la información y seleccione los proyectos en los que labora.")

            with st.form("form_register_clean"):
                col_n, col_a = st.columns(2)
                with col_n:
                    reg_nombres = st.text_input("Nombres:*", placeholder="Ej. Juan Carlos")
                with col_a:
                    reg_apellidos = st.text_input("Apellidos:*", placeholder="Ej. Pérez Gómez")

                reg_email = st.text_input("Correo electrónico:*", placeholder="ejemplo@correo.com", key="reg_email")
                
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    reg_pass = st.text_input("Crear contraseña:*", type="password", key="reg_pass")
                with col_p2:
                    reg_pass_repeat = st.text_input("Repetir contraseña:*", type="password", key="reg_pass_rep")

                reg_cargo = st.selectbox("Cargo / Rol en Obra:*", CARGOS_DISPONIBLES)
                reg_edificios_sel = st.multiselect(
                    "Edificios / Proyectos asignados:*",
                    options=EDIFICIOS_ALPHA,
                    default=[],
                    placeholder="Seleccione los proyectos en los que trabaja...",
                    key="reg_edif_multisel"
                )
                reg_pin = st.text_input("Código de Seguridad de Registro (PIN de 4 dígitos):*", type="password", max_chars=4, placeholder="****", key="reg_pin")
                reg_mantener = st.checkbox("🔒 Mantener sesión iniciada automáticamente", value=True, key="chk_reg_keep")
                btn_reg = st.form_submit_button("Completar Registro", type="primary", use_container_width=True)

            if btn_reg:
                if reg_nombres and reg_apellidos and reg_email and reg_pass and reg_pass_repeat and reg_pin:
                    current_pin = st.session_state.get("access_pin", "1254")
                    try:
                        res_pin_chk = supabase.table("app_config").select("value").eq("key", "access_pin").execute()
                        if res_pin_chk.data and len(res_pin_chk.data) > 0:
                            current_pin = res_pin_chk.data[0]["value"]
                    except Exception:
                        pass

                    if reg_pin.strip() != current_pin.strip():
                        st.error("⚠️ Código de Seguridad (PIN) incorrecto.")
                    elif reg_pass != reg_pass_repeat:
                        st.error("Las contraseñas no coinciden.")
                    else:
                        mail_clean = reg_email.strip().lower()
                        exists = False
                        try:
                            chk_dup = supabase.table("usuarios").select("correo").ilike("correo", mail_clean).execute()
                            if chk_dup.data and len(chk_dup.data) > 0:
                                exists = True
                        except Exception:
                            exists = any(u["Correo"].lower().strip() == mail_clean for u in st.session_state.db_usuarios)

                        if exists:
                            st.warning("Este correo ya se encuentra registrado.")
                        else:
                            try:
                                insert_payload = {
                                    "correo": mail_clean,
                                    "nombres": reg_nombres.strip(),
                                    "apellidos": reg_apellidos.strip(),
                                    "password": reg_pass.strip(),
                                    "cargo": reg_cargo,
                                    "es_admin": False
                                }
                                try:
                                    payload_full = insert_payload.copy()
                                    payload_full["edificios"] = reg_edificios_sel
                                    supabase.table("usuarios").insert(payload_full).execute()
                                except Exception:
                                    supabase.table("usuarios").insert(insert_payload).execute()
                                    supabase.table("app_config").upsert({
                                        "key": f"user_edificios_{mail_clean}",
                                        "value": json.dumps(reg_edificios_sel)
                                    }).execute()

                                supabase.table("app_config").upsert({
                                    "key": f"user_trabajadores_{mail_clean}",
                                    "value": json.dumps([])
                                }).execute()

                                st.session_state.autenticado = True
                                st.session_state.usuario_email = mail_clean
                                st.session_state.usuario_nombres = reg_nombres.strip()
                                st.session_state.usuario_apellidos = reg_apellidos.strip()
                                st.session_state.usuario_cargo = reg_cargo
                                st.session_state.usuario_edificios = reg_edificios_sel
                                
                                st.query_params["u"] = mail_clean
                                if reg_mantener:
                                    components.html(
                                        f"""
                                        <script>
                                        try {{
                                            const win = window.top || window.parent || window;
                                            win.localStorage.setItem('alpha_user_session', '{mail_clean}');
                                        }} catch(e) {{}}
                                        </script>
                                        """,
                                        height=0,
                                        width=0
                                    )
                                
                                if "db_trabajadores_por_usuario" not in st.session_state:
                                    st.session_state.db_trabajadores_por_usuario = {}
                                st.session_state.db_trabajadores_por_usuario[mail_clean] = []

                                st.session_state.db_loaded = False
                                st.success("¡Registro completado exitosamente!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al guardar usuario: {e}")
                else:
                    st.error("Por favor complete todos los campos requeridos (*).")

        with tab_reset:
            st.markdown("### Recuperación de Contraseña")
            st.caption("Restablezca su acceso de forma segura.")

            with st.form("form_reset_clean"):
                reset_email = st.text_input("Ingrese su correo registrado:", placeholder="ejemplo@correo.com", key="rst_email")
                col_rp1, col_rp2 = st.columns(2)
                with col_rp1:
                    new_pass = st.text_input("Nueva contraseña:", type="password", key="rst_pass")
                with col_rp2:
                    new_pass_rep = st.text_input("Repetir nueva contraseña:", type="password", key="rst_pass_rep")
                btn_reset = st.form_submit_button("Restablecer Contraseña", type="primary", use_container_width=True)

            if btn_reset:
                if reset_email and new_pass and new_pass_rep:
                    if new_pass != new_pass_rep:
                        st.error("Las contraseñas no coinciden.")
                    else:
                        mail_clean = reset_email.strip().lower()
                        try:
                            res_rst = supabase.table("usuarios").update({"password": new_pass.strip()}).ilike("correo", mail_clean).execute()
                            if res_rst.data and len(res_rst.data) > 0:
                                st.session_state.db_loaded = False
                                st.success(f"Contraseña actualizada con éxito para {mail_clean}.")
                                st.rerun()
                            else:
                                st.error("El correo ingresado no está registrado.")
                        except Exception as e:
                            st.error(f"Error actualizando contraseña: {e}")
                else:
                    st.error("Complete todos los campos.")

    st.stop()

# ==============================================================================
# 7. BARRA LATERAL (USUARIO AUTENTICADO)
# ==============================================================================
user_email = st.session_state.usuario_email
user_nombres = st.session_state.usuario_nombres
user_apellidos = st.session_state.usuario_apellidos
user_cargo = st.session_state.usuario_cargo
user_edificios = st.session_state.get("usuario_edificios", [])
es_admin = user_email in st.session_state.admin_emails
es_maestro_mayor = (user_cargo == "Maestro Mayor")

if f"foto_user_{user_email}" not in st.session_state:
    try:
        res_f = supabase.table("usuarios").select("foto_b64").ilike("correo", user_email).execute()
        if res_f.data and len(res_f.data) > 0 and res_f.data[0].get("foto_b64"):
            st.session_state[f"foto_user_{user_email}"] = res_f.data[0]["foto_b64"]
        else:
            st.session_state[f"foto_user_{user_email}"] = None
    except Exception:
        st.session_state[f"foto_user_{user_email}"] = None

with st.sidebar:
    logo_filename = "alpha.473f0c2dc3c48a682723-2.webp"
    if not os.path.exists(logo_filename):
        logo_filename = "images.png"

    if os.path.exists(logo_filename):
        ext = "webp" if logo_filename.endswith(".webp") else "png"
        with open(logo_filename, "rb") as image_file:
            encoded_sidebar_logo = base64.b64encode(image_file.read()).decode("utf-8")
        st.markdown(
            f"""
            <div class="sidebar-logo-card">
                <img src="data:image/{ext};base64,{encoded_sidebar_logo}" style="width: 100%; max-width: 100%; pointer-events: none; display: block; margin: 0 auto;">
            </div>
            """,
            unsafe_allow_html=True,
        )

    b64_foto = st.session_state.get(f"foto_user_{user_email}")
    if not b64_foto:
        b64_foto = get_repo_image_b64(["perfil.jpg", "perfil.png", "perfil.jpeg", "avatar.png"])

    img_obj = base64_to_image(b64_foto)
    if img_obj is not None:
        st.image(img_obj, use_container_width=True)

    if len(user_edificios) > 0:
        tags_edif_sidebar = "".join([f"<span class='edificio-tag-badge' style='margin: 1px;'>{e}</span>" for e in user_edificios])
    else:
        tags_edif_sidebar = "<span style='font-size: 0.60rem; color: #64748b;'>Sin proyectos asignados</span>"

    st.markdown(
        f"""
        <div class="sidebar-profile-box">
            <div class="sidebar-user-nombres">{user_nombres}</div>
            <div class="sidebar-user-apellidos">{user_apellidos}</div>
            <div class="sidebar-user-email">{user_email}</div>
            <div style="margin-top: 4px; margin-bottom: 5px;">
                <div class="sidebar-user-cargo">{user_cargo}</div>
            </div>
            <div style="display: flex; flex-wrap: wrap; justify-content: center; gap: 3px; margin-top: 4px;">
                {tags_edif_sidebar}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if es_admin:
        st.markdown("<div style='text-align: center; margin-bottom: 4px; font-size: 0.65rem; color: #ffffff; font-weight: 800; background: #111827; padding: 3px; border-radius: 6px; border: 1px solid #1f2937;'>ADMINISTRADOR GENERAL</div>", unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    with st.expander("⚙️ Configuración de Cuenta", expanded=False):
        edit_nombres = st.text_input("Nombres:", value=st.session_state.usuario_nombres, key="sb_nom")
        edit_apellidos = st.text_input("Apellidos:", value=st.session_state.usuario_apellidos, key="sb_ape")
        
        idx_c = CARGOS_DISPONIBLES.index(user_cargo) if user_cargo in CARGOS_DISPONIBLES else 0
        edit_cargo = st.selectbox("Cargo:", CARGOS_DISPONIBLES, index=idx_c, key="sb_car")

        edit_edificios = st.multiselect(
            "Edificios / Proyectos Asignados:",
            options=EDIFICIOS_ALPHA,
            default=[e for e in user_edificios if e in EDIFICIOS_ALPHA],
            key="sb_edif_edit"
        )
        edit_pass = st.text_input("Nueva Contraseña:", type="password", key="sb_pass")
        edit_pass_rep = st.text_input("Repetir Contraseña:", type="password", key="sb_pass_rep")
        nueva_foto_file = st.file_uploader("Actualizar Foto de Perfil", type=["jpg", "jpeg", "png"], key="sb_foto_file")

        if st.button("Guardar Ajustes", type="primary", use_container_width=True):
            if edit_pass.strip() or edit_pass_rep.strip():
                if edit_pass != edit_pass_rep:
                    st.error("Las nuevas contraseñas no coinciden.")
                    st.stop()

            base_update_data = {
                "nombres": edit_nombres.strip(),
                "apellidos": edit_apellidos.strip(),
                "cargo": edit_cargo
            }
            if edit_pass.strip():
                base_update_data["password"] = edit_pass.strip()
            if nueva_foto_file is not None:
                b64_str = image_to_base64(nueva_foto_file)
                if b64_str:
                    base_update_data["foto_b64"] = b64_str
                    st.session_state[f"foto_user_{user_email}"] = b64_str

            try:
                try:
                    full_update = base_update_data.copy()
                    full_update["edificios"] = edit_edificios
                    supabase.table("usuarios").update(full_update).ilike("correo", user_email).execute()
                except Exception:
                    supabase.table("usuarios").update(base_update_data).ilike("correo", user_email).execute()
                    supabase.table("app_config").upsert({
                        "key": f"user_edificios_{user_email}",
                        "value": json.dumps(edit_edificios)
                    }).execute()
                
                st.session_state.usuario_nombres = edit_nombres.strip()
                st.session_state.usuario_apellidos = edit_apellidos.strip()
                st.session_state.usuario_cargo = edit_cargo
                st.session_state.usuario_edificios = edit_edificios
                st.session_state.db_loaded = False
                st.success("Configuración actualizada correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error actualizando perfil: {e}")

    st.markdown("<hr>", unsafe_allow_html=True)
    if st.button("Cerrar Sesión", use_container_width=True):
        st.session_state.autenticado = False
        st.session_state.usuario_email = ""
        if "u" in st.query_params:
            del st.query_params["u"]
        components.html(
            """
            <script>
            try {
                const win = window.top || window.parent || window;
                win.localStorage.removeItem('alpha_user_session');
                const url = new URL(win.location.href);
                url.searchParams.delete('u');
                win.location.replace(url.href);
            } catch(e) {}
            </script>
            """,
            height=0,
            width=0
        )
        st.rerun()

# ==============================================================================
# 8. SMART DASHBOARD GLASSMORPHISM Y AUTOGUARDADO (DRAFTS)
# ==============================================================================
user_nombre_completo = f"{user_nombres} {user_apellidos}".strip()

DRAFT_STORAGE_KEY = "alpha_draft_v4_" + base64.urlsafe_b64encode(user_email.encode("utf-8")).decode("ascii").rstrip("=")
components.html(
    f"""
    <script>
    (() => {{
        const win = window.top || window.parent || window;
        const KEY = '{DRAFT_STORAGE_KEY}';
        const EXCLUDE = new Set(['log_email','log_pass','log_pin','reg_pass','reg_pass_rep','reg_pin','sb_pass','sb_pass_rep']);
        let saveTimer = null;

        function keyFor(el, idx) {{
            const name = el.getAttribute('name') || '';
            const type = (el.type || '').toLowerCase();
            const aria = el.getAttribute('aria-label') || '';
            return `el|${{name}}|${{type}}|${{aria}}|${{idx}}`;
        }}

        function snapshot() {{
            const data = {{controls: {{}}}};
            const els = Array.from(win.document.querySelectorAll('input, textarea, select')).filter(el => !el.disabled && !EXCLUDE.has(el.getAttribute('name')));
            els.forEach((el, idx) => {{
                const type = (el.type || '').toLowerCase();
                if (type === 'password' || type === 'file') return;
                const k = keyFor(el, idx);
                data.controls[k] = {{
                    value: type === 'checkbox' || type === 'radio' ? !!el.checked : el.value,
                    type: type
                }};
            }});
            return data;
        }}

        function saveNow() {{
            try {{ win.localStorage.setItem(KEY, JSON.stringify(snapshot())); }} catch (e) {{}}
        }}

        function scheduleSave() {{
            clearTimeout(saveTimer);
            saveTimer = setTimeout(saveNow, 300);
        }}

        function restore() {{
            try {{
                const raw = win.localStorage.getItem(KEY);
                if (!raw) return;
                const data = JSON.parse(raw);
                if (!data || !data.controls) return;
                const els = Array.from(win.document.querySelectorAll('input, textarea, select')).filter(el => !el.disabled && !EXCLUDE.has(el.getAttribute('name')));
                els.forEach((el, idx) => {{
                    const k = keyFor(el, idx);
                    const item = data.controls[k];
                    if (!item) return;
                    const type = (el.type || '').toLowerCase();
                    if (type === 'password' || type === 'file') return;
                    if (type === 'checkbox' || type === 'radio') {{
                        if (el.checked !== !!item.value) {{
                            el.checked = !!item.value;
                            el.dispatchEvent(new Event('input', {{bubbles:true}}));
                            el.dispatchEvent(new Event('change', {{bubbles:true}}));
                        }}
                    }} else if (typeof item.value === 'string') {{
                        if (el.value !== item.value) {{
                            el.value = item.value;
                            el.dispatchEvent(new Event('input', {{bubbles:true}}));
                            el.dispatchEvent(new Event('change', {{bubbles:true}}));
                        }}
                    }}
                }});
            }} catch (e) {{}}
        }}

        function boot() {{
            restore();
            win.document.addEventListener('input', scheduleSave, true);
            win.document.addEventListener('change', scheduleSave, true);
            win.addEventListener('beforeunload', saveNow);
        }}
        
        setTimeout(boot, 800);

        win.alphaBuildersClearDraft = function() {{
            try {{ win.localStorage.removeItem(KEY); }} catch (e) {{}}
        }};
    }})();
    </script>
    """,
    height=0,
    width=0
)

local_dt = get_local_datetime_ecuador()
fecha_hoy_iso = local_dt.strftime("%Y-%m-%d")

dias_nombre_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
dia_semana_actual = dias_nombre_es[local_dt.weekday()]
mes_actual = NOMBRES_MESES[local_dt.month]
fecha_widget_texto = f"{dia_semana_actual}, {local_dt.day} de {mes_actual} de {local_dt.year}"
clima_actual_str = get_realtime_weather()

mis_chks_list = st.session_state.get("db_checklists", {}).get(user_email, [])
mis_insps_list = st.session_state.get("db_inspecciones", {}).get(user_email, [])
mis_rnds_list = st.session_state.get("db_rendimientos", {}).get(user_email, [])

chk_hoy_cumplido = any(c.get("Fecha") == fecha_hoy_iso for c in mis_chks_list)
insp_hoy_cumplida = any(i.get("Fecha") == fecha_hoy_iso for i in mis_insps_list)

tag_chk_html = '<span class="milestone-status-done">✓ Cumplido</span>' if chk_hoy_cumplido else '<span class="milestone-status-pending">⏳ Pendiente</span>'
tag_insp_html = '<span class="milestone-status-done">✓ Cumplido</span>' if insp_hoy_cumplida else '<span class="milestone-status-pending">⏳ Pendiente</span>'

if es_maestro_mayor:
    hitos_body_html = f'<div class="milestone-item"><span class="milestone-name">Libro de Obra Maestro</span>{tag_insp_html}</div>'
else:
    hitos_body_html = (
        f'<div class="milestone-item"><span class="milestone-name">Checklist</span>{tag_chk_html}</div>'
        f'<div class="milestone-item"><span class="milestone-name">Libro de Obra</span>{tag_insp_html}</div>'
    )

if len(mis_rnds_list) > 0:
    total_eficientes = sum(1 for r in mis_rnds_list if r.get("Estado") in ["EFICIENTE", "CUMPLE META"])
    porc_rendimiento = int(round((total_eficientes / len(mis_rnds_list)) * 100))
    lbl_rend_prom = f"{porc_rendimiento}% Eficaz"
else:
    porc_rendimiento = 100
    lbl_rend_prom = "100% Óptimo"

color_dona = "#10b981" if porc_rendimiento >= 75 else "#f59e0b" if porc_rendimiento >= 50 else "#ef4444"

todas_las_incidencias_db = st.session_state.get("db_incidencias_all", [])
if len(user_edificios) > 0:
    incs_mis_proyectos = [inc for inc in todas_las_incidencias_db if inc.get("Proyecto") in user_edificios]
else:
    incs_mis_proyectos = [inc for inc in todas_las_incidencias_db if inc.get("Usuario") == user_email]

incs_abiertas_count = sum(1 for inc in incs_mis_proyectos if inc.get("Estado") == "Abierta")
mi_personal_lista = st.session_state.get("db_trabajadores_por_usuario", {}).get(user_email, [])
total_personal_count = len(mi_personal_lista)

tags_edificios_html = "".join([f"<span class='edificio-tag-badge'>{ed}</span>" for ed in user_edificios])

dashboard_html = (
    '<div class="smart-dashboard-container">'
    '<div class="smart-header-bar">'
    '<div>'
    '<div class="smart-title">Alpha Builders | Portal de Obra</div>'
    f'<div class="smart-user-sub">{user_nombre_completo} &bull; <b style="color: #cbd5e1 !important;">{user_cargo}</b> {tags_edificios_html}</div>'
    '</div>'
    '<div style="display: flex; gap: 6px; flex-wrap: wrap;">'
    f'<div class="smart-pill">📅 {fecha_widget_texto}</div>'
    f'<div class="smart-pill">{clima_actual_str}</div>'
    '</div>'
    '</div>'
    '<div class="widgets-grid">'
    '<div class="widget-glass-card">'
    '<div class="w-card-title"><span>🎯 Hitos Diarios</span><span>Hoy</span></div>'
    f'{hitos_body_html}'
    '</div>'
    '<div class="widget-glass-card">'
    '<div class="w-card-title"><span>⚡ Rendimiento</span><span>Promedio</span></div>'
    '<div class="donut-container">'
    f'<svg class="donut-chart-svg" viewBox="0 0 36 36">'
    f'<circle class="donut-bg" cx="18" cy="18" r="15.9155" />'
    f'<circle class="donut-progress" cx="18" cy="18" r="15.9155" stroke="{color_dona}" stroke-dasharray="{porc_rendimiento}, 100" />'
    f'</svg>'
    f'<div><div class="donut-info-val" style="color: {color_dona} !important;">{porc_rendimiento}%</div>'
    f'<div class="donut-info-lbl">{lbl_rend_prom}</div></div>'
    '</div>'
    '</div>'
    '<div class="widget-glass-card">'
    '<div class="w-card-title"><span>🚨 Incidencias</span><span>Abiertas</span></div>'
    f'<div class="stat-hero-number" style="color: #f87171 !important;">{incs_abiertas_count}</div>'
    '<div style="font-size: 0.65rem; color: #94a3b8 !important; font-weight: 600;">En tus proyectos</div>'
    '</div>'
    '<div class="widget-glass-card">'
    '<div class="w-card-title"><span>👷 Personal</span><span>Activos</span></div>'
    f'<div class="stat-hero-number" style="color: #60a5fa !important;">{total_personal_count}</div>'
    '<div style="font-size: 0.65rem; color: #94a3b8 !important; font-weight: 600;">Personal a tu cargo</div>'
    '</div>'
    '</div>'
    '</div>'
)

st.markdown(dashboard_html, unsafe_allow_html=True)
st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)

# Pestañas adaptadas: para Maestro Mayor no se muestra Levantamiento de Incidencias
if es_maestro_mayor:
    pestanas = [
        "Libro de Obra Maestro",
        "Personal a Cargo",
        "Control de Rendimiento",
        "Colaborativo"
    ]
else:
    pestanas = [
        "Checklist", 
        "Libro de Obra", 
        "Personal a Cargo",
        "Levantamiento de Incidencias", 
        "Control de Rendimiento",
        "Colaborativo"
    ]

if es_admin:
    pestanas.append("Panel Admin")

tabs_app = st.tabs(pestanas)
# ==============================================================================
# PARTE 4 DE 5: MÓDULOS DE CONTROL SEGÚN ROL CON SOPORTE COMPLETO A VERSIONES ANTIGUAS
# ==============================================================================

# ==============================================================================
# 9. ENRUTAMIENTO DINÁMICO DE PESTAÑAS SEGÚN EL CARGO
# ==============================================================================
if es_maestro_mayor:
    tab_libro_maestro = tabs_app[0]
    tab_personal = tabs_app[1]
    tab_rend = tabs_app[2]
    tab_colab = tabs_app[3]
else:
    tab_chk = tabs_app[0]
    tab_libro = tabs_app[1]
    tab_personal = tabs_app[2]
    tab_incidencias = tabs_app[3]
    tab_rend = tabs_app[4]
    tab_colab = tabs_app[5]

# ------------------------------------------------------------------------------
# 9.A. MÓDULO EXCLUSIVO PARA MAESTRO MAYOR: LIBRO DE OBRA MAESTRO
# ------------------------------------------------------------------------------
if es_maestro_mayor:
    with tab_libro_maestro:
        st.markdown("### Libro de Obra – Maestro Mayor")
        st.caption("Registro diario de actividades, cuadrillas a cargo y metrajes ejecutados en obra.")

        if "llenando_libro_mm" not in st.session_state:
            st.session_state.llenando_libro_mm = False

        if "edit_mm_id" not in st.session_state:
            st.session_state.edit_mm_id = None

        if "filas_maestro_act" not in st.session_state:
            st.session_state.filas_maestro_act = [
                {"id": 1, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}
            ]

        # Botón inicial para abrir el formulario
        if not st.session_state.llenando_libro_mm and not st.session_state.edit_mm_id:
            if st.button("➕ Llenar Libro de Obra", type="primary", key="btn_open_llenar_mm"):
                st.session_state.llenando_libro_mm = True
                st.session_state.edit_mm_id = None
                st.session_state.filas_maestro_act = [{"id": 1, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}]
                for k in ["mm_edit_fecha_val", "mm_edit_edif_val"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        # Despliegue del formulario únicamente si está activo
        if st.session_state.llenando_libro_mm or st.session_state.edit_mm_id:
            st.markdown("---")
            if st.session_state.edit_mm_id:
                st.info("✏️ **Modo Edición Activo:** Modificando reporte seleccionado.")

            local_today_mm = get_local_datetime_ecuador().date()
            fecha_default_mm = st.session_state.get("mm_edit_fecha_val", local_today_mm)
            es_edicion_mm = bool(st.session_state.edit_mm_id)

            col_m_cfg1, col_m_cfg2, col_m_cfg3 = st.columns([1.5, 2, 1.5])
            with col_m_cfg1:
                fecha_maestro_val = st.date_input(
                    "Fecha de Trabajo:*", 
                    fecha_default_mm, 
                    disabled=es_edicion_mm, 
                    key=f"mm_fecha_input_{st.session_state.get('edit_mm_id', 'new')}"
                )
                if es_edicion_mm:
                    st.caption("🔒 *La fecha no se puede modificar durante la edición.*")
            with col_m_cfg2:
                proyectos_maestro_disp = user_edificios if len(user_edificios) > 0 else EDIFICIOS_ALPHA
                idx_edif_mm = 0
                if "mm_edit_edif_val" in st.session_state and st.session_state.mm_edit_edif_val in proyectos_maestro_disp:
                    idx_edif_mm = proyectos_maestro_disp.index(st.session_state.mm_edit_edif_val) + 1

                edificio_maestro_val = st.selectbox(
                    "Proyecto / Edificio Asignado:*",
                    ["-- Seleccione un Proyecto --"] + proyectos_maestro_disp,
                    index=idx_edif_mm,
                    key="mm_edificio_input"
                )
            with col_m_cfg3:
                st.text_input("Maestro Responsable:", value=user_nombre_completo, disabled=True)

            st.markdown("---")
            st.markdown(f"#### 🔨 Actividades Realizadas, Personal y Metrajes ({len(st.session_state.filas_maestro_act)} registros)")

            mi_personal_propio = st.session_state.get("db_trabajadores_por_usuario", {}).get(user_email, [])
            if edificio_maestro_val != "-- Seleccione un Proyecto --":
                personal_filtrado_edif = [
                    f"{t['nombre']} ({t['cargo']})" for t in mi_personal_propio 
                    if (t.get("edificio") == edificio_maestro_val or t.get("edificio") == "General" or not t.get("edificio"))
                ]
                if not personal_filtrado_edif:
                    personal_filtrado_edif = [f"{t['nombre']} ({t['cargo']})" for t in mi_personal_propio]
            else:
                personal_filtrado_edif = [f"{t['nombre']} ({t['cargo']})" for t in mi_personal_propio]

            indices_del_m = []
            payload_actividades_m = []

            for idx_m, f_data in enumerate(st.session_state.filas_maestro_act, 1):
                f_id = f_data["id"]
                st.markdown(f"""<div class="banner-item-header"><span>Actividad N° {idx_m}</span></div>""", unsafe_allow_html=True)
                st.markdown('<div class="card-item-body-compact">', unsafe_allow_html=True)
                
                c_m1, c_m2 = st.columns([2.5, 1.5])
                with c_m1:
                    act_m_txt = st.text_input(
                        f"Descripción de la Actividad {idx_m}:",
                        value=f_data.get("actividad", ""),
                        placeholder="Ej. Enlucido paleteado en muros de fachada...",
                        key=f"mm_act_txt_{f_id}"
                    )
                with c_m2:
                    cant_m_txt = st.text_input(
                        f"Cantidad / Avance {idx_m}:",
                        value=str(f_data.get("cantidad", "")),
                        placeholder="Ej. 15 m2, 3 puertas, 2.5 tramos...",
                        key=f"mm_cant_txt_{f_id}"
                    )

                c_m3, c_m4, c_m5 = st.columns([2.2, 2.2, 0.4])
                with c_m3:
                    if personal_filtrado_edif:
                        pers_sel_m = st.multiselect(
                            f"Personal a Cargo {idx_m}:",
                            options=personal_filtrado_edif,
                            default=[p for p in f_data.get("personal_a_cargo", []) if p in personal_filtrado_edif],
                            placeholder="Seleccionar personal asignado...",
                            key=f"mm_pers_sel_{f_id}"
                        )
                    else:
                        pers_sel_txt = st.text_input(
                            f"Personal a Cargo manual {idx_m}:",
                            value=", ".join(f_data.get("personal_a_cargo", [])) if isinstance(f_data.get("personal_a_cargo"), list) else str(f_data.get("personal_a_cargo", "")),
                            placeholder="Escriba los nombres de los trabajadores...",
                            key=f"mm_pers_manual_{f_id}"
                        )
                        pers_sel_m = [p.strip() for p in pers_sel_txt.split(",") if p.strip()]

                with c_m4:
                    obs_m_txt = st.text_input(
                        f"Observaciones / Frente {idx_m}:",
                        value=f_data.get("observaciones", ""),
                        placeholder="Ej. Piso 2 departamento 201...",
                        key=f"mm_obs_txt_{f_id}"
                    )

                with c_m5:
                    st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)
                    if st.button("🗑️", key=f"btn_del_mm_row_{f_id}", help="Eliminar fila"):
                        indices_del_m.append(idx_m - 1)

                st.markdown('</div>', unsafe_allow_html=True)
                payload_actividades_m.append({
                    "Actividad": act_m_txt.strip(),
                    "Cantidad": cant_m_txt.strip(),
                    "Personal_A_Cargo": pers_sel_m,
                    "Observaciones": obs_m_txt.strip()
                })

            if indices_del_m:
                for del_i in sorted(indices_del_m, reverse=True):
                    if len(st.session_state.filas_maestro_act) > 1:
                        st.session_state.filas_maestro_act.pop(del_i)
                    else:
                        st.session_state.filas_maestro_act = [{"id": int(datetime.datetime.now().timestamp() * 1000), "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}]
                st.rerun()

            if st.button("➕ Agregar Otra Actividad", key="btn_add_mm_act_row"):
                next_id_mm = (max([x["id"] for x in st.session_state.filas_maestro_act]) + 1) if st.session_state.filas_maestro_act else 1
                st.session_state.filas_maestro_act.append({"id": next_id_mm, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""})
                st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            lbl_btn_mm = "🔄 Actualizar Reporte de Maestro Mayor" if st.session_state.edit_mm_id else "💾 Guardar Reporte de Maestro Mayor"
            if st.button(lbl_btn_mm, type="primary", use_container_width=True):
                if edificio_maestro_val == "-- Seleccione un Proyecto --":
                    st.error("⚠️ Debe seleccionar el Edificio o Proyecto donde realizó los trabajos.")
                else:
                    actividades_validas = [a for a in payload_actividades_m if a["Actividad"]]
                    if not actividades_validas:
                        st.error("⚠️ Ingrese al menos una actividad antes de guardar.")
                    else:
                        dias_es_m = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                        dia_str_m = dias_es_m[fecha_maestro_val.weekday()]

                        payload_final_mm = {
                            "Tipo_Registro": "Libro_Obra_Maestro",
                            "Actividades_Maestro": actividades_validas,
                            "Edificio": edificio_maestro_val,
                            "Maestro": user_nombre_completo,
                            "Fecha": fecha_maestro_val.strftime("%Y-%m-%d")
                        }

                        record_data_mm = {
                            "usuario_email": user_email,
                            "proyecto": edificio_maestro_val,
                            "fecha": fecha_maestro_val.strftime("%Y-%m-%d"),
                            "dia": dia_str_m,
                            "residente": user_nombre_completo,
                            "frente": "Reporte de Maestro Mayor",
                            "clima": "N/A",
                            "hora_inicio": "07:00",
                            "hora_fin": "16:00",
                            "datos": payload_final_mm
                        }

                        try:
                            if st.session_state.edit_mm_id:
                                supabase.table("inspecciones").update(record_data_mm).eq("id", st.session_state.edit_mm_id).execute()
                                st.success(f"¡Reporte actualizado exitosamente para **{edificio_maestro_val}**!")
                            else:
                                supabase.table("inspecciones").insert(record_data_mm).execute()
                                st.success(f"¡Reporte del Maestro guardado exitosamente para **{edificio_maestro_val}**!")

                            components.html("""<script>try { if (window.top.alphaBuildersClearDraft) window.top.alphaBuildersClearDraft(); } catch(e) {}</script>""", height=0, width=0)
                            st.session_state.db_loaded = False
                            st.session_state.llenando_libro_mm = False
                            st.session_state.edit_mm_id = None
                            st.session_state.filas_maestro_act = [{"id": 1, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}]
                            for k in ["mm_edit_fecha_val", "mm_edit_edif_val"]:
                                if k in st.session_state:
                                    del st.session_state[k]
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al procesar reporte: {e}")

            lbl_cancel_mm = "❌ Cancelar Edición" if st.session_state.edit_mm_id else "❌ Cancelar Llenado"
            if st.button(lbl_cancel_mm, key="btn_cancel_mm_bottom", use_container_width=True):
                st.session_state.llenando_libro_mm = False
                st.session_state.edit_mm_id = None
                st.session_state.filas_maestro_act = [{"id": 1, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}]
                for k in ["mm_edit_fecha_val", "mm_edit_edif_val"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        st.markdown("---")
        st.markdown("### Historial de Reportes del Maestro Mayor")
        mis_libros_m = st.session_state.get("db_inspecciones", {}).get(user_email, [])

        if len(mis_libros_m) > 0:
            df_libros_m = pd.DataFrame(mis_libros_m)
            df_libros_m["fecha_dt"] = pd.to_datetime(df_libros_m["Fecha"])
            df_libros_m = df_libros_m.sort_values(by="fecha_dt", ascending=False)
            df_libros_m["año"] = df_libros_m["fecha_dt"].dt.year
            df_libros_m["mes_num"] = df_libros_m["fecha_dt"].dt.month

            grupos_m = df_libros_m.groupby(["año", "mes_num"], sort=False)

            for (anio, mes_num), items_mes in grupos_m:
                nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Reportes)"
                with st.expander(nombre_mes_str, expanded=False):
                    for idx_insp_m, insp_m in items_mes.iterrows():
                        insp_dict_m = insp_m.to_dict()
                        insp_db_id_m = insp_dict_m.get("db_id")

                        with st.expander(f"📌 [{insp_dict_m.get('Proyecto')}] {insp_dict_m.get('Fecha')} ({insp_dict_m.get('Dia')})", expanded=False):
                            raw_dm = insp_dict_m.get("Datos", {})
                            d_parsed = raw_dm if isinstance(raw_dm, dict) else json.loads(raw_dm or "{}") if isinstance(raw_dm, str) else {}
                            
                            # Compatible con versiones antiguas y nuevas
                            if isinstance(d_parsed, dict):
                                acts_guardadas = d_parsed.get("Actividades_Maestro", [])
                            elif isinstance(d_parsed, list):
                                acts_guardadas = d_parsed
                            else:
                                acts_guardadas = []
                            
                            for a_g in acts_guardadas:
                                if isinstance(a_g, dict):
                                    pers_raw = a_g.get("Personal_A_Cargo", a_g.get("personal_a_cargo", []))
                                    pers_str = ", ".join(pers_raw) if isinstance(pers_raw, list) else str(pers_raw or "")
                                    pers_tag = f" | 👷 **Personal:** {pers_str}" if pers_str else ""
                                    act_n = a_g.get('Actividad', a_g.get('actividad', ''))
                                    act_c = a_g.get('Cantidad', a_g.get('cantidad', ''))
                                    act_o = a_g.get('Observaciones', a_g.get('observaciones', 'Sin observaciones'))
                                    st.write(f"• **{act_n}**: `{act_c}`{pers_tag} — *{act_o}*")

                            c_dl_m1, c_dl_m2, c_ed_m, c_del_m = st.columns([2, 2, 1, 1])
                            with c_dl_m1:
                                with st.popover("📊 Exportar Excel", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.xlsx)",
                                        get_cached_libro_maestro_excel(safe_json_dumps(insp_dict_m)),
                                        file_name=f"Libro_Maestro_{insp_dict_m['Proyecto']}_{insp_dict_m['Fecha']}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"dl_mm_xlsx_{idx_insp_m}",
                                        use_container_width=True
                                    )
                            with c_dl_m2:
                                with st.popover("📄 Exportar PDF", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.pdf)",
                                        get_cached_libro_maestro_pdf(safe_json_dumps(insp_dict_m)),
                                        file_name=f"Libro_Maestro_{insp_dict_m['Proyecto']}_{insp_dict_m['Fecha']}.pdf",
                                        mime="application/pdf",
                                        key=f"dl_mm_pdf_{idx_insp_m}",
                                        use_container_width=True
                                    )
                            with c_ed_m:
                                if st.button("✏️", key=f"btn_edit_mm_{idx_insp_m}_{insp_db_id_m}", help="Editar reporte", use_container_width=True):
                                    st.session_state.edit_mm_id = insp_db_id_m
                                    st.session_state.llenando_libro_mm = True
                                    st.session_state.mm_edit_fecha_val = pd.to_datetime(insp_dict_m.get("Fecha")).date()
                                    st.session_state.mm_edit_edif_val = insp_dict_m.get("Proyecto")
                                    raw_dm2 = insp_dict_m.get("Datos", {})
                                    d_parsed2 = raw_dm2 if isinstance(raw_dm2, dict) else json.loads(raw_dm2 or "{}") if isinstance(raw_dm2, str) else {}
                                    
                                    if isinstance(d_parsed2, dict):
                                        acts_rec = d_parsed2.get("Actividades_Maestro", [])
                                    elif isinstance(d_parsed2, list):
                                        acts_rec = d_parsed2
                                    else:
                                        acts_rec = []

                                    st.session_state.filas_maestro_act = [
                                        {
                                            "id": i + 1,
                                            "actividad": it.get("Actividad", it.get("actividad", "")),
                                            "cantidad": it.get("Cantidad", it.get("cantidad", "")),
                                            "personal_a_cargo": it.get("Personal_A_Cargo", it.get("personal_a_cargo", [])),
                                            "observaciones": it.get("Observaciones", it.get("observaciones", ""))
                                        } for i, it in enumerate(acts_rec)
                                    ] if acts_rec else [{"id": 1, "actividad": "", "cantidad": "", "personal_a_cargo": [], "observaciones": ""}]
                                    st.rerun()

                            with c_del_m:
                                if st.button("🗑️", key=f"btn_del_mm_{idx_insp_m}_{insp_db_id_m}", help="Eliminar reporte", use_container_width=True):
                                    try:
                                        supabase.table("inspecciones").delete().eq("id", insp_db_id_m).execute()
                                        st.session_state.db_loaded = False
                                        st.success("Reporte eliminado.")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error al eliminar: {e}")
        else:
            st.info("Aún no tienes reportes guardados como Maestro Mayor.")

# ------------------------------------------------------------------------------
# 9.B. MÓDULOS PARA RESIDENTE Y ASISTENTE: CHECKLIST Y LIBRO DE OBRA OFICIAL
# ------------------------------------------------------------------------------
else:
    # 1. CHECKLIST (CONTROL DIARIO DE OBRA)
    with tab_chk:
        if "creando_jornada" not in st.session_state:
            st.session_state.creando_jornada = False

        if "edit_chk_id" not in st.session_state:
            st.session_state.edit_chk_id = None

        if "filas_supervision" not in st.session_state:
            st.session_state.filas_supervision = [{"id": 1, "actividad": ""}]

        if "chk_obs_counts" not in st.session_state:
            st.session_state.chk_obs_counts = {}

        if "chk_responsables_extra" not in st.session_state:
            st.session_state.chk_responsables_extra = {}

        st.markdown("### Checklist – Control de Obra")
        st.caption("Supervisión técnica y control de ejecución diaria en obra.")

        if not st.session_state.creando_jornada and not st.session_state.edit_chk_id:
            if st.button("➕ Crear Nueva Jornada de Inspección", type="primary"):
                st.session_state.creando_jornada = True
                st.session_state.edit_chk_id = None
                st.session_state.filas_supervision = [{"id": 1, "actividad": ""}]
                st.session_state.chk_obs_counts = {}
                st.session_state.chk_responsables_extra = {}
                for k in ["chk_edit_edif_val", "chk_edit_fecha_val", "chk_edit_hini_val", "chk_edit_hfin_val", "chk_edit_resp_map"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        if st.session_state.creando_jornada or st.session_state.edit_chk_id:
            st.markdown("---")
            if st.session_state.edit_chk_id:
                st.info("✏️ **Modo Edición de Checklist Activo**")

            es_edicion_chk = bool(st.session_state.edit_chk_id)

            with st.container():
                st.markdown("#### Configuración de la Jornada")
                cfg_c1, cfg_c2, cfg_c3, cfg_c4, cfg_c5 = st.columns([2, 2, 2, 1.5, 1.5])
                with cfg_c1:
                    proyectos_disponibles = user_edificios if len(user_edificios) > 0 else EDIFICIOS_ALPHA
                    idx_edif_chk = 0
                    if "chk_edit_edif_val" in st.session_state and st.session_state.chk_edit_edif_val in proyectos_disponibles:
                        idx_edif_chk = proyectos_disponibles.index(st.session_state.chk_edit_edif_val) + 1

                    edificio_val = st.selectbox("Edificio / Proyecto:", ["-- Seleccione --"] + proyectos_disponibles, index=idx_edif_chk, key="sel_edificio")
                with cfg_c2:
                    st.text_input("Responsable:", value=user_nombre_completo, disabled=True)
                with cfg_c3:
                    local_now_chk = get_local_datetime_ecuador().date()
                    fecha_def_chk = st.session_state.get("chk_edit_fecha_val", local_now_chk)
                    fecha_val = st.date_input(
                        "Fecha:", 
                        fecha_def_chk, 
                        disabled=es_edicion_chk, 
                        key=f"sel_fecha_{st.session_state.get('edit_chk_id', 'new')}"
                    )
                    if es_edicion_chk:
                        st.caption("🔒 *Fecha bloqueada en edición.*")
                with cfg_c4:
                    h_ini_def = st.session_state.get("chk_edit_hini_val", datetime.time(7, 0))
                    hora_inicio_val = st.time_input("Hora Inicio:", h_ini_def, key="sel_hora_inicio")
                with cfg_c5:
                    h_fin_def = st.session_state.get("chk_edit_hfin_val", datetime.time(16, 0))
                    hora_fin_val = st.time_input("Hora Fin:", h_fin_def, key="sel_hora_fin")

                st.markdown("---")

                edit_resp_map = st.session_state.get("chk_edit_resp_map", {})

                def buscar_item_retrocompatible(jornada, idx, act_texto):
                    k1 = f"{jornada}_{act_texto}".lower().strip()
                    if k1 in edit_resp_map:
                        return edit_resp_map[k1]
                    k2 = f"{jornada}_{idx}".lower().strip()
                    if k2 in edit_resp_map:
                        return edit_resp_map[k2]
                    for k_map, v_map in edit_resp_map.items():
                        if k_map.startswith(jornada.lower()) and (act_texto.lower()[:8] in k_map or k_map.split("_")[-1][:8] in act_texto.lower()):
                            return v_map
                    return {}

                # JORNADA DE LA MAÑANA
                st.markdown("#### 🌅 Jornada de la Mañana")
                resp_manana = []

                for idx, act in enumerate(ACTIVIDADES_MANANA_CLEAN, 1):
                    item_key = f"m_{idx}"
                    item_guardado = buscar_item_retrocompatible("Mañana", idx, act)
                    
                    default_estado_m = item_guardado.get("Estado")
                    if default_estado_m not in ["✓ Cumple", "✗ No Cumple"]:
                        default_estado_m = None

                    obs_guardadas_m = item_guardado.get("Observaciones", [])
                    if isinstance(obs_guardadas_m, str):
                        obs_guardadas_m = [obs_guardadas_m] if obs_guardadas_m.strip() else []

                    st.markdown(f"""<div class="banner-item-header"><span>N° {idx} — {act}</span></div>""", unsafe_allow_html=True)
                    st.markdown('<div class="card-item-body-compact">', unsafe_allow_html=True)
                    c_col1, c_col2, c_col3 = st.columns([1.1, 2.3, 1.6])

                    with c_col1:
                        st.markdown("<small style='font-weight:700; color:#334155;'>Estado General:</small>", unsafe_allow_html=True)
                        est = st.segmented_control(
                            f"M_Est_{idx}", 
                            ["✓ Cumple", "✗ No Cumple"], 
                            default=default_estado_m, 
                            key=f"m_st_{idx}_{st.session_state.get('edit_chk_id', 'new')}", 
                            label_visibility="collapsed"
                        )

                    with c_col2:
                        st.markdown("<small style='font-weight:700; color:#334155;'>Observaciones del Ítem:</small>", unsafe_allow_html=True)
                        obs_vals_item = []
                        total_obs_count = max(st.session_state.chk_obs_counts.get(item_key, 1), len(obs_guardadas_m))
                        st.session_state.chk_obs_counts[item_key] = max(total_obs_count, 1)

                        for sub_o in range(1, st.session_state.chk_obs_counts[item_key] + 1):
                            val_o_prev = obs_guardadas_m[sub_o - 1] if sub_o <= len(obs_guardadas_m) else ""
                            o_txt = st.text_input(
                                f"Obs {idx}_{sub_o}", 
                                value=val_o_prev, 
                                key=f"m_ob_{idx}_{sub_o}_{st.session_state.get('edit_chk_id', 'new')}", 
                                placeholder=f"Observación {sub_o}...", 
                                label_visibility="collapsed"
                            )
                            if o_txt.strip():
                                obs_vals_item.append(o_txt.strip())
                        
                        c_btn_o, _ = st.columns([2, 2])
                        with c_btn_o:
                            if st.button("➕ Observación", key=f"btn_add_obs_m_{idx}", use_container_width=True):
                                st.session_state.chk_obs_counts[item_key] += 1
                                st.rerun()

                    with c_col3:
                        if idx != 1:
                            st.markdown("<small style='font-weight:700; color:#334155;'>Fotos Evidencia (múltiples):</small>", unsafe_allow_html=True)
                            uploaded_files = st.file_uploader(
                                f"Fotos M_{idx}",
                                type=["jpg", "jpeg", "png"],
                                accept_multiple_files=True,
                                key=f"m_ft_{idx}_{st.session_state.get('edit_chk_id', 'new')}",
                                label_visibility="collapsed"
                            )
                            fotos_b64 = []
                            if uploaded_files:
                                for f in uploaded_files:
                                    b64 = image_to_base64(f)
                                    if b64:
                                        fotos_b64.append(b64)
                            else:
                                fotos_guardadas = item_guardado.get("Fotos", [])
                                if isinstance(fotos_guardadas, list):
                                    fotos_b64 = fotos_guardadas
                                elif item_guardado.get("Foto_B64"):
                                    fotos_b64 = [item_guardado.get("Foto_B64")]
                        else:
                            st.markdown("<small style='font-weight:700; color:#94a3b8;'>Fotos Evidencia:</small>", unsafe_allow_html=True)
                            st.caption("📷 *No requerida*")
                            fotos_b64 = []

                    responsables_list = []
                    if act == "Coordinación con otras especialidades" and est == "✗ No Cumple":
                        st.markdown("---")
                        st.markdown("##### Datos adicionales para No Cumple")
                        
                        resp_guardados = item_guardado.get("Responsables", [])
                        if not resp_guardados:
                            resp_guardados = []
                        
                        extra_key = f"extra_{idx}_{st.session_state.get('edit_chk_id', 'new')}"
                        if extra_key not in st.session_state.chk_responsables_extra:
                            if resp_guardados:
                                st.session_state.chk_responsables_extra[extra_key] = resp_guardados.copy()
                            else:
                                st.session_state.chk_responsables_extra[extra_key] = [{"nombre": "", "area": ""}]
                        
                        resp_list = st.session_state.chk_responsables_extra[extra_key]
                        indices_a_eliminar_resp = []
                        
                        for r_idx, resp_item in enumerate(resp_list):
                            st.markdown(f"**Responsable #{r_idx + 1}**")
                            col_r1, col_r2, col_r3 = st.columns([2.5, 2.5, 0.6])
                            with col_r1:
                                nombre_resp = st.text_input(
                                    f"Nombre {r_idx + 1}",
                                    value=resp_item.get("nombre", ""),
                                    placeholder="Nombre del responsable...",
                                    key=f"resp_nom_{idx}_{r_idx}_{st.session_state.get('edit_chk_id', 'new')}",
                                    label_visibility="collapsed"
                                )
                            with col_r2:
                                area_resp = st.text_input(
                                    f"Área {r_idx + 1}",
                                    value=resp_item.get("area", ""),
                                    placeholder="Área de trabajo...",
                                    key=f"resp_area_{idx}_{r_idx}_{st.session_state.get('edit_chk_id', 'new')}",
                                    label_visibility="collapsed"
                                )
                            with col_r3:
                                if len(resp_list) > 1:
                                    if st.button("🗑️", key=f"del_resp_{idx}_{r_idx}_{st.session_state.get('edit_chk_id', 'new')}", help="Eliminar responsable"):
                                        indices_a_eliminar_resp.append(r_idx)
                            
                            resp_list[r_idx] = {"nombre": nombre_resp, "area": area_resp}
                        
                        if indices_a_eliminar_resp:
                            for r_idx in sorted(indices_a_eliminar_resp, reverse=True):
                                if len(resp_list) > 1:
                                    resp_list.pop(r_idx)
                            st.rerun()
                        
                        if st.button("➕ Agregar Responsable", key=f"add_resp_{idx}_{st.session_state.get('edit_chk_id', 'new')}", use_container_width=True):
                            resp_list.append({"nombre": "", "area": ""})
                            st.rerun()
                        
                        responsables_list = resp_list
                        st.session_state.chk_responsables_extra[extra_key] = resp_list

                    st.markdown('</div>', unsafe_allow_html=True)
                    resp_manana.append({
                        "Jornada": "Mañana",
                        "N°": idx,
                        "Actividad": act,
                        "Estado": est,
                        "Observaciones": obs_vals_item,
                        "Fotos": fotos_b64,
                        "Responsables": responsables_list if responsables_list else []
                    })

                st.markdown("<br>", unsafe_allow_html=True)

                # JORNADA DE LA TARDE
                st.markdown("#### 🌆 Jornada de la Tarde")
                resp_tarde = []

                for idx, act in enumerate(ACTIVIDADES_TARDE_CLEAN, 1):
                    item_key = f"t_{idx}"
                    item_guardado_t = buscar_item_retrocompatible("Tarde", idx, act)

                    default_estado_t = item_guardado_t.get("Estado")
                    if default_estado_t not in ["✓ Cumple", "✗ No Cumple"]:
                        default_estado_t = None

                    obs_guardadas_t = item_guardado_t.get("Observaciones", [])
                    if isinstance(obs_guardadas_t, str):
                        obs_guardadas_t = [obs_guardadas_t] if obs_guardadas_t.strip() else []

                    st.markdown(f"""<div class="banner-item-header"><span>N° {idx} — {act}</span></div>""", unsafe_allow_html=True)
                    st.markdown('<div class="card-item-body-compact">', unsafe_allow_html=True)
                    c_col1, c_col2, c_col3 = st.columns([1.1, 2.3, 1.6])

                    with c_col1:
                        st.markdown("<small style='font-weight:700; color:#334155;'>Estado General:</small>", unsafe_allow_html=True)
                        est = st.segmented_control(
                            f"T_Est_{idx}", 
                            ["✓ Cumple", "✗ No Cumple"], 
                            default=default_estado_t, 
                            key=f"t_st_{idx}_{st.session_state.get('edit_chk_id', 'new')}", 
                            label_visibility="collapsed"
                        )

                    with c_col2:
                        st.markdown("<small style='font-weight:700; color:#334155;'>Observaciones del Ítem:</small>", unsafe_allow_html=True)
                        obs_vals_item = []
                        total_obs_count_t = max(st.session_state.chk_obs_counts.get(item_key, 1), len(obs_guardadas_t))
                        st.session_state.chk_obs_counts[item_key] = max(total_obs_count_t, 1)

                        for sub_o in range(1, st.session_state.chk_obs_counts[item_key] + 1):
                            val_o_prev_t = obs_guardadas_t[sub_o - 1] if sub_o <= len(obs_guardadas_t) else ""
                            o_txt = st.text_input(
                                f"Obs T_{idx}_{sub_o}", 
                                value=val_o_prev_t, 
                                key=f"t_ob_{idx}_{sub_o}_{st.session_state.get('edit_chk_id', 'new')}", 
                                placeholder=f"Observación {sub_o}...", 
                                label_visibility="collapsed"
                            )
                            if o_txt.strip():
                                obs_vals_item.append(o_txt.strip())
                        
                        c_btn_o, _ = st.columns([2, 2])
                        with c_btn_o:
                            if st.button("➕ Observación", key=f"btn_add_obs_t_{idx}", use_container_width=True):
                                st.session_state.chk_obs_counts[item_key] += 1
                                st.rerun()

                    with c_col3:
                        st.markdown("<small style='font-weight:700; color:#334155;'>Fotos Evidencia (múltiples):</small>", unsafe_allow_html=True)
                        uploaded_files_t = st.file_uploader(
                            f"Fotos T_{idx}",
                            type=["jpg", "jpeg", "png"],
                            accept_multiple_files=True,
                            key=f"t_ft_{idx}_{st.session_state.get('edit_chk_id', 'new')}",
                            label_visibility="collapsed"
                        )
                        fotos_b64_t = []
                        if uploaded_files_t:
                            for f in uploaded_files_t:
                                b64 = image_to_base64(f)
                                if b64:
                                    fotos_b64_t.append(b64)
                        else:
                            fotos_guardadas_t = item_guardado_t.get("Fotos", [])
                            if isinstance(fotos_guardadas_t, list):
                                fotos_b64_t = fotos_guardadas_t
                            elif item_guardado_t.get("Foto_B64"):
                                fotos_b64_t = [item_guardado_t.get("Foto_B64")]

                    st.markdown('</div>', unsafe_allow_html=True)
                    resp_tarde.append({
                        "Jornada": "Tarde",
                        "N°": idx,
                        "Actividad": act,
                        "Estado": est,
                        "Observaciones": obs_vals_item,
                        "Fotos": fotos_b64_t
                    })

                st.markdown("<br>", unsafe_allow_html=True)

                # SUPERVISIÓN DE TRABAJOS (TABLA ÁGIL DE ACTIVIDADES)
                st.markdown("#### 📋 Supervisión de la Ejecución de Trabajos")
                st.caption("Ingrese la lista de actividades supervisadas durante la jornada:")

                indices_a_eliminar = []
                supervision_payload_data = []

                col_h_num, col_h_act, col_h_del = st.columns([0.6, 6.8, 0.6])
                with col_h_num:
                    st.markdown("<small style='font-weight:800; color:var(--subtext);'>N°</small>", unsafe_allow_html=True)
                with col_h_act:
                    st.markdown("<small style='font-weight:800; color:var(--subtext);'>Descripción de la Actividad</small>", unsafe_allow_html=True)
                with col_h_del:
                    st.markdown("<small style='font-weight:800; color:var(--subtext);'>Acción</small>", unsafe_allow_html=True)

                for idx_f, f_data in enumerate(st.session_state.filas_supervision, 1):
                    f_id = f_data["id"]
                    c_num, c_act, c_del = st.columns([0.6, 6.8, 0.6])

                    with c_num:
                        st.markdown(f"<div style='padding-top:8px; font-weight:700; text-align:center;'>{idx_f}</div>", unsafe_allow_html=True)

                    with c_act:
                        act_val = st.text_input(
                            f"Actividad {idx_f}:",
                            value=f_data.get("actividad", ""),
                            placeholder=f"Ej. Enlucido paleteado en muros / Pintura en departamento {idx_f}01...",
                            key=f"dyn_act_{f_id}",
                            label_visibility="collapsed"
                        )

                    with c_del:
                        if st.button("🗑️", key=f"btn_del_row_{f_id}", help="Eliminar actividad", use_container_width=True):
                            indices_a_eliminar.append(idx_f - 1)

                    supervision_payload_data.append({
                        "N°": idx_f,
                        "Actividad": act_val.strip(),
                        "Encargados": "",
                        "Observaciones": "",
                        "Fotos": []
                    })

                if indices_a_eliminar:
                    for del_i in sorted(indices_a_eliminar, reverse=True):
                        if len(st.session_state.filas_supervision) > 1:
                            st.session_state.filas_supervision.pop(del_i)
                        else:
                            st.session_state.filas_supervision = [{"id": int(datetime.datetime.now().timestamp() * 1000), "actividad": ""}]
                    st.rerun()

                if st.button("➕ Agregar Actividad", key="btn_add_dyn_supervision_row"):
                    next_id_sup = (max([x["id"] for x in st.session_state.filas_supervision]) + 1) if st.session_state.filas_supervision else 1
                    st.session_state.filas_supervision.append({"id": next_id_sup, "actividad": ""})
                    st.rerun()

                st.markdown("<br>", unsafe_allow_html=True)
                lbl_btn_chk = "🔄 Actualizar Checklist de Obra" if st.session_state.edit_chk_id else "💾 Guardar Checklist de Obra"
                if st.button(lbl_btn_chk, type="primary", use_container_width=True):
                    if edificio_val == "-- Seleccione --" or not edificio_val:
                        st.error("⚠️ Por favor seleccione un Edificio o Proyecto válido.")
                    else:
                        manana_respondida = [item for item in resp_manana if item["Estado"] is not None]
                        tarde_respondida = [item for item in resp_tarde if item["Estado"] is not None]
                        supervisiones_validas = [item for item in supervision_payload_data if item["Actividad"]]

                        if not manana_respondida and not tarde_respondida and not supervisiones_validas:
                            st.error("⚠️ Ingrese información antes de guardar.")
                        else:
                            payload_jornada = {"Verificaciones": manana_respondida + tarde_respondida, "Supervision_Trabajos": supervisiones_validas}
                            chk_record = {
                                "usuario_email": user_email,
                                "edificio": edificio_val,
                                "fecha": fecha_val.strftime("%Y-%m-%d"),
                                "hora_inicio": hora_inicio_val.strftime("%H:%M"),
                                "hora_fin": hora_fin_val.strftime("%H:%M"),
                                "responsable": user_nombre_completo,
                                "cargo": user_cargo,
                                "observacion_general": "",
                                "datos": payload_jornada
                            }

                            try:
                                if st.session_state.edit_chk_id:
                                    supabase.table("checklists").update(chk_record).eq("id", st.session_state.edit_chk_id).execute()
                                    st.success(f"¡Checklist actualizado correctamente para **{edificio_val}**!")
                                else:
                                    supabase.table("checklists").insert(chk_record).execute()
                                    st.success(f"¡Checklist guardado permanentemente para **{edificio_val}**!")

                                components.html("""<script>try { if (window.top.alphaBuildersClearDraft) window.top.alphaBuildersClearDraft(); } catch(e) {}</script>""", height=0, width=0)
                                st.session_state.db_loaded = False
                                st.session_state.creando_jornada = False
                                st.session_state.edit_chk_id = None
                                st.session_state.filas_supervision = [{"id": 1, "actividad": ""}]
                                st.session_state.chk_obs_counts = {}
                                st.session_state.chk_responsables_extra = {}
                                for k in ["chk_edit_edif_val", "chk_edit_fecha_val", "chk_edit_hini_val", "chk_edit_hfin_val", "chk_edit_resp_map"]:
                                    if k in st.session_state:
                                        del st.session_state[k]
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al guardar checklist: {e}")

                lbl_cancel_chk = "❌ Cancelar Edición" if st.session_state.edit_chk_id else "❌ Cancelar Llenado"
                if st.button(lbl_cancel_chk, key="btn_cancel_chk_bottom", use_container_width=True):
                    st.session_state.edit_chk_id = None
                    st.session_state.creando_jornada = False
                    st.session_state.filas_supervision = [{"id": 1, "actividad": ""}]
                    st.session_state.chk_responsables_extra = {}
                    for k in ["chk_edit_edif_val", "chk_edit_fecha_val", "chk_edit_hini_val", "chk_edit_hfin_val", "chk_edit_resp_map"]:
                        if k in st.session_state:
                            del st.session_state[k]
                    st.rerun()

        st.markdown("---")
        st.markdown("### Historial General de Checklists Creados")
        mis_jornadas = st.session_state.get("db_checklists", {}).get(user_email, [])

        if len(mis_jornadas) > 0:
            col_edif_sel, _ = st.columns([2, 2])
            with col_edif_sel:
                edificio_filtro = st.selectbox("🏢 Seleccionar Edificio / Proyecto:", ["-- Todos los Edificios --"] + EDIFICIOS_ALPHA, key="filtro_edificio_historial")

            jornadas_filtradas = [j for j in mis_jornadas if j.get("Edificio") == edificio_filtro] if edificio_filtro != "-- Todos los Edificios --" else mis_jornadas.copy()
            st.caption(f"Mostrando **{len(jornadas_filtradas)}** checklist(s).")

            df_jornadas = pd.DataFrame(jornadas_filtradas)
            df_jornadas["fecha_dt"] = pd.to_datetime(df_jornadas["Fecha"])
            df_jornadas = df_jornadas.sort_values(by="fecha_dt", ascending=False)
            df_jornadas["año"] = df_jornadas["fecha_dt"].dt.year
            df_jornadas["mes_num"] = df_jornadas["fecha_dt"].dt.month

            grupos_chk = df_jornadas.groupby(["año", "mes_num"], sort=False)

            for (anio, mes_num), items_mes in grupos_chk:
                nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Checklists)"
                with st.expander(nombre_mes_str, expanded=False):
                    for orig_idx, j in items_mes.iterrows():
                        j_dict = j.to_dict()
                        chk_db_id = j_dict.get("db_id")

                        with st.expander(f"📌 {j_dict['Edificio']} — {j_dict['Fecha']} (Horario: {j_dict.get('Hora_Inicio', 'N/A')} - {j_dict.get('Hora_Fin', 'N/A')})", expanded=False):
                            c_dl1, c_dl2, c_ed_chk, c_del_chk = st.columns([2, 2, 1, 1])
                            with c_dl1:
                                with st.popover("📊 Exportar Excel", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.xlsx)", 
                                        get_cached_checklist_excel(safe_json_dumps(j_dict)), 
                                        file_name=f"Checklist_{j_dict['Edificio']}_{j_dict['Fecha']}.xlsx", 
                                        key=f"dl_xlsx_{orig_idx}", 
                                        use_container_width=True
                                    )
                            with c_dl2:
                                with st.popover("📄 Exportar PDF", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.pdf)", 
                                        get_cached_checklist_pdf(safe_json_dumps(j_dict)), 
                                        file_name=f"Checklist_{j_dict['Edificio']}_{j_dict['Fecha']}.pdf", 
                                        key=f"dl_pdf_{orig_idx}", 
                                        use_container_width=True
                                    )
                            with c_ed_chk:
                                if st.button("✏️", key=f"btn_edit_chk_{orig_idx}_{chk_db_id}", help="Editar checklist", use_container_width=True):
                                    st.session_state.edit_chk_id = chk_db_id
                                    st.session_state.creando_jornada = True
                                    st.session_state.chk_edit_edif_val = j_dict.get("Edificio")
                                    st.session_state.chk_edit_fecha_val = pd.to_datetime(j_dict.get("Fecha")).date()
                                    try:
                                        st.session_state.chk_edit_hini_val = datetime.datetime.strptime(j_dict.get("Hora_Inicio", "07:00"), "%H:%M").time()
                                        st.session_state.chk_edit_hfin_val = datetime.datetime.strptime(j_dict.get("Hora_Fin", "16:00"), "%H:%M").time()
                                    except Exception:
                                        pass

                                    raw_data_chk = j_dict.get("Datos", {})
                                    parsed_chk_d = raw_data_chk if isinstance(raw_data_chk, (dict, list)) else json.loads(raw_data_chk or "{}") if isinstance(raw_data_chk, str) else {}
                                    
                                    if isinstance(parsed_chk_d, dict):
                                        verifs = parsed_chk_d.get("Verificaciones", [])
                                        sups_rec = parsed_chk_d.get("Supervision_Trabajos", [])
                                    elif isinstance(parsed_chk_d, list):
                                        verifs = parsed_chk_d
                                        sups_rec = []
                                    else:
                                        verifs = []
                                        sups_rec = []

                                    st.session_state.chk_edit_resp_map = {}
                                    for v in verifs:
                                        if isinstance(v, dict):
                                            j_name = str(v.get('Jornada', '')).lower().strip()
                                            a_name = str(v.get('Actividad', '')).lower().strip()
                                            num_item = str(v.get('N°', '')).strip()
                                            if j_name and a_name:
                                                st.session_state.chk_edit_resp_map[f"{j_name}_{a_name}"] = v
                                            if j_name and num_item:
                                                st.session_state.chk_edit_resp_map[f"{j_name}_{num_item}"] = v

                                    st.session_state.filas_supervision = [
                                        {"id": i + 1, "actividad": it.get("Actividad", it.get("actividad", ""))} for i, it in enumerate(sups_rec)
                                    ] if sups_rec else [{"id": 1, "actividad": ""}]
                                    st.rerun()

                            with c_del_chk:
                                if st.button("🗑️", key=f"btn_del_chk_{orig_idx}_{chk_db_id}", help="Eliminar checklist", use_container_width=True):
                                    try:
                                        supabase.table("checklists").delete().eq("id", chk_db_id).execute()
                                        st.session_state.db_loaded = False
                                        st.success("Checklist eliminado correctamente.")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error al eliminar: {e}")
        else:
            st.info("Aún no hay checklists guardados en tu cuenta.")

    # 2. LIBRO DE OBRA OFICIAL (ESTRUCTURA ORIGINAL CON BOTÓN LLENAR Y CIERRE)
    with tab_libro:
        st.markdown("### Libro de Obra – Formato Oficial")
        st.caption("Estructura técnica de control diario con recopilación automática de personal y sincronización de Checklist.")

        if "llenando_libro_oficial" not in st.session_state:
            st.session_state.llenando_libro_oficial = False

        if "edit_lo_id" not in st.session_state:
            st.session_state.edit_lo_id = None

        # Botón inicial para abrir el formulario del libro oficial
        if not st.session_state.llenando_libro_oficial and not st.session_state.edit_lo_id:
            if st.button("➕ Llenar Libro de Obra", type="primary", key="btn_open_llenar_lo_oficial"):
                st.session_state.llenando_libro_oficial = True
                st.session_state.edit_lo_id = None
                st.session_state.filas_lo_actividades = [{"id": 1, "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}]
                for k in ["lo_edit_proy_val", "lo_edit_fecha_val", "lo_edit_ubic_val", "lo_edit_barr_val", "lo_edit_super_val", "lo_edit_fisc_val", "lo_edit_hoja_val", "lo_edit_nov_val", "lo_edit_hini_val", "lo_edit_hfin_val", "lo_edit_clima_val", "lo_edit_clima_obs_val", "lo_edit_nom_map", "lo_edit_rot_map", "lo_edit_maq_map", "lo_edit_seg_map"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        # Despliegue del formulario únicamente si está activo
        if st.session_state.llenando_libro_oficial or st.session_state.edit_lo_id:
            st.markdown("---")
            if st.session_state.edit_lo_id:
                st.info("✏️ **Modo Edición de Libro de Obra Activo**")

            dias_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
            local_today_insp = get_local_datetime_ecuador().date()
            lo_fecha_def = st.session_state.get("lo_edit_fecha_val", local_today_insp)
            es_edicion_lo = bool(st.session_state.edit_lo_id)

            col_lo1, col_lo2 = st.columns([1.5, 2.5])
            with col_lo1:
                lo_fecha = st.date_input(
                    "Fecha de Reporte:", 
                    lo_fecha_def, 
                    disabled=es_edicion_lo, 
                    key=f"lo_official_fecha_{st.session_state.get('edit_lo_id', 'new')}"
                )
                lo_dia = dias_es[lo_fecha.weekday()]
                st.caption(f"Día seleccionado: **{lo_dia}**" + (" *(🔒 Fecha fija en edición)*" if es_edicion_lo else ""))

            fecha_lo_str = lo_fecha.strftime("%Y-%m-%d")
            chk_asociado = next((c for c in mis_jornadas if c.get("Fecha") == fecha_lo_str), None)

            proyectos_libro = user_edificios if len(user_edificios) > 0 else EDIFICIOS_ALPHA
            idx_proy_lo = 0
            if "lo_edit_proy_val" in st.session_state and st.session_state.lo_edit_proy_val in proyectos_libro:
                idx_proy_lo = proyectos_libro.index(st.session_state.lo_edit_proy_val) + 1
            elif chk_asociado and chk_asociado.get("Edificio") in proyectos_libro:
                idx_proy_lo = proyectos_libro.index(chk_asociado.get("Edificio")) + 1

            with col_lo2:
                if chk_asociado:
                    st.success(f"🔗 **Checklist Vinculado:** Se sincronizan las actividades del edificio **{chk_asociado.get('Edificio')}**.")
                else:
                    st.info("💡 Si llenas el Checklist del día, los trabajos y horarios se precargarán automáticamente aquí.")

            lo_proyecto = st.selectbox(
                "🏢 Proyecto Asignado:*",
                ["-- Seleccione --"] + proyectos_libro,
                index=idx_proy_lo,
                key="lo_proy_sel_off"
            )

            mi_personal_obra = st.session_state.get("db_trabajadores_por_usuario", {}).get(user_email, [])
            if lo_proyecto != "-- Seleccione --":
                personal_edificio = [
                    p for p in mi_personal_obra 
                    if (p.get("edificio") == lo_proyecto or p.get("edificio") == "General" or not p.get("edificio"))
                ]
                if not personal_edificio:
                    personal_edificio = mi_personal_obra
            else:
                personal_edificio = mi_personal_obra

            def calcular_conteo_oficio(nombre_oficio, lista_personal):
                ofi_clean = nombre_oficio.upper().strip()
                total = 0
                for persona in lista_personal:
                    cargo = persona.get("cargo", "").upper().strip()
                    if not cargo:
                        continue
                    if ofi_clean == "ALBAÑILES" and ("ALBAÑIL" in cargo or "ALBANIL" in cargo):
                        total += 1
                    elif ofi_clean == "AYUDANTE" and ("AYUDANTE" in cargo or "PEON" in cargo):
                        total += 1
                    elif ofi_clean == "SOLDADOR" and ("SOLDADOR" in cargo or "SOLDA" in cargo):
                        total += 1
                    elif ofi_clean == "OPERADOR" and ("OPERADOR" in cargo or "MAQUINISTA" in cargo):
                        total += 1
                    elif ofi_clean == "FIERRERO" and ("FIERRERO" in cargo or "HIERRERO" in cargo or "ESTRUCTURA" in cargo):
                        total += 1
                    elif ofi_clean == "PINTORES" and ("PINTOR" in cargo or "PINTORES" in cargo):
                        total += 1
                    elif ofi_clean == "MAESTRO SUPERVISOR" and ("MAESTRO" in cargo or "SUPERVISOR" in cargo or "CAPATAZ" in cargo):
                        total += 1
                    elif ofi_clean == "CARPINTERO" and ("CARPINTERO" in cargo or "MADERA" in cargo):
                        total += 1
                    elif ofi_clean == "GUACHIMAN" and ("GUACHIMAN" in cargo or "GUARDIA" in cargo or "SEGURIDAD" in cargo):
                        total += 1
                    elif ofi_clean == "GYPSEROS" and ("GYPS" in cargo or "TABLAYESO" in cargo):
                        total += 1
                    elif ofi_clean == "ELECTRICOS" and ("ELECTRIC" in cargo):
                        total += 1
                    elif ofi_clean == "PLOMEROS" and ("PLOMER" in cargo or "HIDROSANITARI" in cargo or "FONTANER" in cargo):
                        total += 1
                    elif ofi_clean == "ALUMINIO Y VIDRIO" and ("ALUMIN" in cargo or "VIDRI" in cargo):
                        total += 1
                    elif ofi_clean in cargo or cargo in ofi_clean:
                        total += 1
                return total

            st.markdown("#### 1. Datos Generales de la Obra")
            c_lo_a1, c_lo_a2, c_lo_a3 = st.columns([1.5, 1.5, 1])
            with c_lo_a1:
                ubic_def = st.session_state.get("lo_edit_ubic_val", "CALLE LUXEMBURGO Y HOLANDA")
                barr_def = st.session_state.get("lo_edit_barr_val", "BENALCAZAR")
                lo_ubicacion = st.text_input("Ubicación:*", value=ubic_def, key="lo_ubic_in")
                lo_barrio = st.text_input("Barrio:", value=barr_def, key="lo_barr_in")

            with c_lo_a2:
                super_def = st.session_state.get("lo_edit_super_val", "ING. PABLO ESPINOSA")
                fisc_def = st.session_state.get("lo_edit_fisc_val", "ING. DIEGO CHARVET")
                lo_superintendente = st.text_input("Superintendente:*", value=super_def, key="lo_super_in")
                lo_residente = st.text_input("Residente de Obra:*", value=user_nombre_completo, key="lo_res_in")
                lo_fiscalizador = st.text_input("Fiscalizador:*", value=fisc_def, key="lo_fisc_in")

            with c_lo_a3:
                hoja_def = st.session_state.get("lo_edit_hoja_val", "000053")
                hini_lo_def = st.session_state.get("lo_edit_hini_val", datetime.time(7, 0))
                hfin_lo_def = st.session_state.get("lo_edit_hfin_val", datetime.time(16, 0))
                lo_hoja = st.text_input("Hoja N°:*", value=hoja_def, key="lo_hoja_in")
                lo_h_ini = st.time_input("Hora Entrada:*", hini_lo_def, key="lo_hini_off")
                lo_h_fin = st.time_input("Hora Salida:*", hfin_lo_def, key="lo_hfin_off")

            st.markdown("---")

            st.markdown("#### 2. Jornada de Trabajo y Nómina de Personal")
            st.caption(f"Cálculo automático de cuadrilla ({len(personal_edificio)} integrantes activos detectados en tu nómina para este proyecto):")

            saved_nom_map = st.session_state.get("lo_edit_nom_map", {})
            saved_rot_map = st.session_state.get("lo_edit_rot_map", {})

            c_nom1, c_nom2 = st.columns(2)
            with c_nom1:
                st.markdown("##### 👷 Personal Nómina")
                nomina_input_map = {}
                for ofi in OFICIOS_NOMINA_FORMATO:
                    if st.session_state.edit_lo_id and ofi in saved_nom_map:
                        conteo_val = int(saved_nom_map[ofi])
                    else:
                        conteo_val = int(calcular_conteo_oficio(ofi, personal_edificio))

                    col_n_l, col_n_v = st.columns([3, 1])
                    with col_n_l:
                        st.write(f"• {ofi} (7:00AM - 4:00PM)")
                    with col_n_v:
                        nomina_input_map[ofi] = st.number_input(
                            f"N_{ofi}",
                            min_value=0,
                            value=conteo_val,
                            key=f"lo_nom_{ofi}_{lo_proyecto}_{st.session_state.get('edit_lo_id', 'new')}",
                            label_visibility="collapsed"
                        )

            with c_nom2:
                st.markdown("##### 🔄 Personal Rotativo / Subcontratos")
                rotativo_input_map = {}
                for rubro in RUROS_ROTATIVOS_FORMATO:
                    val_rot_prev = int(saved_rot_map.get(rubro, 0)) if st.session_state.edit_lo_id else 0
                    col_r_l, col_r_v = st.columns([3, 1])
                    with col_r_l:
                        st.write(f"• {rubro}")
                    with col_r_v:
                        rotativo_input_map[rubro] = st.number_input(
                            f"R_{rubro}", 
                            min_value=0, 
                            value=val_rot_prev, 
                            key=f"lo_rot_{rubro}_{st.session_state.get('edit_lo_id', 'new')}", 
                            label_visibility="collapsed"
                        )

            st.markdown("---")

            st.markdown("#### 3. Condiciones Climáticas y Maquinaria / Herramientas")
            clima_opts = ["SOL", "NUBLADO", "LLUVIA TENUE", "LLUVIA INTENSA", "GRANIZO"]
            saved_clima_cond = st.session_state.get("lo_edit_clima_val", "SOL")
            idx_clima = clima_opts.index(saved_clima_cond) if saved_clima_cond in clima_opts else 0
            saved_clima_obs = st.session_state.get("lo_edit_clima_obs_val", "")
            saved_maq_map = st.session_state.get("lo_edit_maq_map", {})

            c_cl1, c_cl2 = st.columns(2)
            with c_cl1:
                st.markdown("##### ⛅ Condiciones Climáticas")
                clima_cond_sel = st.selectbox("Estado del Clima:*", clima_opts, index=idx_clima, key="lo_clima_cond")
                clima_obs = st.text_input("Observaciones del Clima en Obra:", value=saved_clima_obs, placeholder="Describa si el clima afectó el rendimiento...", key="lo_clima_obs")

            with c_cl2:
                st.markdown("##### ⚙️ Maquinaria / Herramientas en Operación")
                maq_input_map = {}
                for maq in MAQUINARIAS_FORMATO:
                    val_maq_prev = int(saved_maq_map.get(maq, 0)) if st.session_state.edit_lo_id else 0
                    c_m_l, c_m_v = st.columns([3, 1])
                    with c_m_l:
                        st.write(f"• {maq}")
                    with c_m_v:
                        maq_input_map[maq] = st.number_input(
                            f"M_{maq}", 
                            min_value=0, 
                            value=val_maq_prev, 
                            key=f"lo_maq_{maq}_{st.session_state.get('edit_lo_id', 'new')}", 
                            label_visibility="collapsed"
                        )

            st.markdown("---")

            st.markdown("#### 4. Seguridad Industrial, Señalización y Mitigación")
            saved_seg_map = st.session_state.get("lo_edit_seg_map", {})

            c_ss1, c_ss2, c_ss3 = st.columns(3)
            with c_ss1:
                st.markdown("##### 🛡️ Seguridad")
                seg_casco = st.checkbox("Casco", value=bool(saved_seg_map.get("Casco", False)), key=f"seg_casco_{st.session_state.get('edit_lo_id', 'new')}")
                seg_chaleco = st.checkbox("Chalecos", value=bool(saved_seg_map.get("Chalecos", False)), key=f"seg_chaleco_{st.session_state.get('edit_lo_id', 'new')}")
                seg_guantes = st.checkbox("Guantes", value=bool(saved_seg_map.get("Guantes", False)), key=f"seg_guantes_{st.session_state.get('edit_lo_id', 'new')}")
                seg_gafas = st.checkbox("Gafas", value=bool(saved_seg_map.get("Gafas", False)), key=f"seg_gafas_{st.session_state.get('edit_lo_id', 'new')}")
                seg_mascarilla = st.checkbox("Mascarilla", value=bool(saved_seg_map.get("Mascarilla", False)), key=f"seg_mascarilla_{st.session_state.get('edit_lo_id', 'new')}")
                seg_auditivo = st.checkbox("Auditivo", value=bool(saved_seg_map.get("Auditivo", False)), key=f"seg_auditivo_{st.session_state.get('edit_lo_id', 'new')}")

            with c_ss2:
                st.markdown("##### 🚧 Señalización")
                sen_conos = st.checkbox("Conos", value=bool(saved_seg_map.get("Conos", False)), key=f"sen_conos_{st.session_state.get('edit_lo_id', 'new')}")
                sen_cintas = st.checkbox("Cintas", value=bool(saved_seg_map.get("Cintas", False)), key=f"sen_cintas_{st.session_state.get('edit_lo_id', 'new')}")
                sen_rotulos = st.checkbox("Rótulos", value=bool(saved_seg_map.get("Rótulos", False)), key=f"sen_rotulos_{st.session_state.get('edit_lo_id', 'new')}")
                sen_vallas = st.checkbox("Vallas", value=bool(saved_seg_map.get("Vallas", False)), key=f"sen_vallas_{st.session_state.get('edit_lo_id', 'new')}")
                sen_extintor = st.checkbox("Extintor", value=bool(saved_seg_map.get("Extintor", False)), key=f"sen_extintor_{st.session_state.get('edit_lo_id', 'new')}")
                sen_botiquin = st.checkbox("Botiquín", value=bool(saved_seg_map.get("Botiquin", False)), key=f"sen_botiquin_{st.session_state.get('edit_lo_id', 'new')}")

            with c_ss3:
                mit_polvo = st.checkbox("Control de Polvo", value=bool(saved_seg_map.get("Polvo", False)), key=f"mit_polvo_{st.session_state.get('edit_lo_id', 'new')}")
                mit_ruido = st.checkbox("Control de Ruido", value=bool(saved_seg_map.get("Ruido", False)), key=f"mit_ruido_{st.session_state.get('edit_lo_id', 'new')}")
                mit_liquidos = st.checkbox("Líquidos Contaminantes", value=bool(saved_seg_map.get("Liquidos", False)), key=f"mit_liquidos_{st.session_state.get('edit_lo_id', 'new')}")
                mit_cerramiento = st.checkbox("Cerramiento", value=bool(saved_seg_map.get("Cerramiento", False)), key=f"mit_cerramiento_{st.session_state.get('edit_lo_id', 'new')}")
                mit_limpieza = st.checkbox("Limpieza y Orden", value=bool(saved_seg_map.get("Limpieza", False)), key=f"mit_limpieza_{st.session_state.get('edit_lo_id', 'new')}")

            st.markdown("---")
            st.markdown("#### 5. Actividades Realizadas dentro de la Jornada Laboral")
            st.caption("Estructura completa de actividades con frente, cuadrillas y cantidades:")

            if "filas_lo_actividades" not in st.session_state:
                st.session_state.filas_lo_actividades = []

            checklist_clave_sesion = f"init_chk_lo_{fecha_lo_str}_{lo_proyecto}"
            if chk_asociado and checklist_clave_sesion not in st.session_state and not st.session_state.edit_lo_id:
                raw_dchk = chk_asociado.get("Datos", {})
                d_p_chk = raw_dchk if isinstance(raw_dchk, dict) else json.loads(raw_dchk or "{}") if isinstance(raw_dchk, str) else {}
                sup_items = d_p_chk.get("Supervision_Trabajos", []) if isinstance(d_p_chk, dict) else []
                importados_chk = []
                for s_it in sup_items:
                    if s_it.get("Actividad"):
                        importados_chk.append({
                            "id": len(importados_chk) + 1,
                            "descripcion": s_it.get("Actividad"),
                            "area": "",
                            "unidad": "",
                            "cantidad": 0.0
                        })
                if importados_chk:
                    st.session_state.filas_lo_actividades = importados_chk
                st.session_state[checklist_clave_sesion] = True

            if not st.session_state.filas_lo_actividades:
                st.session_state.filas_lo_actividades = [
                    {"id": 1, "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}
                ]

            indices_eliminar_lo = []
            acts_final_payload = []

            for idx_act_form, item_f in enumerate(st.session_state.filas_lo_actividades, 1):
                f_act_id = item_f["id"]
                st.markdown(f"**Actividad N° {idx_act_form}:**")
                
                c_af1, c_af2 = st.columns([2.5, 1.5])
                with c_af1:
                    d_in = st.text_input(
                        f"Descripción de la Actividad {idx_act_form}:",
                        value=item_f.get("descripcion", ""),
                        placeholder="Ej. Albañilería, Enlucidos, Pintura...",
                        key=f"lo_act_d_{f_act_id}"
                    )
                with c_af2:
                    ar_in = st.text_input(
                        f"Área de Trabajo {idx_act_form}:",
                        value=item_f.get("area", ""),
                        placeholder="Ej. Piso 3, Bloque A...",
                        key=f"lo_act_ar_{f_act_id}"
                    )

                c_af3, c_af4, c_af5 = st.columns([2.0, 1.0, 1.0])
                with c_af3:
                    u_in = st.text_input(
                        f"Unidad {idx_act_form}:",
                        value=item_f.get("unidad", ""),
                        placeholder="Ej. m2, m, glb...",
                        key=f"lo_act_u_{f_act_id}"
                    )
                with c_af4:
                    ct_in = st.number_input(
                        f"Cantidad {idx_act_form}:",
                        min_value=0.0,
                        value=float(item_f.get("cantidad", 0.0)),
                        step=0.5,
                        key=f"lo_act_ct_{f_act_id}"
                    )
                with c_af5:
                    st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)
                    if st.button("🗑️", key=f"btn_del_lo_row_{f_act_id}", help="Eliminar fila"):
                        indices_eliminar_lo.append(idx_act_form - 1)

                acts_final_payload.append({
                    "Descripcion": d_in.strip(),
                    "Area": ar_in.strip(),
                    "Unidad": u_in.strip(),
                    "Cantidad": ct_in
                })
                st.markdown("---")

            if indices_eliminar_lo:
                for del_i in sorted(indices_eliminar_lo, reverse=True):
                    if len(st.session_state.filas_lo_actividades) > 1:
                        st.session_state.filas_lo_actividades.pop(del_i)
                    else:
                        st.session_state.filas_lo_actividades = [{"id": int(datetime.datetime.now().timestamp() * 1000), "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}]
                st.rerun()

            if st.button("➕ Agregar Fila de Trabajo", key="btn_add_lo_actividad_extra"):
                next_id_lo = (max([x["id"] for x in st.session_state.filas_lo_actividades]) + 1) if st.session_state.filas_lo_actividades else 1
                st.session_state.filas_lo_actividades.append({
                    "id": next_id_lo,
                    "descripcion": "",
                    "area": "",
                    "unidad": "",
                    "cantidad": 0.0
                })
                st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("#### 6. Novedades y Recomendaciones")
            nov_def = st.session_state.get("lo_edit_nov_val", "")
            lo_novedades = st.text_area("Observaciones Generales / Recomendaciones de Supervisión:*", value=nov_def, placeholder="Escriba las novedades y recomendaciones del día...", height=90, key="lo_nov_in")

            lbl_btn_lo = "🔄 Actualizar Formato Oficial de Libro de Obra" if st.session_state.edit_lo_id else "💾 Guardar Formato Oficial de Libro de Obra"
            btn_guardar_lo_oficial = st.button(lbl_btn_lo, type="primary", use_container_width=True, key="btn_save_lo_official_main")

            if btn_guardar_lo_oficial:
                if lo_proyecto == "-- Seleccione --":
                    st.error("⚠️ Seleccione un Proyecto o Edificio.")
                else:
                    seguridad_checks = {
                        "Casco": seg_casco,
                        "Chalecos": seg_chaleco,
                        "Guantes": seg_guantes,
                        "Gafas": seg_gafas,
                        "Mascarilla": seg_mascarilla,
                        "Auditivo": seg_auditivo,
                        "Conos": sen_conos,
                        "Cintas": sen_cintas,
                        "Rótulos": sen_rotulos,
                        "Vallas": sen_vallas,
                        "Extintor": sen_extintor,
                        "Botiquin": sen_botiquin,
                        "Polvo": mit_polvo,
                        "Ruido": mit_ruido,
                        "Liquidos": mit_liquidos,
                        "Cerramiento": mit_cerramiento,
                        "Limpieza": mit_limpieza
                    }

                    payload_libro_oficial = {
                        "Superintendente": lo_superintendente,
                        "Fiscalizador": lo_fiscalizador,
                        "Ubicacion": lo_ubicacion,
                        "Barrio": lo_barrio,
                        "Hoja": lo_hoja,
                        "Nomina_Conteo": nomina_input_map,
                        "Rotativo_Conteo": rotativo_input_map,
                        "Clima_Condicion": clima_cond_sel,
                        "Clima_Obs": clima_obs,
                        "Maquinaria_Conteo": maq_input_map,
                        "Seguridad_Check": seguridad_checks,
                        "Actividades_Ejecutadas": acts_final_payload,
                        "Novedades": lo_novedades
                    }

                    lo_record = {
                        "usuario_email": user_email,
                        "proyecto": lo_proyecto,
                        "fecha": lo_fecha.strftime("%Y-%m-%d"),
                        "dia": lo_dia,
                        "residente": lo_residente,
                        "frente": lo_ubicacion,
                        "clima": clima_cond_sel,
                        "hora_inicio": lo_h_ini.strftime("%H:%M"),
                        "hora_fin": lo_h_fin.strftime("%H:%M"),
                        "datos": payload_libro_oficial
                    }

                    try:
                        if st.session_state.edit_lo_id:
                            supabase.table("inspecciones").update(lo_record).eq("id", st.session_state.edit_lo_id).execute()
                            st.success(f"¡Libro de Obra actualizado exitosamente para **{lo_proyecto}**!")
                        else:
                            supabase.table("inspecciones").insert(lo_record).execute()
                            st.success(f"¡Libro de Obra guardado exitosamente para **{lo_proyecto}**!")

                        components.html("""<script>try { if (window.top.alphaBuildersClearDraft) window.top.alphaBuildersClearDraft(); } catch(e) {}</script>""", height=0, width=0)
                        st.session_state.db_loaded = False
                        st.session_state.llenando_libro_oficial = False
                        st.session_state.edit_lo_id = None
                        st.session_state.filas_lo_actividades = [{"id": 1, "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}]
                        for k in ["lo_edit_proy_val", "lo_edit_fecha_val", "lo_edit_ubic_val", "lo_edit_barr_val", "lo_edit_super_val", "lo_edit_fisc_val", "lo_edit_hoja_val", "lo_edit_nov_val", "lo_edit_hini_val", "lo_edit_hfin_val", "lo_edit_clima_val", "lo_edit_clima_obs_val", "lo_edit_nom_map", "lo_edit_rot_map", "lo_edit_maq_map", "lo_edit_seg_map"]:
                            if k in st.session_state:
                                del st.session_state[k]
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al procesar: {e}")

            lbl_cancel_lo = "❌ Cancelar Edición" if st.session_state.edit_lo_id else "❌ Cancelar Llenado"
            if st.button(lbl_cancel_lo, key="btn_cancel_lo_bottom", use_container_width=True):
                st.session_state.llenando_libro_oficial = False
                st.session_state.edit_lo_id = None
                st.session_state.filas_lo_actividades = [{"id": 1, "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}]
                for k in ["lo_edit_proy_val", "lo_edit_fecha_val", "lo_edit_ubic_val", "lo_edit_barr_val", "lo_edit_super_val", "lo_edit_fisc_val", "lo_edit_hoja_val", "lo_edit_nov_val", "lo_edit_hini_val", "lo_edit_hfin_val", "lo_edit_clima_val", "lo_edit_clima_obs_val", "lo_edit_nom_map", "lo_edit_rot_map", "lo_edit_maq_map", "lo_edit_seg_map"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        st.markdown("---")

        # HISTORIAL DE FORMATOS
        st.markdown("### Historial de Formatos en Libro de Obra")
        mis_inspecciones = st.session_state.get("db_inspecciones", {}).get(user_email, [])

        if len(mis_inspecciones) > 0:
            col_edif_insp, _ = st.columns([2, 2])
            with col_edif_insp:
                edif_insp_filtro = st.selectbox("🏢 Filtrar Proyecto:", ["-- Todos los Proyectos --"] + EDIFICIOS_ALPHA, key="filtro_edif_inspecciones")

            insps_filtradas = [i for i in mis_inspecciones if i.get("Proyecto") == edif_insp_filtro] if edif_insp_filtro != "-- Todos los Proyectos --" else mis_inspecciones.copy()
            st.caption(f"Mostrando **{len(insps_filtradas)}** registro(s) en Libro de Obra.")

            df_insps = pd.DataFrame(insps_filtradas)
            df_insps["fecha_dt"] = pd.to_datetime(df_insps["Fecha"])
            df_insps = df_insps.sort_values(by="fecha_dt", ascending=False)
            df_insps["año"] = df_insps["fecha_dt"].dt.year
            df_insps["mes_num"] = df_insps["fecha_dt"].dt.month

            grupos_insp = df_insps.groupby(["año", "mes_num"], sort=False)

            for (anio, mes_num), items_mes in grupos_insp:
                nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Libros de Obra)"
                with st.expander(nombre_mes_str, expanded=False):
                    for idx_insp, insp in items_mes.iterrows():
                        insp_dict = insp.to_dict()
                        insp_db_id = insp_dict.get("db_id")

                        with st.expander(f"📌 {insp_dict['Proyecto']} — {insp_dict['Fecha']} ({insp_dict['Dia']}) | Residente: {insp_dict.get('Residente', 'N/A')}", expanded=False):
                            raw_dinsp = insp_dict.get("Datos", {})
                            d_insp = raw_dinsp if isinstance(raw_dinsp, dict) else json.loads(raw_dinsp or "{}") if isinstance(raw_dinsp, str) else {}
                            
                            st.markdown(f"**Ubicación:** {d_insp.get('Ubicacion', insp_dict.get('Frente', ''))} | **Clima:** {insp_dict.get('Clima', '')}")
                            st.markdown(f"**Superintendente:** {d_insp.get('Superintendente', '')} | **Fiscalizador:** {d_insp.get('Fiscalizador', '')}")

                            acts_mostrar = d_insp.get("Actividades_Ejecutadas", [])
                            if acts_mostrar:
                                st.markdown("**Actividades realizadas:**")
                                for a in acts_mostrar:
                                    st.write(f"• **{a.get('Descripcion', '')}** | Área: {a.get('Area', '')} | {a.get('Cantidad', 0)} {a.get('Unidad', 'm2')}")

                            c_dl_i1, c_dl_i2, c_ed_lo, c_del_lo = st.columns([2, 2, 1, 1])
                            with c_dl_i1:
                                with st.popover("📊 Exportar Excel", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.xlsx)",
                                        get_cached_libro_oficial_excel(safe_json_dumps(insp_dict)),
                                        file_name=f"Libro_Obra_{insp_dict['Proyecto']}_{insp_dict['Fecha']}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"dl_lo_off_x_{idx_insp}",
                                        use_container_width=True
                                    )
                            with c_dl_i2:
                                with st.popover("📄 Exportar PDF", use_container_width=True):
                                    st.download_button(
                                        "Confirmar Descarga (.pdf)",
                                        get_cached_libro_oficial_pdf(safe_json_dumps(insp_dict)),
                                        file_name=f"Libro_Obra_{insp_dict['Proyecto']}_{insp_dict['Fecha']}.pdf",
                                        mime="application/pdf",
                                        key=f"dl_lo_off_p_{idx_insp}",
                                        use_container_width=True
                                    )
                            with c_ed_lo:
                                if st.button("✏️", key=f"btn_edit_lo_{idx_insp}_{insp_db_id}", help="Editar Libro de Obra", use_container_width=True):
                                    st.session_state.edit_lo_id = insp_db_id
                                    st.session_state.llenando_libro_oficial = True
                                    st.session_state.lo_edit_proy_val = insp_dict.get("Proyecto")
                                    st.session_state.lo_edit_fecha_val = pd.to_datetime(insp_dict.get("Fecha")).date()
                                    raw_ed = insp_dict.get("Datos", {})
                                    d_ed = raw_ed if isinstance(raw_ed, dict) else json.loads(raw_ed or "{}") if isinstance(raw_ed, str) else {}
                                    
                                    st.session_state.lo_edit_ubic_val = d_ed.get("Ubicacion", "CALLE LUXEMBURGO Y HOLANDA")
                                    st.session_state.lo_edit_barr_val = d_ed.get("Barrio", "BENALCAZAR")
                                    st.session_state.lo_edit_super_val = d_ed.get("Superintendente", "ING. PABLO ESPINOSA")
                                    st.session_state.lo_edit_fisc_val = d_ed.get("Fiscalizador", "ING. DIEGO CHARVET")
                                    st.session_state.lo_edit_hoja_val = d_ed.get("Hoja", "000053")
                                    st.session_state.lo_edit_nov_val = d_ed.get("Novedades", "")
                                    st.session_state.lo_edit_clima_val = d_ed.get("Clima_Condicion", insp_dict.get("Clima", "SOL"))
                                    st.session_state.lo_edit_clima_obs_val = d_ed.get("Clima_Obs", "")
                                    st.session_state.lo_edit_nom_map = d_ed.get("Nomina_Conteo", {})
                                    st.session_state.lo_edit_rot_map = d_ed.get("Rotativo_Conteo", {})
                                    st.session_state.lo_edit_maq_map = d_ed.get("Maquinaria_Conteo", {})
                                    st.session_state.lo_edit_seg_map = d_ed.get("Seguridad_Check", {})
                                    
                                    try:
                                        st.session_state.lo_edit_hini_val = datetime.datetime.strptime(insp_dict.get("Hora_Inicio", "07:00"), "%H:%M").time()
                                        st.session_state.lo_edit_hfin_val = datetime.datetime.strptime(insp_dict.get("Hora_Fin", "16:00"), "%H:%M").time()
                                    except Exception:
                                        pass
                                        
                                    acts_lo_rec = d_ed.get("Actividades_Ejecutadas", [])
                                    st.session_state.filas_lo_actividades = [
                                        {
                                            "id": i + 1,
                                            "descripcion": it.get("Descripcion", it.get("actividad", "")),
                                            "area": it.get("Area", it.get("area", "")),
                                            "unidad": it.get("Unidad", it.get("unidad", "m2")),
                                            "cantidad": float(it.get("Cantidad", it.get("cantidad", 0.0)))
                                        } for i, it in enumerate(acts_lo_rec)
                                    ] if acts_lo_rec else [{"id": 1, "descripcion": "", "area": "", "unidad": "", "cantidad": 0.0}]
                                    st.rerun()

                            with c_del_lo:
                                if st.button("🗑️", key=f"btn_del_lo_{idx_insp}_{insp_db_id}", help="Eliminar registro", use_container_width=True):
                                    try:
                                        supabase.table("inspecciones").delete().eq("id", insp_db_id).execute()
                                        st.session_state.db_loaded = False
                                        st.success("Libro de Obra eliminado correctamente.")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Error al eliminar: {e}")
        else:
            st.info("Aún no tienes registros guardados en tu Libro de Obra.")
# ==============================================================================
# PARTE 5 DE 5: PERSONAL, INCIDENCIAS, RENDIMIENTO, ESPACIO COLABORATIVO
#                Y PANEL ADMINISTRADOR (COMPATIBILIDAD TOTAL CON HISTÓRICOS)
# ==============================================================================

# ==============================================================================
# 11. MÓDULO 3: PERSONAL A CARGO (ORDENAMIENTO ALFABÉTICO AUTOMÁTICO)
# ==============================================================================
with tab_personal:
    st.markdown("### Nómina de Personal a Cargo")
    st.caption("Administración de personal en obra por edificio, edición de cargos y sincronización colaborativa.")

    proyectos_personal_disp = user_edificios if len(user_edificios) > 0 else EDIFICIOS_ALPHA

    col_btn_ob1, col_btn_ob2, col_btn_ob3 = st.columns([1.5, 1.6, 2.0])

    with col_btn_ob1:
        with st.popover("➕ Registrar Personal", use_container_width=True):
            st.markdown("#### Nuevo Integrante")
            with st.form("form_pop_reg_personal_tab3"):
                nom_pers_in = st.text_input("Nombre Completo:*", placeholder="Ej. Juan Carlos Pérez")
                car_pers_in = st.text_input("Cargo / Especialidad:*", placeholder="Ej. GYPSERO, ALBAÑIL, AYUDANTE")
                edif_pers_in = st.selectbox("Edificio / Proyecto Asignado:*", proyectos_personal_disp, index=0)
                btn_save_pers_pop = st.form_submit_button("💾 Guardar Personal", type="primary", use_container_width=True)

                if btn_save_pers_pop:
                    if nom_pers_in.strip() and car_pers_in.strip():
                        nombre_clean = nom_pers_in.strip().upper()
                        cargo_clean = car_pers_in.strip().upper()
                        
                        try:
                            supabase.table("trabajadores").insert({
                                "usuario_email": user_email,
                                "nombre": nombre_clean,
                                "cargo": cargo_clean,
                                "edificio": edif_pers_in
                            }).execute()
                        except Exception as ex_db:
                            print(f"[Warn] Inserción SQL trabajadores directa: {ex_db}")

                        if "db_trabajadores_por_usuario" not in st.session_state:
                            st.session_state.db_trabajadores_por_usuario = {}
                        if user_email not in st.session_state.db_trabajadores_por_usuario:
                            st.session_state.db_trabajadores_por_usuario[user_email] = []

                        cur_p = st.session_state.db_trabajadores_por_usuario[user_email]
                        new_item = {
                            "id": int(datetime.datetime.now().timestamp() * 1000),
                            "nombre": nombre_clean,
                            "cargo": cargo_clean,
                            "edificio": edif_pers_in,
                            "usuario_email": user_email
                        }
                        cur_p.append(new_item)
                        
                        # Ordenamiento alfabético automático
                        cur_p = sorted(cur_p, key=lambda it_w: str(it_w.get("nombre", "")).upper())
                        st.session_state.db_trabajadores_por_usuario[user_email] = cur_p

                        try:
                            supabase.table("app_config").upsert({
                                "key": f"user_trabajadores_{user_email}",
                                "value": json.dumps(cur_p)
                            }).execute()
                        except Exception as ex_cfg:
                            print(f"[Warn] Respaldo app_config: {ex_cfg}")

                        st.session_state.db_loaded = False
                        st.success(f"¡{nombre_clean} guardado exitosamente en tu nómina!")
                        st.rerun()
                    else:
                        st.error("⚠️ Complete todos los campos obligatorios.")

    with col_btn_ob2:
        with st.popover("📂 Importación Masiva", use_container_width=True):
            st.markdown("#### Cargar Archivo Masivo")
            st.caption("El archivo (.xlsx o .csv) debe contener columnas: Nombre, Cargo y opcionalmente Edificio.")
            edif_default_masivo = st.selectbox("Asignar a este edificio por defecto si no viene en el archivo:", proyectos_personal_disp, key="edif_masivo_sel_p5")
            archivo_excel_pop = st.file_uploader("Seleccione archivo:", type=["xlsx", "csv"], key="upl_personal_modal_p5")
            
            if archivo_excel_pop is not None:
                try:
                    df_sub_pop = pd.read_csv(archivo_excel_pop) if archivo_excel_pop.name.endswith(".csv") else pd.read_excel(archivo_excel_pop)
                    
                    if len(df_sub_pop.columns) >= 2:
                        if st.button("Confirmar Carga Masiva", type="primary", use_container_width=True):
                            registrados_cnt = 0
                            
                            if "db_trabajadores_por_usuario" not in st.session_state:
                                st.session_state.db_trabajadores_por_usuario = {}
                            if user_email not in st.session_state.db_trabajadores_por_usuario:
                                st.session_state.db_trabajadores_por_usuario[user_email] = []
                            
                            cur_p = st.session_state.db_trabajadores_por_usuario[user_email]
                            
                            for _, r_data in df_sub_pop.iterrows():
                                n_val = str(r_data.iloc[0]).strip().upper()
                                c_val = str(r_data.iloc[1]).strip().upper()
                                e_val = str(r_data.iloc[2]).strip() if len(r_data) >= 3 and str(r_data.iloc[2]).strip() != "nan" else edif_default_masivo
                                
                                if n_val and n_val != "NAN":
                                    try:
                                        supabase.table("trabajadores").insert({
                                            "usuario_email": user_email,
                                            "nombre": n_val,
                                            "cargo": c_val,
                                            "edificio": e_val
                                        }).execute()
                                    except Exception:
                                        pass
                                    cur_p.append({
                                        "id": int(datetime.datetime.now().timestamp() * 1000) + registrados_cnt,
                                        "nombre": n_val,
                                        "cargo": c_val,
                                        "edificio": e_val,
                                        "usuario_email": user_email
                                    })
                                    registrados_cnt += 1

                            # Ordenamiento alfabético automático
                            cur_p = sorted(cur_p, key=lambda it_w: str(it_w.get("nombre", "")).upper())
                            st.session_state.db_trabajadores_por_usuario[user_email] = cur_p

                            try:
                                supabase.table("app_config").upsert({
                                    "key": f"user_trabajadores_{user_email}",
                                    "value": json.dumps(cur_p)
                                }).execute()
                            except Exception:
                                pass

                            st.session_state.db_loaded = False
                            st.success(f"¡{registrados_cnt} personas importadas a tu nómina personal!")
                            st.rerun()
                    else:
                        st.error("El archivo debe contener mínimo 2 columnas (Nombre y Cargo).")
                except Exception as e:
                    st.error(f"Error al procesar archivo: {e}")

    with col_btn_ob3:
        with st.popover("👥 Cargar Nómina Colaborativa", use_container_width=True):
            st.markdown("#### Importar Nómina de Compañero")
            st.caption("Copia la lista de personal de un compañero que comparta proyectos contigo a tu propia nómina.")
            
            mis_edifs_norm = set([str(x).strip().lower() for x in st.session_state.get("usuario_edificios", [])])
            companeros_nomina = []
            for u in st.session_state.get("db_usuarios", []):
                u_mail_chk = str(u["Correo"]).lower().strip()
                if u_mail_chk != user_email:
                    u_edifs_norm = set([str(x).strip().lower() for x in u.get("Edificios", [])])
                    if len(mis_edifs_norm.intersection(u_edifs_norm)) > 0:
                        pers_comp = st.session_state.get("db_trabajadores_por_usuario", {}).get(u_mail_chk, [])
                        if len(pers_comp) > 0:
                            companeros_nomina.append((u, pers_comp))

            if len(companeros_nomina) > 0:
                map_comp_nom = {f"{u['Nombres']} {u['Apellidos']} ({len(p_list)} trabajadores)": (u, p_list) for u, p_list in companeros_nomina}
                sel_comp_label = st.selectbox("Seleccione compañero:", list(map_comp_nom.keys()), key="sel_comp_nom_pop_p5")
                comp_elegido_u, lista_p_elegida = map_comp_nom[sel_comp_label]
                
                comunes_n = [e for e in st.session_state.get("usuario_edificios", []) if str(e).strip().lower() in [str(x).strip().lower() for x in comp_elegido_u.get("Edificios", [])]]
                st.info(f"Proyectos en común: **{', '.join(comunes_n)}**")
                
                if st.button("📥 Importar a Mi Nómina Personal", type="primary", use_container_width=True):
                    importados_cnt = 0
                    if user_email not in st.session_state.db_trabajadores_por_usuario:
                        st.session_state.db_trabajadores_por_usuario[user_email] = []
                    
                    cur_p = st.session_state.db_trabajadores_por_usuario[user_email]
                    mi_personal_nombres = [t["nombre"] for t in cur_p]

                    for p_item in lista_p_elegida:
                        if p_item["nombre"] not in mi_personal_nombres:
                            edif_dest = p_item.get("edificio") or (comunes_n[0] if len(comunes_n) > 0 else "General")
                            try:
                                supabase.table("trabajadores").insert({
                                    "usuario_email": user_email,
                                    "nombre": p_item["nombre"],
                                    "cargo": p_item["cargo"],
                                    "edificio": edif_dest
                                }).execute()
                            except Exception:
                                pass
                            cur_p.append({
                                "id": int(datetime.datetime.now().timestamp() * 1000) + importados_cnt,
                                "nombre": p_item["nombre"],
                                "cargo": p_item["cargo"],
                                "edificio": edif_dest,
                                "usuario_email": user_email
                            })
                            mi_personal_nombres.append(p_item["nombre"])
                            importados_cnt += 1

                    # Ordenamiento alfabético automático
                    cur_p = sorted(cur_p, key=lambda it_w: str(it_w.get("nombre", "")).upper())
                    st.session_state.db_trabajadores_por_usuario[user_email] = cur_p

                    try:
                        supabase.table("app_config").upsert({
                            "key": f"user_trabajadores_{user_email}",
                            "value": json.dumps(cur_p)
                        }).execute()
                    except Exception:
                        pass

                    st.session_state.db_loaded = False
                    st.success(f"¡Se copiaron {importados_cnt} personas a tu nómina personal!")
                    st.rerun()
            else:
                st.info("No hay compañeros con proyectos en común que tengan personal registrado aún.")

    st.markdown("---")

    mi_personal_actual = st.session_state.get("db_trabajadores_por_usuario", {}).get(user_email, [])
    # Garantizar orden alfabético
    mi_personal_actual = sorted(mi_personal_actual, key=lambda it_w: str(it_w.get("nombre", "")).upper())
    st.markdown(f"#### Tu Nómina de Personal a Cargo ({len(mi_personal_actual)} integrantes)")

    if len(mi_personal_actual) > 0:
        edificios_con_personal = sorted(list(set([p.get("edificio") or "General" for p in mi_personal_actual])))
        pestanas_edificios = ["🏢 Todos los Edificios"] + [f"📍 {e}" for e in edificios_con_personal]
        tabs_edificios = st.tabs(pestanas_edificios)

        with tabs_edificios[0]:
            st.caption(f"Mostrando el total de **{len(mi_personal_actual)}** integrantes de tu nómina ordenados alfabéticamente.")
            for idx_t, pers in enumerate(mi_personal_actual, 1):
                t_id = pers.get("id", idx_t)
                t_nom = pers["nombre"]
                t_car = pers["cargo"]
                t_edif = pers.get("edificio") or "General"

                c_info, c_edit, c_del = st.columns([7, 1.5, 1.5])
                with c_info:
                    st.markdown(
                        f"""
                        <div class="worker-card-row">
                            <div class="worker-info-block">
                                <div class="worker-name-title">{idx_t}. {t_nom}</div>
                                <div class="worker-meta-tags">
                                    <span class="tag-cargo-chip">{t_car}</span>
                                    <span class="tag-edif-chip">🏢 {t_edif}</span>
                                </div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                with c_edit:
                    with st.popover("✏️", help=f"Editar datos de {t_nom}", use_container_width=True):
                        st.markdown("**Modificar Trabajador**")
                        nuevo_cargo_input = st.text_input("Cargo / Especialidad:", value=t_car, key=f"pop_all_car_{t_id}_{idx_t}_p5")
                        nuevo_edif_input = st.selectbox("Edificio Asignado:", proyectos_personal_disp, index=proyectos_personal_disp.index(t_edif) if t_edif in proyectos_personal_disp else 0, key=f"pop_all_edif_{t_id}_{idx_t}_p5")
                        
                        if st.button("Guardar", key=f"btn_save_all_{t_id}_{idx_t}_p5", type="primary", use_container_width=True):
                            try:
                                try:
                                    supabase.table("trabajadores").update({
                                        "cargo": nuevo_cargo_input.strip().upper(),
                                        "edificio": nuevo_edif_input
                                    }).eq("id", t_id).execute()
                                except Exception:
                                    pass
                                for p in mi_personal_actual:
                                    if p.get("id") == t_id or p.get("nombre") == t_nom:
                                        p["cargo"] = nuevo_cargo_input.strip().upper()
                                        p["edificio"] = nuevo_edif_input
                                supabase.table("app_config").upsert({
                                    "key": f"user_trabajadores_{user_email}",
                                    "value": json.dumps(mi_personal_actual)
                                }).execute()

                                st.session_state.db_loaded = False
                                st.success("Datos actualizados correctamente.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")

                with c_del:
                    if st.button("🗑️", key=f"btn_del_all_{t_id}_{idx_t}_p5", help=f"Remover a {t_nom}", use_container_width=True):
                        try:
                            try:
                                supabase.table("trabajadores").delete().eq("id", t_id).execute()
                            except Exception:
                                pass
                            mi_personal_actual = [p for p in mi_personal_actual if p.get("id") != t_id and p.get("nombre") != t_nom]
                            st.session_state.db_trabajadores_por_usuario[user_email] = mi_personal_actual
                            supabase.table("app_config").upsert({
                                "key": f"user_trabajadores_{user_email}",
                                "value": json.dumps(mi_personal_actual)
                            }).execute()

                            st.session_state.db_loaded = False
                            st.success(f"{t_nom} removido de tu nómina.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al eliminar: {e}")

        for idx_tab, edif_name in enumerate(edificios_con_personal, 1):
            with tabs_edificios[idx_tab]:
                personal_del_edificio = [p for p in mi_personal_actual if (p.get("edificio") or "General") == edif_name]
                st.caption(f"Cuadrilla asignada a **{edif_name}** ({len(personal_del_edificio)} trabajadores):")

                for idx_sub, pers in enumerate(personal_del_edificio, 1):
                    t_id = pers.get("id", idx_sub)
                    t_nom = pers["nombre"]
                    t_car = pers["cargo"]

                    c_info_e, c_edit_e, c_del_e = st.columns([7, 1.5, 1.5])
                    with c_info_e:
                        st.markdown(
                            f"""
                            <div class="worker-card-row">
                                <div class="worker-info-block">
                                    <div class="worker-name-title">{idx_sub}. {t_nom}</div>
                                    <div class="worker-meta-tags">
                                        <span class="tag-cargo-chip">{t_car}</span>
                                        <span class="tag-edif-chip">📍 {edif_name}</span>
                                    </div>
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                    with c_edit_e:
                        with st.popover("✏️", help=f"Editar datos de {t_nom}", key=f"pop_edif_tab_{t_id}_{idx_sub}_p5", use_container_width=True):
                            st.markdown(f"**Modificar — {t_nom}**")
                            n_car = st.text_input("Cargo:", value=t_car, key=f"car_e_{t_id}_{idx_sub}_p5")
                            n_ed = st.selectbox("Cambiar Edificio:", proyectos_personal_disp, index=proyectos_personal_disp.index(edif_name) if edif_name in proyectos_personal_disp else 0, key=f"ed_e_{t_id}_{idx_sub}_p5")
                            
                            if st.button("Guardar", key=f"btn_s_e_{t_id}_{idx_sub}_p5", type="primary", use_container_width=True):
                                try:
                                    try:
                                        supabase.table("trabajadores").update({
                                            "cargo": n_car.strip().upper(),
                                            "edificio": n_ed
                                        }).eq("id", t_id).execute()
                                    except Exception:
                                        pass
                                    for p in mi_personal_actual:
                                        if p.get("id") == t_id or p.get("nombre") == t_nom:
                                            p["cargo"] = n_car.strip().upper()
                                            p["edificio"] = n_ed
                                    supabase.table("app_config").upsert({
                                        "key": f"user_trabajadores_{user_email}",
                                        "value": json.dumps(mi_personal_actual)
                                    }).execute()

                                    st.session_state.db_loaded = False
                                    st.success("Actualizado correctamente.")
                                    st.rerun()
                                except Exception as err:
                                    st.error(f"Error: {err}")

                    with c_del_e:
                        if st.button("🗑️", key=f"btn_d_e_{t_id}_{idx_sub}_p5", help=f"Remover a {t_nom}", use_container_width=True):
                            try:
                                try:
                                    supabase.table("trabajadores").delete().eq("id", t_id).execute()
                                except Exception:
                                    pass
                                mi_personal_actual = [p for p in mi_personal_actual if p.get("id") != t_id and p.get("nombre") != t_nom]
                                st.session_state.db_trabajadores_por_usuario[user_email] = mi_personal_actual
                                supabase.table("app_config").upsert({
                                    "key": f"user_trabajadores_{user_email}",
                                    "value": json.dumps(mi_personal_actual)
                                }).execute()

                                st.session_state.db_loaded = False
                                st.success(f"{t_nom} removido.")
                                st.rerun()
                            except Exception as err:
                                st.error(f"Error: {err}")

        st.markdown("<br>", unsafe_allow_html=True)
        df_export_personal = pd.DataFrame(mi_personal_actual)
        df_export_personal = df_export_personal.drop(columns=["id", "usuario_email"], errors="ignore")
        if "edificio" not in df_export_personal.columns:
            df_export_personal["edificio"] = "General"
        df_export_personal = df_export_personal[["nombre", "cargo", "edificio"]]
        df_export_personal.columns = ["Nombre Completo", "Cargo / Especialidad", "Edificio Asignado"]
        df_export_personal.index = range(1, len(df_export_personal) + 1)
        csv_personal_bytes = export_dataframe_to_excel_csv(df_export_personal)
        
        st.download_button(
            label="📥 Descargar Tu Nómina Completa en CSV (Excel)",
            data=csv_personal_bytes,
            file_name=f"Personal_A_Cargo_{user_email}_{get_local_datetime_ecuador().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="dl_csv_nomina_personal_tab_p5",
            use_container_width=True
        )
    else:
        st.info("Aún no has registrado personal a cargo en tu cuenta. Agrega integrantes con '➕ Registrar Personal' o importa la nómina de un compañero.")

# ==============================================================================
# 12. MÓDULO 4: LEVANTAMIENTO DE INCIDENCIAS (SOLO RESIDENTE Y ASISTENTE)
# ==============================================================================
if not es_maestro_mayor:
    with tab_incidencias:
        st.markdown("### Levantamiento de Incidencias")
        st.caption("Control de no conformidades, responsables, plazos de atención y seguimiento en tus proyectos asignados.")

        proyectos_inc_disp = user_edificios if len(user_edificios) > 0 else EDIFICIOS_ALPHA

        col_fil_inc1, col_fil_inc2 = st.columns([2, 2])
        with col_fil_inc1:
            proy_inc_sel = st.selectbox(
                "🏢 Seleccionar Proyecto / Edificio:",
                ["-- Todos Mis Proyectos --"] + proyectos_inc_disp,
                key="sel_proy_incidencias_p5"
            )
        with col_fil_inc2:
            filtro_estado_vista = st.segmented_control(
                "Filtrar por Estado:",
                ["Todos", "Abierta", "Cerrada"],
                default="Todos",
                key="filtro_estado_inc_view_p5"
            )

        st.markdown("---")

        with st.expander("➕ Registrar Nueva Incidencia en Obra", expanded=False):
            with st.form("form_nueva_incidencia_p5"):
                c_i1, c_i2 = st.columns(2)
                with c_i1:
                    proy_nuevo = st.selectbox(
                        "Proyecto / Edificio:*",
                        proyectos_inc_disp,
                        index=0 if proy_inc_sel == "-- Todos Mis Proyectos --" else (proyectos_inc_disp.index(proy_inc_sel) if proy_inc_sel in proyectos_inc_disp else 0),
                        key="f_inc_proy_p5"
                    )
                    area_nueva = st.text_input("Área / Ubicación:*", placeholder="Ej. Losa Piso 3 / Eje B-4", key="f_inc_area_p5")
                    resp_nuevo = st.text_input("Responsable:*", placeholder="Ej. Cuadrilla Estructura / Ing. Residente", key="f_inc_resp_p5")

                with c_i2:
                    prio_nueva = st.segmented_control("Prioridad:*", ["Alta", "Media", "Baja"], default="Media", key="f_inc_prio_p5")
                    local_f_comp = get_local_datetime_ecuador().date() + datetime.timedelta(days=3)
                    f_comp_nueva = st.date_input("Fecha Compromiso:*", local_f_comp, key="f_inc_fcomp_p5")
                    est_nuevo = st.segmented_control("Estado Inicial:*", ["Abierta", "Cerrada"], default="Abierta", key="f_inc_est_p5")

                desc_nueva = st.text_area("Descripción de la Incidencia / No Conformidad:*", placeholder="Describa a detalle el problema o trabajo por corregir...", key="f_inc_desc_p5")

                btn_guardar_inc = st.form_submit_button("💾 Guardar Incidencia", type="primary", use_container_width=True)

                if btn_guardar_inc:
                    if not area_nueva.strip() or not desc_nueva.strip() or not resp_nuevo.strip():
                        st.error("⚠️ Por favor complete todos los campos obligatorios (*).")
                    else:
                        try:
                            nueva_data = {
                                "usuario_email": user_email,
                                "proyecto": proy_nuevo,
                                "area": area_nueva.strip(),
                                "descripcion": desc_nueva.strip(),
                                "responsable": resp_nuevo.strip(),
                                "prioridad": prio_nueva,
                                "fecha_compromiso": f_comp_nueva.strftime("%Y-%m-%d"),
                                "estado": est_nuevo
                            }
                            supabase.table("incidencias").insert(nueva_data).execute()
                            st.session_state.db_loaded = False
                            st.success("¡Incidencia registrada exitosamente!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al registrar la incidencia: {e}")

        todas_las_incidencias_db = st.session_state.get("db_incidencias_all", [])
        if len(user_edificios) > 0:
            lista_incs_vista = [inc for inc in todas_las_incidencias_db if inc.get("Proyecto") in user_edificios]
        else:
            lista_incs_vista = [inc for inc in todas_las_incidencias_db if inc.get("Usuario") == user_email]

        if proy_inc_sel != "-- Todos Mis Proyectos --":
            lista_incs_vista = [i for i in lista_incs_vista if i.get("Proyecto") == proy_inc_sel]

        if filtro_estado_vista != "Todos":
            lista_incs_vista = [i for i in lista_incs_vista if i.get("Estado") == filtro_estado_vista]

        st.markdown(f"#### Matriz de Incidencias en tus Proyectos ({len(lista_incs_vista)} registros)")

        if len(lista_incs_vista) > 0:
            table_rows_html = ""
            for idx, inc in enumerate(lista_incs_vista, 1):
                prio_val = inc.get('Prioridad', 'Media')
                est_val = inc.get('Estado', 'Abierta')
                
                p_alta_st = "font-weight:800; color:#dc2626;" if prio_val == 'Alta' else "font-weight:400; color:#64748b;"
                p_alta_ck = "☒ Alta" if prio_val == 'Alta' else "☐ Alta"
                p_med_st = "font-weight:800; color:#d97706;" if prio_val == 'Media' else "font-weight:400; color:#64748b;"
                p_med_ck = "☒ Media" if prio_val == 'Media' else "☐ Media"
                p_baj_st = "font-weight:800; color:#16a34a;" if prio_val == 'Baja' else "font-weight:400; color:#64748b;"
                p_baj_ck = "☒ Baja" if prio_val == 'Baja' else "☐ Baja"
                
                prio_cell = f'<div style="font-size:0.75rem; line-height:1.3;"><span style="{p_alta_st}">{p_alta_ck}</span><br/><span style="{p_med_st}">{p_med_ck}</span><br/><span style="{p_baj_st}">{p_baj_ck}</span></div>'
                
                e_ab_st = "font-weight:800; color:#dc2626;" if est_val == 'Abierta' else "font-weight:400; color:#64748b;"
                e_ab_ck = "☒ Abierta" if est_val == 'Abierta' else "☐ Abierta"
                e_ce_st = "font-weight:800; color:#16a34a;" if est_val == 'Cerrada' else "font-weight:400; color:#64748b;"
                e_ce_ck = "☒ Cerrada" if est_val == 'Cerrada' else "☐ Cerrada"
                
                est_cell = f'<div style="font-size:0.75rem; line-height:1.3;"><span style="{e_ab_st}">{e_ab_ck}</span><br/><span style="{e_ce_st}">{e_ce_ck}</span></div>'
                
                creador_tag = f"<br/><small style='color: #2563eb;'>👤 {inc.get('Usuario', '')}</small>" if inc.get('Usuario') != user_email else ""

                row_html = (
                    f'<tr>'
                    f'<td class="center" style="font-weight:700; width:45px;">{idx}</td>'
                    f'<td style="width:140px;"><b>{inc.get("Area", "")}</b><br/><small style="color:#64748b;">{inc.get("Proyecto", "")}</small>{creador_tag}</td>'
                    f'<td>{inc.get("Descripcion", "")}</td>'
                    f'<td style="width:130px;">{inc.get("Responsable", "")}</td>'
                    f'<td style="width:90px;">{prio_cell}</td>'
                    f'<td class="center" style="font-weight:700; width:100px;">{inc.get("Fecha_Compromiso", "")}</td>'
                    f'<td style="width:95px;">{est_cell}</td>'
                    f'</tr>'
                )
                table_rows_html += row_html

            full_table_html = (
                '<div style="overflow-x:auto;">'
                '<table class="incidencias-table">'
                '<thead>'
                '<tr>'
                '<th class="center" style="width:45px;">N°</th>'
                '<th style="width:140px;">Área</th>'
                '<th>Descripción</th>'
                '<th style="width:130px;">Responsable</th>'
                '<th style="width:90px;">Prioridad</th>'
                '<th class="center" style="width:100px;">Fecha Comp.</th>'
                '<th style="width:95px;">Estado</th>'
                '</tr>'
                '</thead>'
                f'<tbody>{table_rows_html}</tbody>'
                '</table>'
                '</div>'
            )
            st.markdown(full_table_html, unsafe_allow_html=True)

            with st.expander("⚙️ Administrar Estado o Eliminar Incidencias"):
                col_g1, col_g2, col_g3 = st.columns([2, 1.5, 1])
                inc_map = {f"N° {i} - {inc.get('Area')} ({inc.get('Proyecto')})": inc for i, inc in enumerate(lista_incs_vista, 1)}
                
                with col_g1:
                    sel_inc_label = st.selectbox("Seleccione el registro:", list(inc_map.keys()), key="sel_inc_gest_p5")
                    inc_target = inc_map[sel_inc_label]
                
                with col_g2:
                    toggle_st = "Cerrada" if inc_target.get("Estado") == "Abierta" else "Abierta"
                    if st.button(f"Cambiar estado a: {toggle_st}", type="primary", use_container_width=True):
                        try:
                            supabase.table("incidencias").update({"estado": toggle_st}).eq("id", inc_target["db_id"]).execute()
                            st.session_state.db_loaded = False
                            st.success(f"Estado actualizado a {toggle_st}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col_g3:
                    if st.button("🗑️ Eliminar", type="secondary", use_container_width=True):
                        try:
                            supabase.table("incidencias").delete().eq("id", inc_target["db_id"]).execute()
                            st.session_state.db_loaded = False
                            st.success("Incidencia eliminada correctamente")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al eliminar: {e}")

            st.markdown("<br>", unsafe_allow_html=True)
            c_exp1, c_exp2 = st.columns(2)
            nombre_proy_rep = proy_inc_sel if proy_inc_sel != "-- Todos Mis Proyectos --" else "Alpha_Builders_Incidencias"
            local_today_str = get_local_datetime_ecuador().strftime('%Y%m%d')
            
            with c_exp1:
                with st.popover("📊 Exportar Incidencias en Excel", use_container_width=True):
                    st.download_button(
                        label="Confirmar Descarga (.xlsx)",
                        data=export_incidencias_to_excel(lista_incs_vista, nombre_proy_rep),
                        file_name=f"Levantamiento_Incidencias_{nombre_proy_rep}_{local_today_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_excel_incidencias_tab_p5",
                        use_container_width=True
                    )
            with c_exp2:
                with st.popover("📄 Exportar Incidencias en PDF", use_container_width=True):
                    st.download_button(
                        label="Confirmar Descarga (.pdf)",
                        data=export_incidencias_to_pdf(lista_incs_vista, nombre_proy_rep),
                        file_name=f"Levantamiento_Incidencias_{nombre_proy_rep}_{local_today_str}.pdf",
                        mime="application/pdf",
                        key=f"dl_pdf_incidencias_tab_p5",
                        use_container_width=True
                    )
        else:
            st.info("No hay incidencias registradas para los proyectos seleccionados.")

# ==============================================================================
# 13. MÓDULO 5: CONTROL DE RENDIMIENTO
# ==============================================================================
with tab_rend:
    st.markdown("### Control de Rendimiento por Personal")
    st.caption("Asignación de rubros, ingreso manual de horario/HH, cantidades ejecutadas y diagnóstico de productividad.")

    mi_personal_propio = st.session_state.get("db_trabajadores_por_usuario", {}).get(user_email, [])
    # Ordenar alfabéticamente
    mi_personal_propio = sorted(mi_personal_propio, key=lambda it_w: str(it_w.get("nombre", "")).upper())
    nombres_personal = [f"{t['nombre']} ({t.get('edificio', 'General')})" for t in mi_personal_propio]

    col1, col2 = st.columns(2)
    with col1:
        opciones_personal = ["-- Seleccione un Integrante --"] + nombres_personal
        personal_sel = st.selectbox(
            f"Seleccionar de tu Personal ({len(nombres_personal)} Activos):*",
            opciones_personal,
            index=0,
            key="sel_trabajador_rend_p5"
        )
        
        if personal_sel != "-- Seleccione un Integrante --":
            nom_p_clean = personal_sel.split(" (")[0]
            cargo_actual = next((t["cargo"] for t in mi_personal_propio if t["nombre"] == nom_p_clean), "PERSONAL")
            st.info(f"**Cargo en obra:** {cargo_actual}")
        else:
            cargo_actual = "PERSONAL"

    with col2:
        opciones_rubros = ["-- Seleccione un Rubro --", "Enlucidos", "Fijos", "Fajas", "Dinteles"]
        rubro_sel = st.selectbox("Seleccionar Rubro:*", opciones_rubros, index=0, key="sel_rubro_rend_p5")
        
        if rubro_sel != "-- Seleccione un Rubro --":
            unidad_medida = UNIDADES_RUBRO[rubro_sel]
            st.caption(f"Unidad de medida: **{unidad_medida}** | Rend. Teórico: **{RENDIMIENTOS_TEORICOS.get(rubro_sel, 1.0)} HH/{unidad_medida}**")
        else:
            unidad_medida = "unid"

    st.markdown("---")
    st.markdown("#### ⏱️ Horario e Intervalo Trabajado")

    c_int1, c_int2 = st.columns(2)
    with c_int1:
        intervalo_manual = st.text_input(
            "Intervalo de Horas Trabajadas:*",
            placeholder="Ej. 07:00 - 12:00 / 13:00 - 16:00",
            key="in_intervalo_manual_rend_p5"
        )
    with c_int2:
        hh_manual = st.number_input(
            "Total Horas-Hombre (HH):*",
            min_value=0.0,
            max_value=24.0,
            step=0.5,
            value=0.0,
            format="%.2f",
            key="in_hh_manual_rend_p5"
        )

    st.markdown("#### 📊 Cantidades de Obra")
    c_cant1, c_cant2 = st.columns(2)
    with c_cant1:
        avance_cant = st.number_input(
            f"Cantidad Ejecutada Real ({unidad_medida}):*",
            min_value=0.0,
            step=0.1,
            format="%.2f",
            key="in_ejec_rend_p5"
        )
    with c_cant2:
        esperado_cant = st.number_input(
            f"Cantidad Esperada / Meta ({unidad_medida}):*",
            min_value=0.0,
            step=0.1,
            format="%.2f",
            key="in_esp_rend_p5"
        )

    if st.button("💾 Registrar Rendimiento", type="primary", use_container_width=True, key="btn_reg_rend_p5"):
        if personal_sel == "-- Seleccione un Integrante --":
            st.error("⚠️ Por favor seleccione un integrante de tu nómina de personal.")
        elif rubro_sel == "-- Seleccione un Rubro --":
            st.error("⚠️ Por favor seleccione un rubro.")
        elif not intervalo_manual.strip():
            st.error("⚠️ Por favor ingrese el intervalo de horas trabajadas.")
        elif hh_manual <= 0:
            st.error("⚠️ Por favor ingrese un valor de Horas-Hombre (HH) mayor a 0.")
        elif avance_cant <= 0:
            st.warning("⚠️ Ingrese una cantidad ejecutada mayor a 0.")
        else:
            rend_real = round(hh_manual / avance_cant, 3)
            rend_teorico = RENDIMIENTOS_TEORICOS.get(rubro_sel, 1.0)
            
            if esperado_cant > 0:
                estado_diag = "CUMPLE META" if avance_cant >= esperado_cant else "BAJO RENDIMIENTO"
            else:
                estado_diag = "EFICIENTE" if rend_real <= rend_teorico else "EXCESO DE HH"

            local_fecha_r = get_local_datetime_ecuador().strftime("%Y-%m-%d")

            try:
                supabase.table("rendimientos").insert({
                    "usuario_email": user_email,
                    "cargo_obrero": cargo_actual,
                    "fecha": local_fecha_r,
                    "trabajador": personal_sel.split(" (")[0],
                    "rubro": rubro_sel,
                    "intervalo": intervalo_manual.strip(),
                    "horas_hh": hh_manual,
                    "avance": avance_cant,
                    "esperado": esperado_cant,
                    "unidad": unidad_medida,
                    "rend_real": rend_real,
                    "rend_teorico": rend_teorico,
                    "estado": estado_diag
                }).execute()

                st.session_state.db_loaded = False
                st.success(f"¡Rendimiento registrado exitosamente para {personal_sel.split(' (')[0]}!")
                st.rerun()
            except Exception as e:
                st.error(f"Error registrando rendimiento: {e}")

    st.markdown("---")
    st.markdown("### Tabla de Resultados de Rendimiento Propios")

    mis_rendimientos = st.session_state.get("db_rendimientos", {}).get(user_email, [])

    if len(mis_rendimientos) > 0:
        for idx_r, r_item in enumerate(mis_rendimientos, 1):
            r_db_id = r_item.get("db_id")
            badge_r = render_estado_badge(r_item.get('Estado'))

            st.markdown('<div class="card-item-body-compact">', unsafe_allow_html=True)
            c_r1, c_r2, c_r3, c_r4, c_r5, c_r6, c_r7 = st.columns([0.5, 1.2, 2.5, 1.5, 1.5, 1.5, 0.5])

            with c_r1:
                st.markdown(f"**{idx_r}.**")
            with c_r2:
                st.caption(r_item.get('Fecha'))
            with c_r3:
                st.markdown(f"**{r_item.get('Trabajador')}** ({r_item.get('Cargo_Obrero', 'PERSONAL')})")
            with c_r4:
                st.write(f"{r_item.get('Rubro')} - {r_item.get('Horas Trabajadas (HH)')} HH")
            with c_r5:
                st.write(f"Ejec: {r_item.get('Avance')} {r_item.get('Unidad')}")
            with c_r6:
                st.markdown(badge_r, unsafe_allow_html=True)
            with c_r7:
                if st.button("🗑️", key=f"del_rnd_btn_{idx_r}_{r_db_id}_p5", help="Eliminar registro"):
                    try:
                        supabase.table("rendimientos").delete().eq("id", r_db_id).execute()
                        st.session_state.db_loaded = False
                        st.success("Registro de rendimiento eliminado.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al eliminar: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        df_mis_r = pd.DataFrame(mis_rendimientos)
        df_display = df_mis_r.drop(columns=["db_id", "Usuario_Registro", "Cargo_Registrador"], errors="ignore")
        if not df_display.empty:
            df_display.index = range(1, len(df_display) + 1)

        csv_bytes_r = export_dataframe_to_excel_csv(df_display)
        st.download_button(
            label="📥 Descargar Rendimientos en CSV (Excel)", 
            data=csv_bytes_r, 
            file_name=f"Rendimientos_{user_email}.csv", 
            mime="text/csv", 
            key="dl_csv_rend_tab_p5", 
            use_container_width=True
        )
    else:
        st.info("Aún no existen registros de rendimiento en tu cuenta.")

# ==============================================================================
# 14. MÓDULO 6: ESPACIO COLABORATIVO (AGRUPACIÓN MENSUAL DESPLEGABLE)
# ==============================================================================
with tab_colab:
    st.markdown("### Espacio de Trabajo Colaborativo")
    st.caption("Visualización cruzada de datos, incidencias, personal y actividades del Maestro Mayor entre compañeros que comparten proyectos.")

    mis_edificios_raw = st.session_state.get("usuario_edificios", [])
    mis_proyectos_norm = set([str(x).strip().lower() for x in mis_edificios_raw if str(x).strip()])

    if len(mis_proyectos_norm) == 0:
        st.warning("⚠️ No tienes proyectos asignados a tu cuenta. Agrega tus edificios en '⚙️ Configuración de Cuenta' en la barra lateral para unirte a grupos colaborativos.")
    else:
        companeros_colab = []
        for u in st.session_state.get("db_usuarios", []):
            u_correo_norm = str(u["Correo"]).lower().strip()
            if u_correo_norm != user_email:
                u_edifs_raw = u.get("Edificios", [])
                u_edifs_norm = set([str(x).strip().lower() for x in u_edifs_raw if str(x).strip()])
                interseccion_norm = mis_proyectos_norm.intersection(u_edifs_norm)
                
                if len(interseccion_norm) > 0:
                    nombres_comunes = [e for e in mis_edificios_raw if str(e).strip().lower() in interseccion_norm]
                    companeros_colab.append({
                        "usuario": u,
                        "proyectos_comunes": nombres_comunes if nombres_comunes else list(interseccion_norm)
                    })

        if len(companeros_colab) > 0:
            st.markdown("#### 👥 Integrantes de tu Grupo de Trabajo")
            
            c_sel_col1, c_sel_col2 = st.columns([2.5, 1.5])
            with c_sel_col1:
                map_colegas = {
                    f"{item['usuario']['Nombres']} {item['usuario']['Apellidos']} ({item['usuario']['Cargo']}) — {item['usuario']['Correo']}": item
                    for item in companeros_colab
                }
                sel_colega_str = st.selectbox(
                    "Selecciona un compañero para ver sus registros:",
                    list(map_colegas.keys()),
                    key="sel_colega_colab_p5"
                )
                item_colega_sel = map_colegas[sel_colega_str]
                colega_u = item_colega_sel["usuario"]
                c_mail = str(colega_u["Correo"]).lower().strip()
                c_cargo = colega_u.get("Cargo", "Residente")

            with c_sel_col2:
                st.markdown(
                    f"""
                    <div style="background: var(--card-bg); border: 1px solid var(--card-border); border-radius: 8px; padding: 6px 10px; margin-top: 24px;">
                        <small style="color: var(--subtext); font-weight: 700;">Proyectos en Común:</small><br/>
                        <b>{', '.join(item_colega_sel['proyectos_comunes'])}</b>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            st.markdown("---")

            if c_cargo == "Maestro Mayor":
                sub_tabs_colab = st.tabs([
                    "🔨 Libro de Obra del Maestro",
                    "⚡ Rendimientos del Compañero"
                ])

                with sub_tabs_colab[0]:
                    insps_maestro = st.session_state.get("db_inspecciones", {}).get(c_mail, [])
                    if len(insps_maestro) > 0:
                        df_col_mm = pd.DataFrame(insps_maestro)
                        df_col_mm["fecha_dt"] = pd.to_datetime(df_col_mm["Fecha"])
                        df_col_mm = df_col_mm.sort_values(by="fecha_dt", ascending=False)
                        df_col_mm["año"] = df_col_mm["fecha_dt"].dt.year
                        df_col_mm["mes_num"] = df_col_mm["fecha_dt"].dt.month
                        grupos_col_mm = df_col_mm.groupby(["año", "mes_num"], sort=False)

                        st.caption(f"Mostrando **{len(insps_maestro)}** reporte(s) de Maestro Mayor registrados por **{colega_u['Nombres']}** agrupados por mes:")
                        for (anio, mes_num), items_mes in grupos_col_mm:
                            nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Reportes)"
                            with st.expander(nombre_mes_str, expanded=False):
                                for idx_c_m, m_row in items_mes.iterrows():
                                    m_rep = m_row.to_dict()
                                    with st.expander(f"📌 [{m_rep.get('Proyecto', '')}] Fecha: {m_rep.get('Fecha', '')} ({m_rep.get('Dia', '')})", expanded=False):
                                        raw_dm = m_rep.get("Datos", {})
                                        d_m = raw_dm if isinstance(raw_dm, dict) else json.loads(raw_dm or "{}") if isinstance(raw_dm, str) else {}
                                        acts_m = d_m.get("Actividades_Maestro", []) if isinstance(d_m, dict) else d_m if isinstance(d_m, list) else []
                                        
                                        if acts_m:
                                            st.markdown("**Actividades, Personal y Metrajes Reportados:**")
                                            for a_it in acts_m:
                                                if isinstance(a_it, dict):
                                                    act_nombre = a_it.get('Actividad', a_it.get('actividad', ''))
                                                    act_cant = a_it.get('Cantidad', a_it.get('cantidad', ''))
                                                    act_obs = a_it.get('Observaciones', a_it.get('observaciones', 'Sin observaciones'))
                                                    pers_raw = a_it.get('Personal_A_Cargo', a_it.get('personal_a_cargo', []))
                                                    pers_str = ", ".join(pers_raw) if isinstance(pers_raw, list) else str(pers_raw or "")
                                                    pers_tag = f" | 👷 **Personal:** {pers_str}" if pers_str else ""
                                                    st.write(f"• **{act_nombre}**: `{act_cant}`{pers_tag} — *{act_obs}*")
                                        
                                        col_dl_mm1, col_dl_mm2 = st.columns(2)
                                        with col_dl_mm1:
                                            with st.popover("📊 Exportar Excel", use_container_width=True):
                                                st.download_button("Confirmar (.xlsx)", get_cached_libro_maestro_excel(safe_json_dumps(m_rep)), file_name=f"Libro_Maestro_{c_mail}_{m_rep['Fecha']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_col_mm_x_{idx_c_m}_p5", use_container_width=True)
                                        with col_dl_mm2:
                                            with st.popover("📄 Exportar PDF", use_container_width=True):
                                                st.download_button("Confirmar (.pdf)", get_cached_libro_maestro_pdf(safe_json_dumps(m_rep)), file_name=f"Libro_Maestro_{c_mail}_{m_rep['Fecha']}.pdf", mime="application/pdf", key=f"dl_col_mm_p_{idx_c_m}_p5", use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} aún no ha cargado actividades en su Libro de Obra.")

                with sub_tabs_colab[1]:
                    rnds_colega = st.session_state.get("db_rendimientos", {}).get(c_mail, [])
                    if len(rnds_colega) > 0:
                        df_r_col = pd.DataFrame(rnds_colega)
                        df_r_col["fecha_dt"] = pd.to_datetime(df_r_col["Fecha"])
                        df_r_col = df_r_col.sort_values(by="fecha_dt", ascending=False)
                        df_r_col["año"] = df_r_col["fecha_dt"].dt.year
                        df_r_col["mes_num"] = df_r_col["fecha_dt"].dt.month
                        grupos_r_col = df_r_col.groupby(["año", "mes_num"], sort=False)

                        st.caption(f"Mostrando **{len(rnds_colega)}** registros de rendimiento de **{colega_u['Nombres']}** organizados por mes:")
                        for (anio, mes_num), items_mes in grupos_r_col:
                            nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Rendimientos)"
                            with st.expander(nombre_mes_str, expanded=False):
                                df_disp_m = items_mes.drop(columns=["db_id", "Usuario_Registro", "Cargo_Registrador", "fecha_dt", "año", "mes_num"], errors="ignore")
                                if not df_disp_m.empty:
                                    df_disp_m.index = range(1, len(df_disp_m) + 1)
                                st.dataframe(df_disp_m, use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} aún no ha ingresado rendimientos.")

            else:
                sub_tabs_colab = st.tabs([
                    "📋 Checklists del Compañero",
                    "📖 Libro de Obra del Compañero",
                    "🚨 Incidencias de la Persona",
                    "⚡ Rendimientos del Compañero"
                ])

                with sub_tabs_colab[0]:
                    chks_colega = st.session_state.get("db_checklists", {}).get(c_mail, [])
                    if len(chks_colega) > 0:
                        df_chks_col = pd.DataFrame(chks_colega)
                        df_chks_col["fecha_dt"] = pd.to_datetime(df_chks_col["Fecha"])
                        df_chks_col = df_chks_col.sort_values(by="fecha_dt", ascending=False)
                        df_chks_col["año"] = df_chks_col["fecha_dt"].dt.year
                        df_chks_col["mes_num"] = df_chks_col["fecha_dt"].dt.month
                        grupos_chks_col = df_chks_col.groupby(["año", "mes_num"], sort=False)

                        st.caption(f"Mostrando **{len(chks_colega)}** checklist(s) registrados por **{colega_u['Nombres']}** organizados por mes:")
                        for (anio, mes_num), items_mes in grupos_chks_col:
                            nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Checklists)"
                            with st.expander(nombre_mes_str, expanded=False):
                                for idx_c_j, j_row in items_mes.iterrows():
                                    j_col = j_row.to_dict()
                                    with st.expander(f"📌 [{j_col.get('Edificio', '')}] {j_col.get('Fecha', '')} (Horario: {j_col.get('Hora_Inicio', '')} - {j_col.get('Hora_Fin', '')})", expanded=False):
                                        col_dl_c1, col_dl_c2 = st.columns(2)
                                        with col_dl_c1:
                                            with st.popover("📊 Exportar Excel", use_container_width=True):
                                                st.download_button("Confirmar (.xlsx)", get_cached_checklist_excel(safe_json_dumps(j_col)), file_name=f"Checklist_{c_mail}_{j_col['Fecha']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_colab_chk_x_{idx_c_j}_p5", use_container_width=True)
                                        with col_dl_c2:
                                            with st.popover("📄 Exportar PDF", use_container_width=True):
                                                st.download_button("Confirmar (.pdf)", get_cached_checklist_pdf(safe_json_dumps(j_col)), file_name=f"Checklist_{c_mail}_{j_col['Fecha']}.pdf", mime="application/pdf", key=f"dl_colab_chk_p_{idx_c_j}_p5", use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} aún no ha registrado checklists.")

                with sub_tabs_colab[1]:
                    insps_colega = st.session_state.get("db_inspecciones", {}).get(c_mail, [])
                    if len(insps_colega) > 0:
                        df_insps_col = pd.DataFrame(insps_colega)
                        df_insps_col["fecha_dt"] = pd.to_datetime(df_insps_col["Fecha"])
                        df_insps_col = df_insps_col.sort_values(by="fecha_dt", ascending=False)
                        df_insps_col["año"] = df_insps_col["fecha_dt"].dt.year
                        df_insps_col["mes_num"] = df_insps_col["fecha_dt"].dt.month
                        grupos_insps_col = df_insps_col.groupby(["año", "mes_num"], sort=False)

                        st.caption(f"Mostrando **{len(insps_colega)}** registro(s) en Libro de Obra organizados por mes:")
                        for (anio, mes_num), items_mes in grupos_insps_col:
                            nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Libros de Obra)"
                            with st.expander(nombre_mes_str, expanded=False):
                                for idx_c_i, i_row in items_mes.iterrows():
                                    i_col = i_row.to_dict()
                                    with st.expander(f"📌 [{i_col.get('Proyecto', '')}] {i_col.get('Fecha', '')} ({i_col.get('Dia', '')}) | Residente: {i_col.get('Residente', '')}", expanded=False):
                                        col_dl_i1, col_dl_i2 = st.columns(2)
                                        with col_dl_i1:
                                            with st.popover("📊 Exportar Excel", use_container_width=True):
                                                st.download_button("Confirmar (.xlsx)", get_cached_libro_oficial_excel(safe_json_dumps(i_col)), file_name=f"Libro_Obra_{c_mail}_{i_col['Fecha']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_colab_insp_x_{idx_c_i}_p5", use_container_width=True)
                                        with col_dl_i2:
                                            with st.popover("📄 Exportar PDF", use_container_width=True):
                                                st.download_button("Confirmar (.pdf)", get_cached_libro_oficial_pdf(safe_json_dumps(i_col)), file_name=f"Libro_Obra_{c_mail}_{i_col['Fecha']}.pdf", mime="application/pdf", key=f"dl_colab_insp_p_{idx_c_i}_p5", use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} aún no ha registrado formatos de Libro de Obra.")

                with sub_tabs_colab[2]:
                    todas_las_incidencias_db = st.session_state.get("db_incidencias_all", [])
                    incs_de_la_persona = [inc for inc in todas_las_incidencias_db if str(inc.get("Usuario", "")).lower().strip() == c_mail]
                    if len(incs_de_la_persona) > 0:
                        st.caption(f"Mostrando **{len(incs_de_la_persona)}** incidencia(s) levantadas por **{colega_u['Nombres']}**:")
                        df_incs_p = pd.DataFrame(incs_de_la_persona).drop(columns=["db_id"], errors="ignore")
                        st.dataframe(df_incs_p, use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} no tiene incidencias registradas.")

                with sub_tabs_colab[3]:
                    rnds_colega = st.session_state.get("db_rendimientos", {}).get(c_mail, [])
                    if len(rnds_colega) > 0:
                        df_r_col = pd.DataFrame(rnds_colega)
                        df_r_col["fecha_dt"] = pd.to_datetime(df_r_col["Fecha"])
                        df_r_col = df_r_col.sort_values(by="fecha_dt", ascending=False)
                        df_r_col["año"] = df_r_col["fecha_dt"].dt.year
                        df_r_col["mes_num"] = df_r_col["fecha_dt"].dt.month
                        grupos_r_col = df_r_col.groupby(["año", "mes_num"], sort=False)

                        st.caption(f"Mostrando **{len(rnds_colega)}** registros de rendimiento de **{colega_u['Nombres']}** organizados por mes:")
                        for (anio, mes_num), items_mes in grupos_r_col:
                            nombre_mes_str = f"📅 {NOMBRES_MESES.get(mes_num, 'Mes')} {anio} ({len(items_mes)} Rendimientos)"
                            with st.expander(nombre_mes_str, expanded=False):
                                df_disp_m = items_mes.drop(columns=["db_id", "Usuario_Registro", "Cargo_Registrador", "fecha_dt", "año", "mes_num"], errors="ignore")
                                if not df_disp_m.empty:
                                    df_disp_m.index = range(1, len(df_disp_m) + 1)
                                st.dataframe(df_disp_m, use_container_width=True)
                    else:
                        st.info(f"{colega_u['Nombres']} aún no ha ingresado rendimientos.")
        else:
            st.info("No se encontraron compañeros de trabajo que compartan tus mismos proyectos. Cuando otros usuarios seleccionen los mismos edificios, aparecerán automáticamente en tu grupo colaborativo.")

# ==============================================================================
# 15. MÓDULO 7: PANEL DE CONTROL ADMINISTRADOR
# ==============================================================================
if es_admin:
    tab_admin = tabs_app[-1]
    with tab_admin:
        st.markdown("### Panel de Control Administrador")
        st.caption("Módulo exclusivo para supervisar checklists, libro de obra, incidencias, rendimientos, usuarios y configuración.")

        st.markdown("#### 🔐 Código de Seguridad de Acceso y Registro (PIN)")
        col_pin1, col_pin2 = st.columns([2, 1])

        with col_pin1:
            pin_actual = st.session_state.get("access_pin", "1254")
            with st.form("form_pin_clean_p5"):
                nuevo_pin_input = st.text_input("Nuevo Código PIN (4 dígitos):", value=pin_actual, max_chars=4, type="password", help="Código requerido para iniciar sesión y registrar cuentas nuevas.")
                btn_pin_save = st.form_submit_button("Guardar Nuevo Código PIN", type="primary")

            if btn_pin_save:
                if len(nuevo_pin_input.strip()) == 4 and nuevo_pin_input.strip().isdigit():
                    try:
                        supabase.table("app_config").update({"value": nuevo_pin_input.strip()}).eq("key", "access_pin").execute()
                        st.session_state.access_pin = nuevo_pin_input.strip()
                        st.success(f"¡Código PIN de acceso actualizado exitosamente a: **{nuevo_pin_input.strip()}**!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al actualizar PIN: {e}")
                else:
                    st.error("El PIN debe constar exactamente de 4 números.")

        with col_pin2:
            st.info(f"**PIN Actual Configurado:** `{st.session_state.get('access_pin', '1254')}`")

        st.markdown("---")

        st.markdown("#### 👁️ Checklists Subidos por Todos los Participantes")
        todas_las_jornadas_admin = []
        for u_mail, j_lista in st.session_state.get("db_checklists", {}).items():
            for j_item in j_lista:
                j_copy = j_item.copy()
                j_copy["Usuario_Correo"] = u_mail
                todas_las_jornadas_admin.append(j_copy)

        if len(todas_las_jornadas_admin) > 0:
            col_adm_f1, col_adm_f2 = st.columns(2)
            with col_adm_f1:
                filtro_edif_admin = st.selectbox("Filtrar por edificio / proyecto:", ["-- Todos los Edificios --"] + EDIFICIOS_ALPHA, key="admin_filter_edif_chk_p5")
            with col_adm_f2:
                usuarios_lista_chk = ["-- Todos los Usuarios --"] + sorted(list(set([j["Usuario_Correo"] for j in todas_las_jornadas_admin])))
                filtro_usr_chk = st.selectbox("Filtrar por participante:", usuarios_lista_chk, key="admin_filter_usr_chk_p5")

            jornadas_admin_filtradas = todas_las_jornadas_admin.copy()
            if filtro_edif_admin != "-- Todos los Edificios --":
                jornadas_admin_filtradas = [j for j in jornadas_admin_filtradas if j.get("Edificio") == filtro_edif_admin]
            if filtro_usr_chk != "-- Todos los Usuarios --":
                jornadas_admin_filtradas = [j for j in jornadas_admin_filtradas if j["Usuario_Correo"] == filtro_usr_chk]

            jornadas_admin_filtradas.sort(key=lambda x: x['Fecha'], reverse=True)

            for idx_adm, j_adm in enumerate(jornadas_admin_filtradas, 1):
                resp_str = j_adm.get("Responsable", "") or j_adm["Usuario_Correo"]
                with st.expander(f"📌 [{j_adm.get('Edificio', 'N/A')}] {j_adm['Fecha']} — {resp_str} ({j_adm['Usuario_Correo']})", expanded=False):
                    c_ad_dl1, c_ad_dl2 = st.columns(2)
                    with c_ad_dl1:
                        with st.popover("📊 Exportar Excel", use_container_width=True):
                            st.download_button(
                                label="Confirmar (.xlsx)",
                                data=get_cached_checklist_excel(safe_json_dumps(j_adm)),
                                file_name=f"Checklist_{j_adm['Usuario_Correo']}_{j_adm['Edificio']}_{j_adm['Fecha']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_xlsx_adm_{idx_adm}_p5",
                                use_container_width=True
                            )
                    with c_ad_dl2:
                        with st.popover("📄 Exportar PDF", use_container_width=True):
                            st.download_button(
                                label="Confirmar (.pdf)",
                                data=get_cached_checklist_pdf(safe_json_dumps(j_adm)),
                                file_name=f"Checklist_{j_adm['Usuario_Correo']}_{j_adm['Edificio']}_{j_adm['Fecha']}.pdf",
                                mime="application/pdf",
                                key=f"dl_pdf_adm_{idx_adm}_p5",
                                use_container_width=True
                            )
        else:
            st.info("Ningún participante ha registrado checklists aún.")

        st.markdown("---")

        st.markdown("#### 📑 Registros de Libro de Obra Subidos por Todos los Participantes")
        todas_las_inspecciones_admin = []
        for u_mail, i_lista in st.session_state.get("db_inspecciones", {}).items():
            for i_item in i_lista:
                i_copy = i_item.copy()
                i_copy["Usuario_Correo"] = u_mail
                todas_las_inspecciones_admin.append(i_copy)

        if len(todas_las_inspecciones_admin) > 0:
            col_adm_i1, col_adm_i2 = st.columns(2)
            with col_adm_i1:
                filtro_edif_insp_adm = st.selectbox("Filtrar por edificio / proyecto:", ["-- Todos los Edificios --"] + EDIFICIOS_ALPHA, key="admin_filter_edif_insp_p5")
            with col_adm_i2:
                usuarios_lista_insp = ["-- Todos los Usuarios --"] + sorted(list(set([i["Usuario_Correo"] for i in todas_las_inspecciones_admin])))
                filtro_usr_insp_adm = st.selectbox("Filtrar por usuario:", usuarios_lista_insp, key="admin_filter_usr_insp_p5")

            insps_admin_filtradas = todas_las_inspecciones_admin.copy()
            if filtro_edif_insp_adm != "-- Todos los Edificios --":
                insps_admin_filtradas = [i for i in insps_admin_filtradas if i.get("Proyecto") == filtro_edif_insp_adm]
            if filtro_usr_insp_adm != "-- Todos los Usuarios --":
                insps_admin_filtradas = [i for i in insps_admin_filtradas if i["Usuario_Correo"] == filtro_usr_insp_adm]

            insps_admin_filtradas.sort(key=lambda x: x['Fecha'], reverse=True)

            for idx_i_adm, i_adm in enumerate(insps_admin_filtradas, 1):
                raw_adm_d = i_adm.get("Datos", {})
                d_i_adm = raw_adm_d if isinstance(raw_adm_d, dict) else json.loads(raw_adm_d or "{}") if isinstance(raw_adm_d, str) else {}
                es_tipo_mm = (d_i_adm.get("Tipo_Registro") == "Libro_Obra_Maestro")
                titulo_exp = f"🔨 [MAESTRO MAYOR] [{i_adm.get('Proyecto', 'N/A')}] {i_adm['Fecha']} — Usuario: {i_adm['Usuario_Correo']}" if es_tipo_mm else f"📌 [{i_adm.get('Proyecto', 'N/A')}] {i_adm['Fecha']} ({i_adm['Dia']}) — Usuario: {i_adm['Usuario_Correo']}"

                with st.expander(titulo_exp, expanded=False):
                    if es_tipo_mm:
                        acts_mm = d_i_adm.get("Actividades_Maestro", []) if isinstance(d_i_adm, dict) else d_i_adm if isinstance(d_i_adm, list) else []
                        for a in acts_mm:
                            pers_str = ", ".join(a.get("Personal_A_Cargo", [])) if isinstance(a.get("Personal_A_Cargo"), list) else str(a.get("Personal_A_Cargo", ""))
                            pers_tag = f" | 👷 **Personal:** {pers_str}" if pers_str else ""
                            st.write(f"• **{a.get('Actividad', a.get('actividad', ''))}**: `{a.get('Cantidad', a.get('cantidad', ''))}`{pers_tag} — *{a.get('Observaciones', a.get('observaciones', ''))}*")
                        c_ad_idl1, c_ad_idl2 = st.columns(2)
                        with c_ad_idl1:
                            with st.popover("📊 Exportar Excel", use_container_width=True):
                                st.download_button("Confirmar (.xlsx)", get_cached_libro_maestro_excel(safe_json_dumps(i_adm)), file_name=f"Libro_Maestro_{i_adm['Proyecto']}_{i_adm['Fecha']}.xlsx", key=f"dl_mm_adm_x_{idx_i_adm}", use_container_width=True)
                        with c_ad_idl2:
                            with st.popover("📄 Exportar PDF", use_container_width=True):
                                st.download_button("Confirmar (.pdf)", get_cached_libro_maestro_pdf(safe_json_dumps(i_adm)), file_name=f"Libro_Maestro_{i_adm['Proyecto']}_{i_adm['Fecha']}.pdf", key=f"dl_mm_adm_p_{idx_i_adm}", use_container_width=True)
                    else:
                        st.markdown(f"**Ubicación:** {d_i_adm.get('Ubicacion', i_adm.get('Frente', ''))} | **Clima:** {i_adm.get('Clima', '')}")
                        st.markdown(f"**Superintendente:** {d_i_adm.get('Superintendente', '')} | **Fiscalizador:** {d_i_adm.get('Fiscalizador', '')}")
                        
                        acts_adm_lo = d_i_adm.get("Actividades_Ejecutadas", [])
                        if acts_adm_lo:
                            st.markdown("**Actividades:**")
                            for a in acts_adm_lo:
                                st.write(f"• **{a.get('Descripcion', '')}** | {a.get('Area', '')} | {a.get('Cantidad', 0)} {a.get('Unidad', 'm2')}")

                        c_ad_idl1, c_ad_idl2 = st.columns(2)
                        with c_ad_idl1:
                            with st.popover("📊 Exportar Excel", use_container_width=True):
                                st.download_button("Confirmar (.xlsx)", get_cached_libro_oficial_excel(safe_json_dumps(i_adm)), file_name=f"Libro_Obra_{i_adm['Proyecto'].replace(' ', '_')}_{i_adm['Fecha']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_insp_xlsx_adm_{idx_i_adm}_p5", use_container_width=True)
                        with c_ad_idl2:
                            with st.popover("📄 Exportar PDF", use_container_width=True):
                                st.download_button("Confirmar (.pdf)", get_cached_libro_oficial_pdf(safe_json_dumps(i_adm)), file_name=f"Libro_Obra_{i_adm['Proyecto'].replace(' ', '_')}_{i_adm['Fecha']}.pdf", mime="application/pdf", key=f"dl_insp_pdf_adm_{idx_i_adm}_p5", use_container_width=True)
        else:
            st.info("Ningún participante ha registrado formatos de Libro de Obra aún.")

        st.markdown("---")

        st.markdown("#### 🚨 Levantamiento de Incidencias - Vista Administrador")
        todas_las_incidencias_admin = st.session_state.get("db_incidencias_all", [])
        if len(todas_las_incidencias_admin) > 0:
            df_inc_admin = pd.DataFrame(todas_las_incidencias_admin)
            st.dataframe(df_inc_admin, use_container_width=True)
            csv_inc_admin_bytes = export_dataframe_to_excel_csv(df_inc_admin)
            st.download_button(
                label="📥 Descargar Todas las Incidencias (CSV)",
                data=csv_inc_admin_bytes,
                file_name=f"Incidencias_Globales_{get_local_datetime_ecuador().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                key="dl_csv_inc_admin_all_p5",
                use_container_width=True
            )
        else:
            st.info("No hay incidencias registradas en la obra actualmente.")

        st.markdown("---")

        st.markdown("#### 📊 Rendimientos Subidos por Todos los Participantes")
        todos_los_rendimientos = []
        for u_mail, r_lista in st.session_state.get("db_rendimientos", {}).items():
            for r_item in r_lista:
                r_copy = r_item.copy()
                r_copy["Usuario_Correo"] = u_mail
                todos_los_rendimientos.append(r_copy)

        if len(todos_los_rendimientos) > 0:
            df_rend_admin = pd.DataFrame(todos_los_rendimientos)
            cols_first = ["Usuario_Correo", "Fecha", "Trabajador", "Cargo_Obrero", "Rubro", "Intervalo", "Horas Trabajadas (HH)", "Avance", "Esperado", "Unidad", "Rend. Real (HH/Unid)", "Rend. Teórico", "Estado"]
            df_rend_admin = df_rend_admin.reindex(columns=[c for c in cols_first if c in df_rend_admin.columns])
            if not df_rend_admin.empty:
                df_rend_admin.index = range(1, len(df_rend_admin) + 1)

            st.dataframe(df_rend_admin, use_container_width=True)

            csv_rend_admin_bytes = export_dataframe_to_excel_csv(df_rend_admin)
            st.download_button(
                label="📥 Descargar Todos los Rendimientos (Excel CSV)",
                data=csv_rend_admin_bytes,
                file_name=f"Rendimientos_Globales_{get_local_datetime_ecuador().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                key="dl_csv_rend_admin_p5",
                use_container_width=True
            )
        else:
            st.info("Ningún participante ha registrado rendimientos aún.")

        st.markdown("---")

        st.markdown("#### Gestión de Administradores de la Plataforma")
        col_adm1, col_adm2 = st.columns([2, 1])

        with col_adm1:
            with st.form("form_admin_add_clean_p5"):
                nuevo_admin_mail = st.text_input("Ingrese correo para conceder permisos de Administrador:", placeholder="usuario@correo.com")
                btn_admin_add = st.form_submit_button("Otorgar Acceso Administrador", use_container_width=True)

            if btn_admin_add:
                if nuevo_admin_mail:
                    mail_clean = nuevo_admin_mail.strip().lower()
                    try:
                        supabase.table("usuarios").update({"es_admin": True}).ilike("correo", mail_clean).execute()
                        st.session_state.db_loaded = False
                        st.success(f"Se otorgaron permisos de administrador a: {mail_clean}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error actualizando administrador: {e}")

        with col_adm2:
            st.markdown("**Administradores Registrados:**")
            for admin_email_item in st.session_state.get("admin_emails", []):
                st.write(f"• `{admin_email_item}`")